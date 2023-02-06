
from pathlib import Path

import pandas as pd
import win32com.client as client
import xlsxwriter
from openpyxl.utils import get_column_letter
from utils.formats import taxYear
from utils.functions import PAYNO_Check, tax_calcs

year = taxYear().Year("-")
yearp = taxYear().Yearp("-")
week = tax_calcs().tax_week()

homePath = Path.home() / "advance.online"

user = str(Path.home()).split("\\")[-1]

marginsPath = homePath / rf"J Drive - Exec Reports/Margins Reports/Margins {year}"
#marginspPath = homePath / rf"J Drive - Exec Reports/Margins Reports/Margins {yearp}"

dataPath = marginsPath / rf"Data/Week {week - 1}"

marginsReport = pd.read_excel(marginsPath / rf"Margins Report {year}.xlsx", sheet_name=['Core Data', "Accounts 2", "Clients"])
#marginspReport = pd.read_excel(marginspPath / rf"Margins Report {yearp}.xlsx", sheet_name=['Core Data', "Accounts 2", "Clients"])

joiners  = pd.read_csv(Path.home() / "advance.online/J Drive - Operations/Reports/MCR/Joiners Error Report.csv",usecols=['Pay No',"Email Address", "Sdc Option"], encoding = 'latin', low_memory=False)

joiners = joiners[joiners["Pay No"].apply(lambda x: PAYNO_Check(x))].drop_duplicates(subset=["Pay No"])
joiners['Pay No'] = joiners['Pay No'].astype(int)

joinersp  = pd.read_csv(Path.home() / "advance.online/J Drive - Operations/Reports/MCR/Joiners Error Report p.csv",usecols=['Pay No',"Email Address", "Sdc Option"], encoding = 'latin', low_memory=False)

joinersp = joinersp[joinersp["Pay No"].apply(lambda x: PAYNO_Check(x))].drop_duplicates(subset=["Pay No"])
joinersp['Pay No'] = joinersp['Pay No'].astype(int)

core_accounts = ["Martin Brown", "Gerry Hunnisett", "Adam Shaw", "Sam Amos", "Dave Levenston"]

accounts = marginsReport["Accounts 2"][["Office Number", "Account Owner"]]
accounts = accounts.dropna(subset=['Office Number'])
accounts['Office Number'] = accounts['Office Number'].astype(int)
accounts = accounts.drop_duplicates(subset=['Office Number'], keep='first')

core = marginsReport["Core Data"]
#corep = marginspReport["Core Data"]

clients = marginsReport["Clients"][["Company Name                   ","OFFNO"]].rename(columns = {"Company Name                   ":"Client Name"})
clients.sort_values("Client Name", inplace = True)
clients["Client Name"] = clients["Client Name"].str.upper()
clients.drop_duplicates(subset ="Client Name",
                        keep = "last", inplace = True)

prevWeek = core[(core["Week Number"] == week - 1) & (core["Margins"] > 0)]
prevWeek.loc[~prevWeek["CRM"].isin(core_accounts), "CRM"] = "Unmanned Account"

#prevYear = corep[(corep["Week Number"] == week) & (corep["Margins"] > 0)]
#prevYear.loc[~prevYear["CRM"].isin(core_accounts), "CRM"] = "Unmanned Account"

prevWPivot = pd.pivot_table(data= prevWeek, values="Margins", index=["CRM","Client Name"], aggfunc={"Margins": len}).rename(columns={"Margins": "Prev Week"}).reset_index()

prevWeek = prevWeek[prevWeek['PAYNO'].apply(lambda x: PAYNO_Check(x))]
prevWeek['PAYNO'] = prevWeek['PAYNO'].astype(int)

#prevYPivot = pd.pivot_table(data= prevYear, values="Margins", index=["CRM","Client Name"], aggfunc={"Margins": len}).rename(columns={"Margins": "Prev Year"}).reset_index().drop(columns = "CRM")

#prevYear = prevYear[prevYear['PAYNO'].apply(lambda x: PAYNO_Check(x))]
#prevYear['PAYNO'] = prevYear['PAYNO'].astype(int)

date = tax_calcs().chqdate(week)

io = pd.read_csv("margins.csv",encoding = 'latin')
axm = pd.read_csv("margins axm.csv",encoding = 'latin')
paye = pd.read_csv("margins paye.csv",encoding = 'latin')

paye.loc[paye["FREQ"] == "W", "MANAGEMENT FEE"] = 1

margins = pd.concat([io, axm, paye]).rename(columns={"COMPNAME": "Client Name"})
margins = margins[margins["PAYNO"].apply(lambda x: PAYNO_Check(x))]
zero_margins = margins[margins["MANAGEMENT FEE"] <= 0].reset_index().fillna("")
margins = margins[margins["MANAGEMENT FEE"] > 0]
margins['PAYNO'] = margins['PAYNO'].astype(int)

margins = margins.merge(joiners, how = "left", left_on="PAYNO", right_on="Pay No").drop(columns = "Pay No").merge(clients, how="left").merge(accounts, left_on="OFFNO", right_on="Office Number", how="left").drop(columns=["OFFNO", "Office Number"])

margins.loc[~margins["Account Owner"].isin(core_accounts), "Account Owner"] = "Unmanned Account"

margins["Solution"] = "Umbrella"

margins.loc[ margins["Sdc Option"] == "Mileage Only", "Solution"] = "Umbrella with Mileage"
margins.loc[ margins["Sdc Option"] == "Fixed Expenses", "Solution"] = "Umbrella with Expenses"
margins.loc[ margins["TYPE"] == "CIS", "Solution"] = "CIS"

missingWorkers = prevWeek[~prevWeek["PAYNO"].isin(margins["PAYNO"].unique())][["Client Name", "PAYNO","Surname","Forename","Solution.1"]].drop_duplicates(subset="PAYNO").reset_index(drop=True).fillna('')

missingAgencies = prevWPivot[~prevWPivot["Client Name"].isin(margins["Client Name"].unique())][["Client Name"]].drop_duplicates(subset="Client Name").reset_index(drop=True).fillna('')

pivot = pd.pivot_table(data= margins, values="MANAGEMENT FEE", index=["Account Owner", "Client Name"], aggfunc={"MANAGEMENT FEE": len}).reset_index().rename(columns={"MANAGEMENT FEE": "Total"}).merge(prevWPivot, how="outer", left_on="Client Name", right_on="Client Name")#.merge(prevYPivot, how="outer", left_on="Client Name", right_on="Client Name")

pivot.loc[pivot["Account Owner"].isna(), "Account Owner"] = pivot.loc[pivot["Account Owner"].isna(), "CRM"]

pivot = pivot.drop(columns = "CRM").fillna(0)

pivot["Account Owner"] = pivot["Account Owner"].astype(str)

pivot = pivot.sort_values("Account Owner")

pivot["Difference"] = pivot["Total"] - pivot["Prev Week"]

pivot.loc[len(pivot)] = pivot.sum(numeric_only=True)

pivot.at[len(pivot) - 1, "Client Name"] = "Total"

marginsTotal = round(pivot.at[len(pivot) - 1, "Total"])

pivot = pivot.fillna("")

wb = xlsxwriter.Workbook("margins.xlsx")

format1 = wb.add_format({'bg_color': '#FFC7CE',
                        'font_color': '#9C0006'})
        
format2 = wb.add_format({'bg_color': '#C6EFCE',
                        'font_color': '#006100'})
        
cell_format_column = wb.add_format({'font_size' : 16,
                                    'align': 'center',
                                    'bg_color': '#FFFF00',
                                    'border':1})

ws = wb.add_worksheet('Summary')

for j, column in enumerate(pivot.columns.values):
    col = get_column_letter(j + 1)
    rowend = len(pivot)+1
    REF = f'{col}{1}'
    ws.write(REF,column,cell_format_column)
    ws.set_column(f'{col}:{col}', 15)
    if column in ["Difference", "Variance"]:
        ws.conditional_format(f'{col}{2}:{col}{rowend}', {'type': 'cell',
                                        'criteria': '<',
                                        'value': 0,
                                        'format': format1})
        
        ws.conditional_format(f'{col}{2}:{col}{rowend}', {'type': 'cell',
                                        'criteria': '>',
                                        'value': 0,
                                        'format': format2})
        
for i, row in pivot.iterrows():
    j = 0
    for item in row:
        REF = rf'{get_column_letter(j + 1)}{i + 2}'
        ws.write(REF,item)
        j += 1

ws = wb.add_worksheet('Missing Workers')

for j, column in enumerate(missingWorkers.columns.values):
    rowend = len(missingWorkers) + 1
    REF = f'{get_column_letter(j + 1)}{1}'
    ws.write(REF,column,cell_format_column)
    ws.set_column(f'{get_column_letter(j + 1)}:{get_column_letter(j + 1)}', 15)
        
for i, row in missingWorkers.iterrows():
    j = 0
    for item in row:
        REF = f'{get_column_letter(j + 1)}{i + 2}'
        ws.write(REF,item)
        j += 1
        
ws = wb.add_worksheet('Missing Agencies')

for j, column in enumerate(missingAgencies.columns.values):
    rowend = len(missingWorkers)+1
    REF = f'{get_column_letter(j + 1)}{1}'
    ws.write(REF,column,cell_format_column)
    ws.set_column(f'{get_column_letter(j + 1)}:{get_column_letter(j + 1)}', 15)
        
for i, row in missingAgencies.iterrows():
    j = 0
    for item in row:
        REF = f'{get_column_letter(j + 1)}{i + 2}'
        ws.write(REF,item)
        j += 1

ws = wb.add_worksheet('0 Margins')
for j, column in enumerate(margins.columns.values):
    rowend = len(margins)+1
    REF = f'{get_column_letter(j + 1)}{1}'
    ws.write(REF,column,cell_format_column)
    ws.set_column(f'{get_column_letter(j + 1)}:{get_column_letter(j + 1)}', 15)
        
for i, row in (zero_margins).iterrows():
    j = 0
    for item in row:
        REF = f'{get_column_letter(j + 1)}{i + 2}'
        ws.write(REF,item)
        j += 1
        
ws.hide()

ws = wb.add_worksheet('Workers Paid')

margins = margins[["Account Owner","Client Name","PAYNO","FIRSTNAME", "LASTNAME","Email Address","TOTHRS", "TOTPAY", "Basic", "Solution"]].fillna('')

for j, column in enumerate(margins.columns.values):
    rowend = len(margins)+1
    REF = f'{get_column_letter(j + 1)}{1}'
    ws.write(REF,column,cell_format_column)
    ws.set_column(f'{get_column_letter(j + 1)}:{get_column_letter(j + 1)}', 15)
        
for i, row in margins.iterrows():
    j = 0
    for item in row:
        REF = f'{get_column_letter(j + 1)}{i + 2}'
        ws.write(REF,item)
        j += 1

wb.close()

outlook = client.Dispatch('Outlook.Application')
email = outlook.CreateItem(0)
email.To = 'Joshua.Richards@advance.online; Adam.Shaw@advance.online; Gerry.Hunnisett@advance.online; Sam.Amos@advance.online; Dave.Levenston@advance.online; Martin.Brown@advance.online; Jodie.Beeston@advance.online; Alwyn.Barrow@advance.online; harriet.murray@advance.online; jake.price@advance.online; anna.sills@advance.online;'
email.CC = 'jacob.sterling@advance.online'
email.Subject = ('Margins per agency')

email.HTMLBody = rf"""
    <div>
        Hi All, 
        <br>
        <br>
        See the updated margins report
        <br>
        <br>
        Margins: {marginsTotal}
        <br>
        <br>
        <b>Kind regards</b>,
        <br>
        <br>
        <font color='#3F43AD'>
        <b>Jacob Sterling</b>
        <br>
        Reporting and IT Trainee 
        <br>
        <br>
        Office: 01244 564 564
        <br>
        Email: jacob.sterling@advance.online
        <br>
        Visit: <a href='https://www.advance.online/'>www.advance.online</a>
        <br>
        <br>
        <em>Service is important to us, and we value your feedback. Please tell us how we did today by clicking <a href='https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab\#lrd=0x487ac350478ae187:0xc129ea0eaf860aee,3,,,https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab'>here</a> to leave us a review on Google</em>
        </font>
        <br>
        <br>
        <img src="{str(Path.home() / 'OneDrive - advance.online/Documents/signature.png')}">
    <\div>
"""

email.Attachments.Add(Source=str(Path().absolute() / "margins.xlsx"))

email.Display()