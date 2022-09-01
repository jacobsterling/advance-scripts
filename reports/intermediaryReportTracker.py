# -*- coding: utf-8 -*-
"""
Created on Fri Jan 28 14:21:33 2022

@author: jacob.sterling
"""

class intermediaryReportScript():
    def __init__(self):
        import numpy as np
        import pandas as pd
        from pathlib import Path
        from functions import tax_calcs
        from formats import taxYear
        import win32com.client as client
        
        self.pd = pd
        self.np = np
        self.QuarterRange, self.Quarter, self.Quarterp = tax_calcs().Quarter_Calc()
        
        self.outlook = client.Dispatch('Outlook.Application')
        
        namespace = self.outlook.GetNamespace("MAPI")
        
        Items = namespace.Folders["jacob.sterling@advance.online"].Folders["Sent Items"].Items
        
        subject = f'Intermediate Report - {self.Quarter}'
        
        self.sentItems = []
        
        for i in range(Items.Count -1, 0, -1):
            if Items[i].Subject == subject:
                self.sentItems.append(Items[i].To)
            
        #To answer your question regarding the lack of information on the intermediary report, 
        #all category F contractors are paid on an employment basis. 
        #All of their payment details will be captured in our FPS that we submit each week to HMRC and are therefore not needed on the intermediary report. 
        #All other categories do require this additional information, 
        #however when looking at your report all contractors contained have been paid on an employment basis.
        
        
        #Weekc = tax_calcs().tax_week_calc()
        Weekc = 14

        YearFormat = taxYear().Year('-')
        YearcFormat1 = taxYear().Year_format1('-')
        
        YearpFormat1 = taxYear().Yearp_format1('-')#change to Year_format1
        
        self.report_prefix = 'IntermediaryReport_083_FA45839_'
        trackerName = f'{YearcFormat1}-{self.Quarter} Inter Report Tracker Python.xlsx'
        
        homePath = Path.home() / "advance.online/J Drive - Operations/Reports/Intermediary Reports"
        dataPath = Path.home() / rf"advance.online/J Drive - Exec Reports/Margins Reports/Margins {YearFormat}/Data/Week {Weekc}"
        
        filePathP = homePath / rf"{YearpFormat1}/{self.Quarterp}"

        self.filePath = homePath / rf"{YearcFormat1}/{self.Quarter}"

        self.reportsPath = self.filePath / "Outstanding Reports"
        self.trackerPath = self.filePath / trackerName
        
        for file in filePathP.glob('*.xlsx'):
            if file.name.__contains__('racker'):
                self.trackerP = pd.read_excel(file, usecols=['OFFNO','Merit Email', 'CRM Email'])
                break#change column names to OFFNO and Merit Email
        
        self.c = 1 if input('Type "y" make Changes to files ?: ') == "y" else 0
        
        df_CLIENTS_IO = dataPath / "clients io.csv"
        df_CLIENTS_AXM = dataPath / "clients axm.csv"
        df_ACCOUNTS = dataPath / "Accounts+Office.csv"
        groups_path = Path.home() / "OneDrive - advance.online/Documents/Data/Groups.xlsx"
        
        df_CLIENTS = pd.concat([pd.read_csv(df_CLIENTS_IO,encoding = 'latin',
                                            usecols = ['Company Name                   ','OFFNO','EMAIL_DETAILS_INTER']),
                                pd.read_csv(df_CLIENTS_AXM,encoding = 'latin',
                                            usecols = ['Company Name                   ','OFFNO','EMAIL_DETAILS_INTER'])])
        
        df_CLIENTS.columns = ['Client Name','OFFNO','Email']
        df_CLIENTS['Client Name'] = df_CLIENTS['Client Name'].str.upper()
        df_CLIENTS.sort_values("Client Name", inplace = True)
        df_CLIENTS.drop_duplicates(subset ="Client Name",
                             keep = "last", inplace = True)
        self.df_CLIENTS = df_CLIENTS
        
        df_ACCOUNTS = pd.read_csv(df_ACCOUNTS, na_values = '-', skiprows=6)
        df_ACCOUNTS = df_ACCOUNTS[['Office Number','Account Owner','Send Intermediary Report to','Account Type']]
        df_ACCOUNTS = df_ACCOUNTS.dropna(subset=['Office Number'])
        df_ACCOUNTS['Office Number'] = df_ACCOUNTS['Office Number'].astype(int)
        self.df_ACCOUNTS = df_ACCOUNTS.drop_duplicates(subset=['Office Number'], keep='first')
        
        df_groups = pd.read_excel(groups_path)
        df_groups['Client Name'] = df_groups['Client Name'].str.upper()
        df_groups['Name Change'] = df_groups['Name Change'].str.upper()
        self.df_groups = df_groups
        
        self.missingData = pd.DataFrame([],columns = ['Name','Missing Info'])
    
    def age(self, birthdate):
        from datetime import date
        return int((date.today() - self.pd.to_datetime(birthdate).date()).days / 365)
    
    def nullQuery(self, query, row, k, df_report, notif = None):
        if self.pd.isnull(row[query]) or row[query] == 'N/A':
            self.missingInfo += 1
            self.details += f'missing {query}, ' if not notif else notif
            if self.c == 1:
                try:
                    if query == 'Worker address line 2':
                        print(rf"Address Line 1: {row['Worker address line 1']}: ")
                        
                    user = input(f"Enter missing {query} for {row['Worker forename']} {row['Worker surname']}: ")
                    if user == '':
                        self.missingData.append([[f"{row['Worker forename']} {row['Worker surname']}", query]])
                    else:
                        df_report.at[k, query] = user
                        self.changes += 1
                except KeyError:
                    user = input(f"Enter {notif} for {df_report.columns[1]}: ")
                    if user != '':
                        df_report.at[k, query] = user
                        self.changes += 1
                    else:
                        self.missingInfo += 1
                    self.missingInfo -= 1
        return df_report
                
    def rangeQuery(self, startCol, endCol, df_report_row, k, values, df_report, notif):
        if (df_report.loc[k,startCol:endCol].values != values).all():
            self.missingInfo += 1
            self.details += f'missing {notif}, '
            if self.c == 1:
                df_report.loc[k,startCol:endCol] = values
                self.changes += 1
                print('')
                print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
                print(f"Changed {notif} for {df_report_row['Worker forename']} {df_report_row['Worker surname']}.")
                print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
        return df_report
    
    def matchingQuery(self, condition1, condition2, notif = None):
        if condition1 != condition2:
            self.details += f'{condition1} does not match {condition2}, '
            return False
        return True
    
    def readReport(self,file):
        try:
            df_info = self.pd.read_csv(file,nrows=5, on_bad_lines='skip').iloc[:,:2]
            df_info[df_info.columns[1]] = df_info[df_info.columns[1]].astype(str)
            df_info.loc[df_info[df_info.columns[1]] == 'nan',df_info.columns[1]] = self.np.nan
            return df_info, self.pd.read_csv(file,skiprows=7, on_bad_lines='skip')
        except UnicodeDecodeError:
            df_info = self.pd.read_csv(file,nrows=5,encoding='latin', on_bad_lines='skip').iloc[:,:2]
            df_info[df_info.columns[1]] = df_info[df_info.columns[1]].astype(str)
            df_info.loc[df_info[df_info.columns[1]] == 'nan',df_info.columns[1]] = self.np.nan
            return df_info, self.pd.read_csv(file,skiprows=7,encoding='latin', on_bad_lines='skip')

    def matchReport(self, offno):
        try:
            clientName = self.df_CLIENTS.loc[self.df_CLIENTS['OFFNO']==offno,'Client Name'].values[0]
            clientEmail = self.df_CLIENTS.loc[self.df_CLIENTS['OFFNO']==offno,'Email'].values[0]
        except IndexError:
            clientName = None
            clientEmail = None
            self.missingInfo += 2
            print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
            print(f'{offno} is missing from clients.csv')
            self.details += f'{offno} is missing from clients.csv, '
            print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
            
        try:
            accountOwner = self.df_ACCOUNTS.loc[self.df_ACCOUNTS['Office Number']==offno,'Account Owner'].values[0]
            accountType = self.df_ACCOUNTS.loc[self.df_ACCOUNTS['Office Number']==offno,'Account Type'].values[0]
            crmEmail = self.df_ACCOUNTS.loc[self.df_ACCOUNTS['Office Number']==offno,'Send Intermediary Report to'].values[0]
        except IndexError:
            accountOwner = None
            accountType = None
            crmEmail = None
    
        self.nullVar('account owner',accountOwner, 1)
        self.nullVar('account type',accountType, 1)
        self.nullVar('client email',clientEmail)
        if self.nullVar('crm email',crmEmail, 1):
            try:
                crmEmail = self.trackerP.loc[self.trackerP['OFFNO']==offno,'CRM Email'].values[0]
                if crmEmail == 0:
                    crmEmail = None 
                    self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
                else:
                    self.details += f'Used email from previous {self.Quarterp} tracker, '
            except IndexError:
                try:
                    crmEmail = self.trackerP.loc[self.trackerP['OFFNO']==offno,'Merit Email'].values[0]
                    if crmEmail == 0:
                        crmEmail = self.np.nan 
                        self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
                    else:
                        self.details += f'Used merit email from previous {self.Quarterp} tracker, '
                except IndexError:
                    self.missingInfo += 1
                    self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
            
        elif clientEmail != crmEmail and not self.pd.isnull(clientEmail):
            self.missingInfo += 1
            self.details += 'Merit Email does not match CRM, '
        
        for i, row in self.df_groups.iterrows():
            if not self.pd.isnull(clientName):
                if row['Client Name'] == clientName or (clientName.__contains__(row['Name Change']) and len(row['Name Change']) > 3):
                    clientName = row['Name Change']
                    accountType = 'Group'
                    self.missingInfo += 1
                    self.details += 'Needs to be Grouped, '
                    break
    
        return clientName, accountOwner, accountType, clientEmail, crmEmail

    def queryReport(self, df_info, df_report):
        df_iter_info = df_info
        for j, info_row in df_iter_info.iterrows():
            if j == 0:
                df_info = self.nullQuery(df_info.columns[1], info_row, j, df_info,'Missing Address, ')
        df_info = self.nullQuery(df_info.columns[1], info_row, j, df_info,'Missing Post Code, ')

        df_iter_report = df_report

        for k, df_report_row in df_iter_report.iterrows():
            
            df_report = self.nullQuery('Worker date of birth', df_report_row, k, df_report)
            
            df_report = self.nullQuery('Worker gender', df_report_row, k, df_report)
            
            df_report = self.nullQuery('Worker National Insurance number', df_report_row, k, df_report)

            df_report = self.nullQuery('Start date of engagement',df_report_row, k, df_report)
            
            df_report = self.nullQuery('Worker address line 2',df_report_row, k, df_report)
            
            df_report = self.nullQuery('Worker date of birth', df_report_row, k, df_report)
            
            df_report = self.nullQuery('Worker postcode', df_report_row, k, df_report)
            
            
            if self.age(df_report_row['Worker date of birth']) < 16 or self.age(df_report_row['Worker date of birth']) > 75 :
                self.missingInfo += 1
                self.details += 'Potentially Incorrect Date Of Birth, '
            
            if self.pd.to_datetime(df_report_row['End date of engagement']) == self.pd.to_datetime("19/04/2022" ,format="%d/%m/%Y"):
                self.details += 'Potentially Incorrect End Date, '
                if self.c == 1:
                    endDate = input("Enter new End date of engagament (in format dd/mm/YYYY): ")
                    if endDate != "":
                        df_report.at[k, 'End date of engagement'] = self.pd.to_datetime(endDate, format="%d/%m/%Y")
                        self.changes += 1
                    else:
                        self.missingInfo += 1
                else:
                    self.missingInfo += 1
                    
            if df_report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'D' or df_report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'B':
                
                df_report = self.rangeQuery("Name of party paid by intermediary for worker's services",
                                "Postcode of party paid by intermediary for worker's services",
                                df_report_row, k,
                                ["Advance Contracting Solutions Ltd","First Floor VISTA","St David's Park","Ewloe","Chester","CH5 3DT"],
                                df_report,
                                'Intermediary Details, ')

                df_report = self.nullQuery("Amount paid for the worker's services" , df_report_row, k, df_report)

                df_report = self.nullQuery("Companies House registration number of party paid by intermediary for worker's services", df_report_row, k, df_report)
                    
            if df_report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'A':
                
                df_report = self.rangeQuery("Name of party paid by intermediary for worker's services",
                                "Postcode of party paid by intermediary for worker's services",
                                df_report_row, k, 
                                ["Advance Contracting Solutions Ltd","First Floor VISTA","St David's Park","Ewloe","Chester","CH5 3DT"],
                                df_report,
                                'Intermediary Details, ')
                        
                df_report = self.nullQuery("Amount paid for the worker's services", df_report_row, k, df_report)

                df_report = self.nullQuery("Worker unique taxpayer reference (UTR)", df_report_row, k, df_report)
                    #add groups together
                    #email potentially ?
        
        numberOfErrors = len(df_report[df_report["Amount paid for the worker's services"].astype(float) == 0])
        if numberOfErrors > 0:
            self.missingInfo += numberOfErrors
            self.details += "Amount paid for the worker's services <= 0"
            if self.c == 1:
                if input(rf'Amount paid for the workers services <= 0, Remove {numberOfErrors} rows ? ("y"): ') == 'y':
                    df_report = df_report[df_report["Amount paid for the worker's services"].astype(float) != 0]
                    self.changes += 1
        
        numberOfErrors = len(df_report[ df_report["Worker forename"].apply(lambda x: x.upper().__contains__(" DNU"))])
        if numberOfErrors > 0:
            self.missingInfo += numberOfErrors
            self.details += rf"Contains {numberOfErrors} DNU Worker(s), "
            if self.c == 1:
                if input(rf'Report contains DNU Worker(s), Remove {numberOfErrors} rows ? ("y"): ') == 'y':
                    df_report = df_report[ ~df_report["Worker forename"].apply(lambda x: x.upper().__contains__(" DNU")) ]
                    self.changes += 1
        return df_info, df_report
    
    def createTracker(self):
        
        self.df = self.pd.DataFrame([],columns=['File','OFFNO','Client','Account','Account Type','Merit Email','CRM Email','Changes Made','Missing Info','A','F','B','D', 'Details'])
        for file in self.reportsPath.glob('*'):
            if file.is_file():
                if file.suffix == ".csv" and self.report_prefix in file.name: 
                    offno = int(file.name.split('_')[3])
                    
                    print('___________________________________________________________________')
                    print('')
                    print(f'Reading {file.name}')
                    
                    
                    df_info, df_report = self.readReport(file)
                    
                    self.changes, self.missingInfo, self.details  = 0, 0, ''
                    
                    clientName, accountOwner, accountType, clientEmail, crmEmail = self.matchReport(offno)

                    if self.pd.isnull(clientName):
                        clientName = df_info.columns[1].upper()
                        self.details += 'Client name used from file, '
                        
                    if not self.matchingQuery(df_info.columns[1].upper(), clientName):
                        if self.c == 1 and input(f"Miss Matching Client Name {df_info.columns[1].upper()}, type 'y' to change to {clientName}: ") == 'y':
                            self.changes += 1
                            df_info.columns = ['Employment intermediary name',clientName.title()]
                    
                    print('')
                    print(f'Showing Client - {df_info.columns[1].upper()}')
                    print('')
                    
                    df_info, df_report = self.queryReport(df_info, df_report)
                                
                    self.saveChanges(file, df_info, df_report)
                    
                    
                    
                    if accountType != 'Agency':
                        self.missingInfo += 1
                        self.details += 'Not agency, '

                    df_temp = self.pd.DataFrame([[file.name,
                                             offno,
                                             clientName,
                                             accountOwner,
                                             accountType,
                                             clientEmail,
                                             crmEmail,
                                             self.changes,
                                             self.missingInfo,
                                             len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'A']),
                                             len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'F']),
                                             len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'B']),
                                             len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'D']),
                                             self.details]],
                                             columns=['File','OFFNO','Client','Account','Account Type','Merit Email','CRM Email','Changes Made','Missing Info','A','F','B','D','Details'])
                    
                    self.df = self.df.append(df_temp)
                    
        self.exportTracker()
        if input('Type "y" to edit the tracker ?: ') == 'y':
            self.editTracker()
        else:
            return self.df, self.missingData
    
    def nullVar(self, name, var, i = 0):
        if self.pd.isnull(var):
            self.missingInfo += i
            self.details += f'Missing {name}, '
            return True
        else:
            return False
        
    def updateReport(self, i, clientName, accountOwner, accountType, clientEmail, crmEmail, df_report):
        
        try:
            A = len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'A'])
            F = len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'F'])
            B = len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'B'])
            D = len(df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'D'])  
        except TypeError:
            A = 0
            F = 0
            B = 0
            D = 0
            
        self.df.at[i, ['Client','Account','Account Type','Merit Email','CRM Email','Changes Made','Missing Info','A','F','B','D','Details']] = [clientName,
        accountOwner,
        accountType,
        clientEmail,
        crmEmail,
        self.changes,
        self.missingInfo,
        A,
        F,
        B,
        D,
        self.details]
        
    def editTracker(self):
        try:
            self.send = True if input('Type "y" to setup emails ?: ') == "y" else False
            self.df = self.pd.read_excel(self.trackerPath)
            for i, row in self.df.iterrows():
                clientName = self.df.at[i, 'Client']
                accountOwner = self.df.at[i,'Account']
                accountType = self.df.at[i,'Account Type']
                clientEmail = self.df.at[i,'Merit Email']
                crmEmail = self.df.at[i,'CRM Email']
                self.details = '' # if self.pd.isnull(row['Details']) else row['Details']
                self.missingInfo = self.df.at[i,'Missing Info']
                self.changes = 'Sent' if crmEmail in self.sentItems else row['Changes Made']
                
                if self.changes not in ['Sent']: #,'Send'
                    file = self.reportsPath / f"{row['File']}"
                    
                    print('___________________________________________________________________')
                    print('')
                    print(f'Reading {file.name}')
                
                    df_info, df_report = self.readReport(file)
                    self.missingInfo = 0
                    
                    manualSend = True if self.changes == 'Send' else False
                    self.changes = 0 if self.changes == 'Send' else self.changes
                
                    if accountOwner == 'OFFNO not found in accounts office':
                        self.missingInfo += 1
                        self.details += 'Missing Account Owner, '
                    
                    self.nullVar('account type', accountType, 1)                    
                        
                    if accountType in ['Group', 'End Client'] and self.c == 1:
                        group = self.df.loc[self.df['Client'] == clientName] if accountType == 'Group' else self.df.loc[self.df['Account Type'] == 'End Client']
                        if len(group) > 1:
                            if input(f"Type 'y' to group {clientName} with groupee's: ") == 'y':
                                changes = self.changes
                                for j, groupee in group.iterrows():
                                    gPath = self.filePath / f"Outstanding Reports/{groupee['File']}"
                                    if gPath != file:
                                        g_info, g_report = self.readReport(gPath)
                                        self.changes = groupee['Changes Made']

                                        self.changes = 0 if self.changes == 'Send' else self.changes
                                        
                                        print('')
                                        print(f'Showing {accountType} Groupee - {g_info.columns[1].upper()}')
                                        print('')
                                        
                                        g_info, g_report = self.queryReport(g_info, g_report)
                                        
                                        self.df.at[j, 'Account Type'] = 'Groupee' if accountType == 'Group' else 'End Client Groupee'
                                        
                                        self.saveChanges(file, g_info, g_report)
    
                                        df_report = self.pd.concat([df_report,g_report])
                                        
                                self.changes = changes + 1
                                if accountType == 'Group':
                                    accountType = 'Grouped'
                                else:
                                    accountType = 'End Client Grouped'
                                    clientName = rf'Advance Contracting Solutions - Formally {clientName}'
                                    df_info.columns = ["Employment intermediary name","Advance Contracting Solutions"]
                                    df_info.at[0, "Advance Contracting Solutions"] = "Ground Floor"
                                    df_info.at[1, "Advance Contracting Solutions"] = "VISTA"
                                    df_info.at[2, "Advance Contracting Solutions"] = "St David’s Park"
                                    df_info.at[3, "Advance Contracting Solutions"] = "Ewloe"
                                    df_info.at[4, "Advance Contracting Solutions"] = "CH5 3DT"
                                    df_report = df_report[df_report["Worker engagement details where intermediary didn't operate PAYE"] == 'A']
                                    
                                self.details += 'Grouped, '
                            elif input("Type 'y' if already grouped: ") == 'y':
                                accountType = 'Grouped'
                                self.changes += 1
                    
                    if accountType not in ["Agency", 'Grouped']:
                        if accountType == "Other" and self.send:
                            if input("Type 'y' to treat as agency ?: ") != "y":
                                self.missingInfo += 1
                                self.details += 'Not Agency, '
                            else:
                                self.missingInfo += 1
                                self.details += 'Not Agency, '
                    
                    if accountType != 'Groupee':
                        if not self.matchingQuery(df_info.columns[1].upper(), clientName):
                            if self.c == 1 and input(f"Miss Matching Client Name {df_info.columns[1].upper()}, type 'y' to change to {clientName}: ") == 'y':
                                self.changes += 1
                                df_info.columns = ['Employment intermediary name',clientName.title()]
                        
                        print('')
                        print(f'Showing {accountType} - {df_info.columns[1].upper()}')
                        print('')
                            
                        df_info, df_report = self.queryReport(df_info, df_report)
                                    
                        self.saveChanges(file, df_info, df_report)
                        
                        if self.send and accountType in ["Agency", 'Grouped', "Other"] and self.changes != 'Send' and self.changes > 0:
                            print('')
                            print(f'Details - {self.details}')
                            print('')
                            self.changes = 'Send' if input(f'Type "y" to allow file with {self.changes} changes to be sent to agency (does not send): ') == 'y' else self.changes
                        
                        if self.nullVar('crm email',crmEmail, 1) and not self.nullVar('client email',clientEmail) and input(f'Type "y" to update {clientName} crm email {crmEmail} with {clientEmail}: ') == 'y':
                            crmEmail = clientEmail
                            self.missingInfo -= 1
                            self.details = self.details.replace('Missing CRM Email, ','')
                            
                        self.updateReport(i, clientName, accountOwner, accountType, clientEmail, crmEmail, df_report)
                    
                        if self.send and (manualSend or self.changes == 'Send') and accountType in ["Agency", 'Grouped', "Other"] and not self.pd.isnull(crmEmail):
                            print('')
                            print(f"A: {self.df.at[i,'A']}, F: {self.df.at[i,'F']}, B: {self.df.at[i,'B']}, D: {self.df.at[i,'D']}")
                            print('')
                            if self.missingInfo > 0:
                                s = True if input('Type "y" to if you want to compose email despite missing information: ') == 'y' else False
                            elif input(f'Type "y" to compose email to {clientName}: ') == 'y':
                                s = True
                            else:
                                s = False
                            if s:
                                self.emailReport(crmEmail,file)
                                self.df.at[i,'Changes Made'] = 'Sent'
                        elif self.changes == 'Sent':
                            print('')
                            print('Sent')
                else:
                    self.updateReport(i, clientName, accountOwner, accountType, clientEmail, crmEmail, None)
        except KeyboardInterrupt: 
            pass
        self.exportTracker()
        return self.df, self.missingData
    
    def saveChanges(self, file, info, report):
        if self.c == 1 and self.changes > 0:
            if input('Type "y" to remove duplicates on NI number: ') == 'y':
                report.drop_duplicates(subset=['Worker National Insurance number',"Worker engagement details where intermediary didn't operate PAYE"],keep = "first", inplace = True)
                
            if input(f'Type "y" to save {self.changes} changes to {info.columns[1].upper()} - {file.name} with {self.missingInfo} missing info: ') == 'y':
                
                reportCol = self.pd.DataFrame([report.columns])
                
                report.columns = range(0, len(report.columns))
                
                emptyRow = self.pd.DataFrame([[self.np.nan for i in report.columns]])
                
                infoCol = self.pd.DataFrame([info.columns])
                
                for i in range(len(info.columns), len(report.columns)):
                    info[i] = self.np.nan
                
                info.columns = report.columns
                
                df = self.pd.concat([infoCol, info, emptyRow, reportCol, report]).reset_index(drop=True)
                
                df.to_csv(file, index=False, header=False)
                
                info = info.iloc[:,:2]
                
                info.columns = infoCol.iloc[0]
                
                report.columns = reportCol.iloc[0]
            
        else:
            self.changes = 0
            if self.missingInfo == 0:
                self.changes = 'Send'

    def emailReport(self, crmEmail, report):
        email = self.outlook.CreateItem(0)
        email.To = crmEmail
        email.Subject = self.subject

        email.SentOnBehalfOfName = 'info@advance.online'
        #info
        html = rf"""
            Hi,
            <br>
            <br>
            Please find attached your intermediary report for {self.Quarter}.
            <br>
            <br>
            Please remember to report on any limited companies you may have contracted with in {self.Quarter}. 
            <br>
            <br>
            I have also attached a PDF document to that you may find useful to review on intermediary reporting.
            <br>
            <br>
            <u><i>Assumptions</i></u>
            <br>
            <br>
            Start date is the date we set the contractor up
            <br>
            <br>
            End date is the day the contractor informed us they are no longer working with ADVANCE
            <br>
            <br>
            If you have any questions please let me know. 
            <br>
            <br>
            <b>Kind regards</b>,
            <br>
            <br>
            <font color='#7098F0'>
            <b>ADVANCE</b>
            <br>
            <br>
            Office: 01244 564 564
            <br>
            Email: info@advance.online
            <br>
            Visit: <a href='https://www.advance.online/'>www.advance.online</a>
            <br>
            <br>
            <em>Service is important to us, and we value your feedback. Please tell us how we did today by clicking <a href='https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab#lrd=0x487ac350478ae187:0xc129ea0eaf860aee,3,,,https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab'>here</a> to leave us a review on Google</em>
            </font>
            <br>
            <br>
            <img src="{str(self.filePath / 'signature.png')}">
            <br>
            <br>
            <font style='color:#9C9C9C' face = 'Ariel' size='0.5'>
            The information in this email is confidential, privileged and protected by copyright. It is intended solely for the addressee. If you are not the intended recipient any disclosure, copying or distribution of this email is prohibited and may not be lawful. If you have received this transmission in error, please notify the sender by replying by email and deleting the email from all of your computer systems. The transmission of this email cannot be guaranteed to be secure or unaffected by a virus. The contents may be corrupted, delayed, intercepted or lost in transmission, and Advance cannot accept any liability for errors, omissions or consequences which may arise.
            <br>
            Registered Office; Ground Floor, VISTA, St David's Park, Ewloe, CH5 3DT. Registered in England and Wales.
            </font>
        """
        email.HTMLBody = html
        email.Attachments.Add(Source=str(report))
        email.Attachments.Add(Source=str(self.filePath / "Onshore employment intermediaries - FAQs.pdf"))
        email.Display(True)
        
    def exportTracker(self):
        from openpyxl.utils import get_column_letter
        import xlsxwriter
        
        df = self.df.fillna('').reset_index(drop=True)
        
        wb = xlsxwriter.Workbook(self.trackerPath)
        
        format1 = wb.add_format({'bg_color': '#FFC7CE',
                                       'font_color': '#9C0006'})
        
        format2 = wb.add_format({'bg_color': '#C6EFCE',
                                       'font_color': '#006100'})
        
        cell_format_column = wb.add_format({'font_size' : 16,
                                            'align': 'center',
                                            'bg_color': '#FFFF00',
                                            'border':1})
        
        ws = wb.add_worksheet('Tracker')
        
        for j, column in enumerate(df.columns.values):
            col = get_column_letter(j + 1)
            row = 1
            rowend = len(df)+1
            ws.write(f'{col}{row}',column,cell_format_column)
            ws.set_column(f'{col}:{col}', 15)
            if column == 'Missing Info':
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': '>',
                                              'value': 0,
                                              'format': format1})
                
        for i, row in df.iterrows():
            j = 0
            for item in row:
                REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
                if row.index[j] == 'File':
                    #ws.write_url(REF_1, str(self.filePath / "Outstanding Reports") + '\\' + item,string=item)
                    ws.write_url(REF_1, "Outstanding Reports/" + item, string=item)
                elif row.index[j] == 'Changes Made':
                    if item == 'Sent':
                        ws.write(REF_1,item,format2)
                    else:
                        ws.write(REF_1,item)
                        
                elif row.index[j] == 'Account' or row.index[j] == 'Account Type':
                    if item == '':
                        ws.write(REF_1,item,format1)
                    else:
                        ws.write(REF_1,item)
                else:
                    ws.write(REF_1,item)
                j += 1
        wb.close()
        
#df, missingData = intermediaryReportScript().createTracker()
df, missingData = intermediaryReportScript().editTracker()
