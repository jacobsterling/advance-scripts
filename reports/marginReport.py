# -*- coding: utf-8 -*-
"""
Created on Thu Oct 21 13:35:49 2021

@author: jacob.sterling
"""

#importing external modules
import pandas as pd
from datetime import date
import datetime
import numpy as np
import win32com.client as win32
import xlsxwriter
win32c = win32.constants
from openpyxl.utils import get_column_letter

datetime.datetime.now().time()
today = date.today()

def tax_week_calc(week):
    if week < 15:
        week = week + 38
    else:
        week = week - 14
    return week

Week = tax_week_calc(today.isocalendar()[1])

if Week > 1:
    Weekp = Week - 1
else:
    Weekp = 52
    
Year = today.year
yearppp = Year - 3
yearpp = Year - 2
yearp = Year - 1
yearc = Year
yearcc = Year + 1
    
if Week < 39:
    Yearpp = (f'{yearpp}-{yearp}')
    Yearp = (f'{yearp}-{yearc}')
    Year = (f'{yearc}-{yearcc}')
    Year2 = ('{yearc}{yearcc}').format(yearc = yearc, yearcc = yearcc-2000)
else:
    Yearpp = (f'{yearppp}-{yearpp}')
    Yearp = (f'{yearpp}-{yearp}')
    Year = (f'{yearp}-{yearc}')
    Year2 = ('{yearp}{yearc}').format(yearp = yearp, yearc = yearc-2000)

CORE_ACCOUNTS = ['Adam Shaw','Dave Levenston','Gerry Hunnisett','Sam Amos']

print('___________________________________________________________________')
print('')
print('Loading Paths...')
print(datetime.datetime.now().time())

file_path: str = rf"C:\Users\jacob.sterling\OneDrive - advance.online\Exec Reports\Margins Reports\Margins {Year}\Data"
df_path: str = f"{file_path}\Margin Report {Year} Week {Weekp}.xlsx"
df_path_new: str = f"{file_path}\Margin Report {Year} Week {Week}.xlsx"

last_paid_path_io: str = f"{file_path}\Last Paid Import IO - {today}.csv"
last_paid_path_axm: str = f"{file_path}\Last Paid Import AXM - {today}.csv"
latest_start_date_path: str = f"{file_path}\Latest Start Date Import - {today}.csv"

file_path: str = f"{file_path}\Week {Week}"

df_FEES_RETAINED_IO: str = f"{file_path}\\fees retained io.csv"
df_FEES_RETAINED_AXM: str = f"{file_path}\\fees retained axm.csv"
df_CLIENTS_IO: str = f"{file_path}\\clients io.csv"
df_CLIENTS_AXM: str = f"{file_path}\\clients axm.csv"
df_ACCOUNTS: str = f"{file_path}\\accounts office.csv"
df_PAYE_DATA: str = f"{file_path}\\emp paid.csv"

df_CORRIE_SUMMARY: str = f"J:\Operations\Remittances and invoices\Corrie\Tax Year {Year}\Week {Week}\PAYE\Corrie Summary {Year} Week {Week}.xlsx"

df_MRRR: str = f"{file_path}\\Margins Report Registrations Report (MRRR).csv"
df_0_MARGINS_IO: str = f"{file_path}\\£0 Margins io.csv"
df_0_MARGINS_AXM: str = f"{file_path}\\£0 Margins axm.csv"
df_JOINERS_ERROR_IO: str = f"{file_path}\\joiners error report io.csv"
df_JOINERS_ERROR_AXM: str = f"{file_path}\\joiners error report axm.csv"
latest_start_date: str = f"{file_path}\\latest start date.csv"

print('___________________________________________________________________')
print('')
print('Producing latest start date csv for crm upload...')
print(datetime.datetime.now().time())

latest_start_date = pd.read_csv(latest_start_date ,encoding = 'latin')
latest_start_date.columns = latest_start_date.iloc[0, :]
latest_start_date = latest_start_date.drop(0, axis = 0)
latest_start_date = latest_start_date[:-3]
latest_start_date = latest_start_date.drop(['POTENTIALID'], axis = 1)
latest_start_date = latest_start_date.drop(['Created Time'], axis = 1)
latest_start_date = latest_start_date.drop(['Opportunity Name'], axis = 1)
latest_start_date['Start Date on Site'] = pd.to_datetime(latest_start_date['Start Date on Site'], format='%d %b, %Y').dt.strftime('%d/%m/%Y').astype(str)
latest_start_date.to_csv(latest_start_date_path, encoding='utf-8', index=False)

print('___________________________________________________________________')
print('')
print('Produced latest start date csv ready for crm upload.')
print(datetime.datetime.now().time())

print('___________________________________________________________________')
print('')
print('Loading Core Data...')
print(datetime.datetime.now().time())

df = pd.read_excel(df_path, sheet_name= 'Core Margin Data')
    
df = df[df['Margins'] > 0]

df['Client Name'] = df['Client Name'].str.upper()

for i, row in df.iterrows():
    if isinstance(row['CHQDATE'],str) == False:
        df.at[i,'CHQDATE'] = row['CHQDATE'].strftime('%d/%m/%Y')

df['CHQDATE'] = pd.to_datetime(df['CHQDATE'], format= '%d/%m/%Y')

df = df.sort_values(by='CHQDATE')

df_MARGINS = pd.read_excel(df_path, sheet_name='Core Totals Data', 
                            usecols = ['Tax Week',f'Margins {Yearpp}',f'Margins {Yearp}',f'Margins {Year}'])

df_REGISTRATIONS = pd.read_excel(df_path, sheet_name= 'Core Totals Data', 
                                  usecols = ['Tax Week',
                                            f'New {Yearpp}',f'Conversion {Yearpp}',f'Dormant {Yearpp}',f'Total {Yearpp}',
                                            f'New {Yearp}',f'Conversion {Yearp}',f'Dormant {Yearp}',f'Total {Yearp}',
                                            f'New {Year}',f'Conversion {Year}',f'Dormant {Year}',f'Total {Year}'])

df_REGISTRATIONS[f'New {Year}'] = 0
df_REGISTRATIONS[f'Conversion {Year}'] = 0
df_REGISTRATIONS[f'Dormant {Year}'] = 0
df_REGISTRATIONS[f'Total {Year}'] = 0

print('___________________________________________________________________')
print('')
print('Loading Registration Data...')
print(datetime.datetime.now().time())

df_CONTRACTOR_RELATIONS = pd.read_excel(df_path, sheet_name= 'Core Totals Data', 
                                  usecols = ['Tax Week',f'CR New {Year}',f'CR Conversion {Year}',f'CR Dormant {Year}',f'CR Total {Year}'])

df_CONTRACTOR_RELATIONS[f'CR New {Year}'] = 0
df_CONTRACTOR_RELATIONS[f'CR Conversion {Year}'] = 0
df_CONTRACTOR_RELATIONS[f'CR Dormant {Year}'] = 0
df_CONTRACTOR_RELATIONS[f'CR Total {Year}'] = 0

df_INTERNAL_SALES = pd.read_excel(df_path, sheet_name= 'Core Totals Data', 
                                  usecols = ['Tax Week',f'IS New {Year}',f'IS Conversion {Year}',f'IS Dormant {Year}',f'IS Total {Year}'])

df_INTERNAL_SALES[f'IS New {Year}'] = 0
df_INTERNAL_SALES[f'IS Conversion {Year}'] = 0
df_INTERNAL_SALES[f'IS Dormant {Year}'] = 0
df_INTERNAL_SALES[f'IS Total {Year}'] = 0

print('___________________________________________________________________')
print('')
print('Loading Fees Retained...')
print(datetime.datetime.now().time())

df_FEES_RETAINED_IO = pd.read_csv(df_FEES_RETAINED_IO,encoding = 'latin')

df_FEES_RETAINED_IO = df_FEES_RETAINED_IO.rename(columns={"Management Fee":"Margins"})
df_FEES_RETAINED_IO = df_FEES_RETAINED_IO.rename(columns={"OFF_PAYNO":"PAYNO"})
df_FEES_RETAINED_IO['PAYNO'] = df_FEES_RETAINED_IO['PAYNO'].astype(str)
df_FEES_RETAINED_IO['Client Name'] = df_FEES_RETAINED_IO['Client Name'].str.upper()
df_FEES_RETAINED_IO = df_FEES_RETAINED_IO[df_FEES_RETAINED_IO['Margins'] > 0]
df_FEES_RETAINED_IO['CHQDATE'] = pd.to_datetime(df_FEES_RETAINED_IO['CHQDATE'], format='%d/%m/%Y')

df_FEES_RETAINED_AXM = pd.read_csv(df_FEES_RETAINED_AXM,encoding = 'latin')

df_FEES_RETAINED_AXM = df_FEES_RETAINED_AXM.rename(columns={"Management Fee":"Margins"})
df_FEES_RETAINED_AXM = df_FEES_RETAINED_AXM.rename(columns={"OFF_PAYNO":"PAYNO"})
df_FEES_RETAINED_AXM['PAYNO'] = df_FEES_RETAINED_AXM['PAYNO'].astype(str)
df_FEES_RETAINED_AXM['Client Name'] = df_FEES_RETAINED_AXM['Client Name'].str.upper()
df_FEES_RETAINED_AXM = df_FEES_RETAINED_AXM[df_FEES_RETAINED_AXM['Margins'] > 0]
df_FEES_RETAINED_AXM['CHQDATE'] = pd.to_datetime(df_FEES_RETAINED_AXM['CHQDATE'], format='%d/%m/%Y')

print('___________________________________________________________________')
print('')
print('Loading Clients...')
print(datetime.datetime.now().time())

df_CLIENTS = pd.concat([pd.read_csv(df_CLIENTS_IO,encoding = 'latin',
                                    usecols = ['Company Name                   ','OFFNO']),
                        pd.read_csv(df_CLIENTS_AXM,encoding = 'latin',
                                    usecols = ['Company Name                   ','OFFNO'])])

df_CLIENTS.columns = ['Client Name','OFFNO']
df_CLIENTS['Client Name'] = df_CLIENTS['Client Name'].str.upper()
df_CLIENTS.sort_values("Client Name", inplace = True)
df_CLIENTS.drop_duplicates(subset ="Client Name",
                     keep = "last", inplace = True)

print('___________________________________________________________________')
print('')
print('Loading MRRR...')
print(datetime.datetime.now().time())

df_MRRR = pd.read_csv(df_MRRR ,encoding = 'latin')
df_MRRR.columns = df_MRRR.iloc[0, :]
df_MRRR = df_MRRR.drop(0, axis = 0)
df_MRRR = df_MRRR[:-3]
df_MRRR = df_MRRR[['NI Number (Contact)','Email (Contact)','Solutions','Agency','Consultant',
                   'Dormant/Conversion','Created Time','Job Title','Holiday Pay Entitlement',
                   "Has the worker been told they are 'Inside IR35'?"]]

df_MRRR['Created Time'] = pd.to_datetime(df_MRRR['Created Time'], format='%d %b, %Y %H:%M')
df_MRRR["Has the worker been told they are 'Inside IR35'?"] = df_MRRR["Has the worker been told they are 'Inside IR35'?"].fillna('No')
df_MRRR.loc[df_MRRR["Has the worker been told they are 'Inside IR35'?"] == 'No', 'Contractor Relations'] = 'Internal Sales'
df_MRRR.loc[df_MRRR["Has the worker been told they are 'Inside IR35'?"].str.contains("Yes"), 'Contractor Relations'] = 'Contractor Relations'

print('___________________________________________________________________')
print('')
print('Formating MRRR and calculating totals...')
print(datetime.datetime.now().time())

for i, items in df_MRRR.iterrows():
    week = tax_week_calc(items['Created Time'].isocalendar()[1] + 1)
    df_MRRR.at[i, 'Tax Week'] = week
    df_REGISTRATIONS.at[week -1, f'Total {Year}'] += 1

    if items['Dormant/Conversion'] == 'New Registration':
        df_REGISTRATIONS.at[week -1, f'New {Year}'] += 1
        
        if items['Contractor Relations'] == 'Contractor Relations':
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR New {Year}'] += 1
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR Total {Year}'] += 1
        if items['Contractor Relations'] == 'Internal Sales':
            df_INTERNAL_SALES.at[week -1, f'IS New {Year}'] += 1
            df_INTERNAL_SALES.at[week -1, f'IS Total {Year}'] += 1
            
    if items['Dormant/Conversion'] == 'Conversion':
        df_REGISTRATIONS.at[week -1, f'Conversion {Year}'] += 1
        
        if items['Contractor Relations'] == 'Contractor Relations':
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR Conversion {Year}'] += 1
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR Total {Year}'] += 1
        if items['Contractor Relations'] == 'Internal Sales':
            df_INTERNAL_SALES.at[week -1, f'IS Conversion {Year}'] += 1
            df_INTERNAL_SALES.at[week -1, f'IS Total {Year}'] += 1
            
    if items['Dormant/Conversion'] == 'Dormant':
        df_REGISTRATIONS.at[week -1, f'Dormant {Year}'] += 1
        
        if items['Contractor Relations'] == 'Contractor Relations':
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR Dormant {Year}'] += 1
            df_CONTRACTOR_RELATIONS.at[week -1, f'CR Total {Year}'] += 1
        if items['Contractor Relations'] == 'Internal Sales':
            df_INTERNAL_SALES.at[week -1, f'IS Dormant {Year}'] += 1
            df_INTERNAL_SALES.at[week -1, f'IS Total {Year}'] += 1

df_MRRR = df_MRRR.sort_values(by='Created Time')

df_MRRR['Created Time'] = df_MRRR['Created Time'].dt.strftime('%d/%m/%Y').astype(str)

df_REGISTRATIONS = df_REGISTRATIONS.replace(0, np.nan)
df_INTERNAL_SALES = df_INTERNAL_SALES.replace(0, np.nan)
df_CONTRACTOR_RELATIONS = df_CONTRACTOR_RELATIONS.replace(0, np.nan)

print('___________________________________________________________________')
print('')
print('Loading Accounts...')
print(datetime.datetime.now().time())

df_ACCOUNTS = pd.read_csv(df_ACCOUNTS ,encoding = 'latin')
df_ACCOUNTS.columns = df_ACCOUNTS.iloc[0]
df_ACCOUNTS = df_ACCOUNTS.drop(0 ,axis = 0)
df_ACCOUNTS.reset_index(drop=True)
df_ACCOUNTS = df_ACCOUNTS[['Office Number','Account Owner']]
df_ACCOUNTS = df_ACCOUNTS.dropna()
df_ACCOUNTS['Office Number'] = df_ACCOUNTS['Office Number'].astype(int)
df_ACCOUNTS = df_ACCOUNTS.drop_duplicates(subset=['Office Number'], keep='first')

print('___________________________________________________________________')
print('')
print('Loading PAYE Data for Master Peace Workers...')
print(datetime.datetime.now().time())

df_PAYE_DATA = pd.read_csv(df_PAYE_DATA, usecols = ['Agency','Payno', 'Name', 'Type','Week'])
df_PAYE_DATA.drop_duplicates(subset ="Payno",
                     keep = "last", inplace = True)
df_PAYE_DATA['Agency'] = df_PAYE_DATA['Agency'].str.upper()
df_PAYE_DATA.loc[df_PAYE_DATA['Agency'] == 'MASTER PEACE', 'Agency'] = 'MASTER PEACE RECRUITMENT'
df_PAYE_DATA = df_PAYE_DATA[df_PAYE_DATA['Agency'] != 'CORRIE']
df_PAYE_DATA = df_PAYE_DATA.dropna()

print('___________________________________________________________________')
print('Loading Corrie Summary...')
print(datetime.datetime.now().time())

CORRIE_SUMMARY_TOTAL = None

try:
    df_CORRIE_SUMMARY = pd.read_excel(df_CORRIE_SUMMARY, sheet_name= Year2)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY[['ADVANCE','Unnamed: 1']]
    df_CORRIE_SUMMARY.columns = df_CORRIE_SUMMARY.iloc[3, :]
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY.drop(0, axis = 0)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY.drop(1, axis = 0)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY.drop(2, axis = 0)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY.drop(3, axis = 0)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY.drop(4, axis = 0)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY[:-1].reset_index(drop=True)
    df_CORRIE_SUMMARY = df_CORRIE_SUMMARY[['WEEK','WORKERS']].fillna(0).astype(int)
    CORRIE_SUMMARY_TOTAL = df_CORRIE_SUMMARY.at[Week - 1,'WORKERS']
except FileNotFoundError:
    while isinstance(CORRIE_SUMMARY_TOTAL, int) == False:
        print('___________________________________________________________________')
        try:
            CORRIE_SUMMARY_TOTAL = int(input('No Corrie Summary Found, Enter number of Corrie workers: '))
        except ValueError:
            print('')
            print('Enter an integer number.')


df_JOINERS_ERROR_IO = pd.read_csv(df_JOINERS_ERROR_IO,encoding = 'latin',usecols = ['OFF_PAYNO','Sdc Option','Email Address','WEEKS_PAID','REWARDS'])

df_JOINERS_ERROR_IO = df_JOINERS_ERROR_IO.rename(columns={"OFF_PAYNO":"PAYNO"})
df_JOINERS_ERROR_IO = df_JOINERS_ERROR_IO.rename(columns={"Sdc Option":"Type"})

df_JOINERS_ERROR_AXM = pd.read_csv(df_JOINERS_ERROR_AXM,encoding = 'latin',usecols = ['OFF_PAYNO','Sdc Option','Email Address','WEEKS_PAID','REWARDS'])

df_JOINERS_ERROR_AXM = df_JOINERS_ERROR_AXM.rename(columns={"OFF_PAYNO":"PAYNO"})
df_JOINERS_ERROR_AXM = df_JOINERS_ERROR_AXM.rename(columns={"Sdc Option":"Type"})

df_FEES_RETAINED_IO = pd.merge(df_FEES_RETAINED_IO, df_JOINERS_ERROR_IO)
df_FEES_RETAINED_AXM = pd.merge(df_FEES_RETAINED_AXM, df_JOINERS_ERROR_AXM)

print('___________________________________________________________________')
print('')
print('Creating Last Paid csv for crm upload...')
print(datetime.datetime.now().time())

df_LAST_PAID_IO = df_FEES_RETAINED_IO.drop(['REWARDS'], axis = 1)

df_LAST_PAID_IO = df_LAST_PAID_IO.drop(['Client Name'], axis = 1)
df_LAST_PAID_IO = df_LAST_PAID_IO.drop(['Margins'], axis = 1)
df_LAST_PAID_IO = df_LAST_PAID_IO.drop(['Solution'], axis = 1)
df_LAST_PAID_IO = df_LAST_PAID_IO.drop(['Type'], axis = 1)
df_LAST_PAID_IO.columns = [['Payno','Last Name','First Name','Date Last Paid','Email Address','Weeks Paid']]

for i, item in df_LAST_PAID_IO.iterrows():
    count = str(item['Weeks Paid']).count(',') + 1
    df_LAST_PAID_IO.at[i,'Weeks Paid'] = count

df_LAST_PAID_AXM = df_FEES_RETAINED_AXM.drop(['REWARDS'], axis = 1)

df_LAST_PAID_AXM = df_LAST_PAID_AXM.drop(['Client Name'], axis = 1)
df_LAST_PAID_AXM = df_LAST_PAID_AXM.drop(['Margins'], axis = 1)
df_LAST_PAID_AXM = df_LAST_PAID_AXM.drop(['Solution'], axis = 1)
df_LAST_PAID_AXM = df_LAST_PAID_AXM.drop(['Type'], axis = 1)
df_LAST_PAID_AXM.columns = [['Payno','Last Name','First Name','Date Last Paid','Email Address','Weeks Paid']]

for i, item in df_LAST_PAID_AXM.iterrows():
    count = str(item['Weeks Paid']).count(',') + 1
    df_LAST_PAID_AXM.at[i,'Weeks Paid'] = count
    
df_LAST_PAID_IO.to_csv(last_paid_path_io, encoding='utf-8', index=False)
df_LAST_PAID_AXM.to_csv(last_paid_path_axm, encoding='utf-8',index=False)

print('___________________________________________________________________')
print('')
print('Created Last Paid csv ready for crm upload.')
print(datetime.datetime.now().time())

df_FEES_RETAINED_IO = df_FEES_RETAINED_IO.drop(['WEEKS_PAID'], axis = 1)
df_FEES_RETAINED_AXM = df_FEES_RETAINED_AXM.drop(['WEEKS_PAID'], axis = 1)
df_FEES_RETAINED = df_FEES_RETAINED_IO.append(df_FEES_RETAINED_AXM)
df_REWARDS = df_FEES_RETAINED[df_FEES_RETAINED['REWARDS'] == 'Yes']
df_FEES_RETAINED = df_FEES_RETAINED.drop(['Email Address'], axis = 1)

print('___________________________________________________________________')
print('')
print('Loading £0 Margins...')
print(datetime.datetime.now().time())

df_0_MARGINS = pd.concat([pd.read_csv(df_0_MARGINS_IO,encoding = 'latin'),                           
                          pd.read_csv(df_0_MARGINS_AXM,encoding = 'latin')])

df_0_MARGINS = df_0_MARGINS.rename(columns={"OFF_PAYNO":"PAYNO"})

df_0_MARGINS.drop_duplicates(subset ="PAYNO",
                     keep = "first", inplace = True)

df_0_MARGINS = df_0_MARGINS[df_0_MARGINS['PAYNO'].notna()]
df_0_MARGINS['CHQDATE'] = pd.to_datetime(df_0_MARGINS['CHQDATE'], format='%d/%m/%y')

df_0_MARGINS['PAYNO'] = df_0_MARGINS['PAYNO'].astype(str)
df_0_MARGINS["Reason"] = 0
df_0_MARGINS.loc[df_0_MARGINS["SMP"] > 0, "Reason"] = 'SMP'
df_0_MARGINS.loc[df_0_MARGINS["Spp"] > 0, "Reason"] = 'Spp'
df_0_MARGINS.loc[df_0_MARGINS["SSP"] > 0, "Reason"] = 'SSP'
df_0_MARGINS.loc[df_0_MARGINS["Paydesc"] == "Expenses", "Reason"] = "Expenses"
df_0_MARGINS.loc[df_0_MARGINS["Fee Code"] == 0, "Reason"] = "£0 Margin Agreed"
df_0_MARGINS.loc[df_0_MARGINS["Paydesc"] == "Additional Pay", "Reason"] = "Additional Pay"
df_0_MARGINS.loc[df_0_MARGINS["Fee Code"] == "GSL", "Reason"] = 'GSL'
df_0_MARGINS.loc[df_0_MARGINS["GROSS"] < 100, "Reason"] = "Low Pay"

df_0_MARGINS = df_0_MARGINS[df_0_MARGINS['Reason'].isin([0,"£0 Margin Agreed"])]
df_0_MARGINS = df_0_MARGINS[df_0_MARGINS['Paydesc'] == 'Basic Pay']
df_0_MARGINS = df_0_MARGINS[df_0_MARGINS['Pay Rates'] >= 0]
df_0_MARGINS = df_0_MARGINS.rename(columns={'COMPNAME':'Client Name'})
df_0_MARGINS = df_0_MARGINS.rename(columns={'SURNAME':'Surname'})
df_0_MARGINS = df_0_MARGINS.rename(columns={'FIRSTNAME':'Forename'})
df_0_MARGINS = df_0_MARGINS.rename(columns={'MANAGEMENT FEE':'Margins'})
df_0_MARGINS = df_0_MARGINS.rename(columns={'TYPE':'Type'})
df_0_MARGINS = df_0_MARGINS.rename(columns={'Sdc Option':'Solution'})

df_PREV_WEEK = df_FEES_RETAINED[df_FEES_RETAINED['Margins'] > 0]
df_0_MARGINS = df_0_MARGINS[~df_0_MARGINS["PAYNO"].isin(df_PREV_WEEK['PAYNO'])]
df_0_MARGINS['REWARDS'] = 'No'
df_0_MARGINS_ADDED = pd.DataFrame([],columns=df_0_MARGINS.columns)
    
for i, item in df_0_MARGINS.iterrows():
    print('___________________________________________________________________')
    print('')
    print(f'Listing {len(df_0_MARGINS)} £0 Margins...')
    print('___________________________________________________________________')
    print('')
    print(item)
    print('___________________________________________________________________')
    option1 =input('Type "y" if is a Agreed 0 margin, Type "add" to add as a margin manually (for invoiced margins) and type "n" to remove from £0 margins reports (Other reasons)("y" / "n" / "add"): ')
    if option1  == 'n':
        df_0_MARGINS = df_0_MARGINS.drop(i ,axis = 0)
    elif option1  == 'add':
        print('')
        Quantity = input('Enter number of margins this entry is equal too: ')
        item['Margins'] = 1
        for j in range(0, int(Quantity), 1):
            df_FEES_RETAINED = df_FEES_RETAINED.append(item[['Client Name','PAYNO','Surname','Forename','Margins','CHQDATE','Type','REWARDS']])
            df_0_MARGINS_ADDED = df_0_MARGINS_ADDED.append(item[['Client Name','PAYNO','Surname','Forename','Margins','CHQDATE','Type','REWARDS']])
        df_0_MARGINS = df_0_MARGINS.drop(i ,axis = 0)
    else:
        continue

choice1 = input('Do you want to add any margins manually ? ("y" / "n"): ')

c_CHQDATE = df_FEES_RETAINED.iloc[1][5]

while choice1 == 'y':
        print('')
        Agency = input('Enter Workers Agency: ').upper()
        print('')
        Solution = input('Enter Workers Solution: ').upper()
        print('')
        Quantity = input('Enter number of margins this entry is equal too: ')
        while Quantity.isnumeric() == False:
            print('')
            print('Input is not an integer number')
            print('')
            Quantity = input('Enter number of margins this entry is equal too: ')
            print('')
        print('___________________________________________________________________')
        print('')
        df_ITEM_temp = pd.DataFrame([[Agency, '','','',1 , c_CHQDATE, Solution, Solution, 'No']], columns=df_FEES_RETAINED.columns)
        print(df_ITEM_temp)
        print('___________________________________________________________________')
        print('')
        choice2 = input('Are you sure you want to add this margin ? ("y" / "n"): ')
        
        if choice2 == 'y':
            for i in range(0, int(Quantity), 1):
                df_FEES_RETAINED = pd.concat([df_FEES_RETAINED, df_ITEM_temp])
        else: 
            while choice2 != 'n':
                print('')
                print('Input only accepts "y" for yes and "n" for no.')
                print('')
                choice2 = input('Are you sure you want to add this margin ? ("y" / "n"): ')
        choice1 = input('Do you want to add any margins manually ? ("y" / "n"): ')


                     
for i, item in df_PAYE_DATA.iterrows():
    SURNAME = item['Name'].split()[1]
    FORENAME = item['Name'].split()[0]
    df_ITEM_temp = pd.DataFrame([[item['Agency'],item['Payno'],SURNAME,FORENAME,1 , c_CHQDATE,'PAYE','Advance PAYE','No']], columns=df_FEES_RETAINED.columns)
    df_FEES_RETAINED = pd.concat([df_FEES_RETAINED, df_ITEM_temp])

if CORRIE_SUMMARY_TOTAL > 0:
    for i in range(1, CORRIE_SUMMARY_TOTAL +1, 1):
        df_ITEM_temp = pd.DataFrame([['CORRIE',np.nan,np.nan,np.nan,1 , c_CHQDATE,'PAYE','Advance PAYE','No']], columns=df_FEES_RETAINED.columns)
        df_FEES_RETAINED = pd.concat([df_FEES_RETAINED, df_ITEM_temp])
    
df_FEES_RETAINED = pd.merge(df_FEES_RETAINED, df_CLIENTS, how = 'left')
df_FEES_RETAINED = pd.merge(df_FEES_RETAINED, df_ACCOUNTS , left_on = 'OFFNO', right_on = 'Office Number', how = 'left')
df_FEES_RETAINED = df_FEES_RETAINED.drop(['Office Number'], axis = 1)
df_FEES_RETAINED = df_FEES_RETAINED.rename(columns={'Account Owner':'CRM'})

df_FEES_RETAINED.loc[(df_FEES_RETAINED["Client Name"]).str.contains('SEARCH CONSULTANCY') , 'Client Name'] = 'SEARCH CONSULTANCY'
df_FEES_RETAINED.loc[(df_FEES_RETAINED["Client Name"]) == 'MANPOWER INVOICE', 'Client Name'] = 'MANPOWER'
df_FEES_RETAINED.loc[(df_FEES_RETAINED["Client Name"]).str.contains('THE EDUCATION NETWORK') , 'Client Name'] = 'THE EDUCATION NETWORK'
df_FEES_RETAINED.loc[(df_FEES_RETAINED["Client Name"]) == 'JAMES GRAY TRADES LTD', 'Client Name'] = 'JAMES GRAY RECRUITMENT LTD'

df_FEES_RETAINED['Tax Week'] = int(Week)

df = pd.concat([df, df_FEES_RETAINED])

print('___________________________________________________________________')
print('')
print('Formating Core Margin data...')
print(datetime.datetime.now().time())

df.loc[(df["Client Name"] == 'ALEXANDER MANN') & (df['Solution'] != np.nan), 'Type'] = 'Umbrella'
df.loc[(df["Client Name"] == 'ALEXANDER MANN') & (df['Solution'] != np.nan), 'Solution'] = 'PAYE'

df.loc[(df['Solution'] == 'PAYE') & (df['Type'] == 'Not Under SDC'), 'Type'] = 'Umbrella'
df.loc[(df['Solution'] == 'PAYE') & (df['Type'] == 'Under SDC'), 'Type'] = 'Umbrella'

df.loc[(df['Solution'] == 'SE') | (df['Type'] == 'SE'), ['Solution','Type']] = 'Self Employed Gross'

df.loc[(df['Type'] == 'Mileage Only') & (df['Tax Week'] > 39) , 'Type'] = 'Umbrella w/ Mileage'

df.loc[df['Type'] == 'Mileage Only', 'Type'] = 'Umbrella w/ Expenses'

df.loc[df['Type'] == 'Fixed Expenses', 'Type'] = 'Umbrella w/ Expenses'

df.loc[(df['Type'] == 'CIS') | (df['Solution'] == 'CIS'), 'Solution'] = 'CIS'
df.loc[(df['Type'] == 'CIS') | (df['Solution'] == 'CIS'), 'Type'] = 'CIS'

df.loc[(df['Solution'] == 'Under SDC'), 'Solution'] = 'Umbrella no Expenses'
df.loc[(df['Type'] == 'Under SDC'), 'Type'] = 'Umbrella no Expenses'

df.loc[df['CRM'].isin(CORE_ACCOUNTS) == True, 'Account'] = 'Core Account'
df.loc[df['CRM'].isin(CORE_ACCOUNTS) == False, 'Account'] = 'Unmanned Account'
df.loc[df['CRM'].isin(CORE_ACCOUNTS) == False, 'CRM'] = 'Other Account'

df['Margins'] = df['Margins'].astype(float)
df = df[df['Margins'] > 0]
df_null = df[df['OFFNO'].isnull()]
 
df_MARGINS.at[Week - 1, f'Margins {Year}'] = len(df_FEES_RETAINED)

df_ACTUAL = pd.pivot_table(df, values=['Margins'], index=['Client Name','CRM'],
                    columns=['CHQDATE'], aggfunc=len, fill_value=0)
df_ACTUAL.columns = df_ACTUAL.columns.droplevel(0)
df_ACTUAL.columns.name = None
df_ACTUAL.columns = df_ACTUAL.columns.to_list()
df_ACTUAL = df_ACTUAL[sorted(df_ACTUAL.columns)]
df_ACTUAL.columns = df_ACTUAL.columns.strftime('%d/%m/%Y').astype(str)
#df_ACTUAL['Grand Total'] = df_ACTUAL.sum(axis=1)

df_FORECAST = pd.read_excel(df_path, sheet_name= 'Forecast').fillna(0)
df_FORECAST = df_FORECAST.set_index(['Client Name','CRM'], drop=True).drop(('Grand Total','All Accounts'),axis = 0).drop('Grand Total',axis = 1)
#df_FORECAST.columns = pd.to_datetime(df_FORECAST.columns.tolist())
#df_FORECAST = df_FORECAST[sorted(df_FORECAST.columns)]
#df_FORECAST.columns = df_FORECAST.columns.strftime('%d/%m/%Y').astype(str)
#df_FORECAST = df_FORECAST.sort_values(by='Client Name')
df['CHQDATE'] = df['CHQDATE'].dt.strftime('%d/%m/%Y').astype(str)

columns = pd.Series(df_FORECAST.columns)

columns = pd.MultiIndex.from_product([df_FORECAST.columns,["Forecast","Actual","Variance"]],
                                     names=['CHQDATE','Margins'])

df_VARIANCE = pd.DataFrame(np.zeros((df_FORECAST.shape[0],df_FORECAST.shape[1]*3)),
                    columns=columns,
                    index=df_FORECAST.index)

df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Forecast'] = df_FORECAST.values


        
for column, value in df_ACTUAL.iteritems():
    for i, item in value.iteritems():
        if item != 0 and column in df_FORECAST.columns:
                if i[0] in df_FORECAST.index.get_level_values(0):
                    df_VARIANCE.loc[i[0],(column,'Actual')] = df_VARIANCE.loc[i[0],(column,'Actual')].values + item
                else:
                    df_VARIANCE.loc[('Other Account','Other Account'),(column,'Actual')] = df_VARIANCE.loc[('Other Account','Other Account'),(column,'Actual')] + item


for column in df_FORECAST.columns:
    if column not in df_ACTUAL.columns:
        df_VARIANCE = df_VARIANCE.drop(column,axis = 1)
        

df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Variance'] = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Actual'].values - df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Forecast'].values
df_VARIANCE[('Grand Total','Forecast')] = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Forecast'].values.sum(axis=1)
df_VARIANCE[('Grand Total','Actual')] = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Actual'].values.sum(axis=1)
df_VARIANCE[('Grand Total','Variance')] = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Variance'].values.sum(axis=1)


for column, value in df_VARIANCE.iteritems():
    df_VARIANCE.loc[('Grand Total','All Accounts'),column] = df_VARIANCE.loc[:,column].sum()
    
for column, value in df_FORECAST.iteritems():
    df_FORECAST.loc[('Grand Total','All Accounts'),column] = df_FORECAST.loc[:,column].sum()

df_FORECAST['Grand Total'] = df_FORECAST.sum(axis=1)

df_VARIANCE_TOTALS = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Variance'].drop(('Grand Total'), axis = 1)

df_CRM_TOTALS = df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Variance']
df_CRM_TOTALS.columns = df_CRM_TOTALS.columns.droplevel(1)
df_CRM_TOTALS = df_CRM_TOTALS.rename(columns = {'Grand Total':'Total'})
df_CRM_TOTALS = df_CRM_TOTALS['Total'].reset_index().drop(('Client Name'), axis = 1)
df_CRM_TOTALS = df_CRM_TOTALS.groupby(['CRM'])['Total'].sum().reset_index()
df_CRM_TOTALS = df_CRM_TOTALS[df_CRM_TOTALS['CRM'] != 'All Accounts']
CRM_GrandTotal = df_CRM_TOTALS.sum()
CRM_GrandTotal.loc[('CRM')] = 'Grand Total'
df_CRM_TOTALS = df_CRM_TOTALS.append(CRM_GrandTotal,ignore_index=True)

df.loc[(df["Client Name"] == 'ALEXANDER MANN') & (df['Solution'] != np.nan), 'Solution'] = 'PAYE'

df_VARIANCE_TOTALS.columns = df_VARIANCE_TOTALS.columns.droplevel(1)

df_VARIANCE_TOTALS = df_VARIANCE_TOTALS.reset_index(drop = True)

df_FORECAST_TOTALS = df_FORECAST.loc[('Grand Total','All Accounts'),:].drop('Grand Total', axis = 0).reset_index()

df_FORECAST_TOTALS.columns = df_FORECAST_TOTALS.columns.droplevel(1)

df_FORECAST_TOTALS = df_FORECAST_TOTALS.rename(columns = {'index':'Tax Week'})

df_FORECAST_TOTALS['Tax Week'] = pd.to_datetime(df_FORECAST_TOTALS['Tax Week'], format='%d/%m/%Y')

for i, items in df_FORECAST_TOTALS.iterrows():
    df_FORECAST_TOTALS.at[i, 'Tax Week'] = tax_week_calc(items['Tax Week'].isocalendar()[1] + 1)
    
df_FORECAST_TOTALS = df_FORECAST_TOTALS.astype(int).sort_values(by=['Tax Week'])
df_FORECAST_TOTALS = df_FORECAST_TOTALS.reset_index(drop=True).rename(columns = {'Grand Total':'Forecast Totals'})

df_TOTALS = pd.merge(df_MARGINS.reset_index(drop=True), df_FORECAST_TOTALS.reset_index(drop=True),how='left')

df_REG_TOTALS = pd.merge(df_REGISTRATIONS.reset_index(drop=True), df_CONTRACTOR_RELATIONS.reset_index(drop=True))
df_REG_TOTALS = pd.merge(df_REG_TOTALS, df_INTERNAL_SALES.reset_index(drop=True))

print('___________________________________________________________________')
print('')
print('Adding Data to New Excel Document...')
print(datetime.datetime.now().time())

wb = xlsxwriter.Workbook(df_path_new)

print('___________________________________________________________________')
print('')
print('Writing Margin Summary...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Margin Summary')
ws.set_tab_color('#3F43AD')

title_merge_format = wb.add_format({'align': 'center',
                              'font_size' : 30,
                              'bg_color': '#3F43AD',
                              'font_color': '#FFFFFF',
                              'bold':     True})

cell_merge1 = wb.add_format({'bold':     True,
                             'bg_color': '#F2F2F2'})

cell_merge1_border = wb.add_format({'bold':     True,
                             'bg_color': '#F2F2F2',
                             'border':1})

cell_merge2 = wb.add_format({'bold':     False,
                             'align': 'center',
                             'bg_color': '#DAEEF3'})

cell_merge3 = wb.add_format({'bold':     True,
                             'align': 'center',
                             'bg_color': '#3F43AD',
                             'font_color': '#FFFFFF'})

cell_format = wb.add_format({'bold':     False,
                             'align': 'center',
                             'bg_color': '#DAEEF3'})

cell_format1 = wb.add_format({'bold':     False,
                             'align': 'center',
                             'bg_color': '#DAEEF3'})

cell_format1.set_bottom()
cell_format1.set_border_color('#3E95AC')

result_format1 = wb.add_format({'bold': True,
                             'align': 'center',
                             'bg_color': '#DAEEF3',
                             'font_size':30})

format1 = wb.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006'})

format2 = wb.add_format({'bg_color': '#C6EFCE',
                               'font_color': '#006100'})

border_format=wb.add_format({'border':1,
                            'align':'left',
                            'font_size':10})

border_format.set_border_color('#3E95AC')
# ws.add_table(tab_REF, {'name': 'Clients',
#                        'style': 'Table Style Dark 2'})

ws.merge_range('A1:T1', 'Margin Summary', title_merge_format)

ws.write('B3', 'SELECT TAX WEEK', cell_merge1)
ws.write('C3', '', cell_merge1)
# ws.data_validation('C3', {'validate':'list', 'source': list(pd.unique(df['Tax Week']))})
ws.write('D3', '', cell_merge1_border)                  
ws.data_validation('D3', {'validate':'list', 'source': list(pd.unique(df['Tax Week']))})

CORE_ACCOUNTS.insert(0,'All')
CORE_ACCOUNTS.append('Other Account')

ws.write('F3', 'SELECT ACCOUNT', cell_merge1)
ws.write('G3', '', cell_merge1)
ws.write('H3', '', cell_merge1_border)
ws.data_validation('H3',{'validate':'list', 'source': CORE_ACCOUNTS})

ws.write('J3', 'SELECT CLIENT', cell_merge1)
ws.write('K3', '', cell_merge1)
ws.write('L3', '', cell_merge1_border)
ws.data_validation('L3', {'validate':'list', 'source' : "='Margin Summary'!Y:Y"})

ws.write('N3', 'SELECT SOLUTION', cell_merge1)
ws.write('O3', '', cell_merge1)
ws.write('P3', '', cell_merge1_border)
ws.data_validation('P3', {'validate':'list', 'source': list(pd.unique(df['Type']))})

ws.write('R3', 'SELECT REWARDS', cell_merge1)
ws.write('S3', '', cell_merge1)
ws.write('T3', '', cell_merge1_border)
ws.data_validation('T3', {'validate':'list', 'source': ['Yes, No']})

ws.set_column('C:C', 20)
ws.set_column('G:G', 20)
ws.set_column('K:K', 20)
ws.set_column('O:O', 20)
ws.set_column('S:S', 20)

ws.set_row(6, 50)
ws.set_row(7, 6)
ws.set_row(8, 3)

ws.write('C6', 'Margins by Date Range', cell_format)
ws.merge_range('B5:D5', '', cell_merge2)
ws.merge_range('B8:D8', '', cell_merge2)
ws.merge_range('B9:D9', '', cell_merge3)
ws.write('B7','', cell_format)
ws.write('B6','', cell_format)
ws.write('D7','', cell_format)
ws.write('D6','', cell_format)
ws.write('C7',('=COUNTIF({sheet}!K:K, D3)').format(sheet = "'Core Margin Data'"), result_format1)
ws.write('C6','Margins by Week', cell_format)

ws.merge_range('G6', 'Margins by Account', cell_format)
ws.merge_range('F5:H5', '', cell_merge2)
ws.merge_range('F8:H8', '', cell_merge2)
ws.merge_range('F9:H9', '', cell_merge3)
ws.write('F7','', cell_format)
ws.write('F6','', cell_format)
ws.write('H7','', cell_format)
ws.write('H6','', cell_format)
ws.write('G7',('=IF(H3="All",COUNTIFS({sheet}!K:K, D3),COUNTIFS({sheet}!K:K, D3, {sheet}!J:J, H3))').format(sheet = "'Core Margin Data'"), result_format1)
#">=" & C3, {sheet}!K:K, "<=" &
ws.write('G6','Margins by Account', cell_format)

ws.merge_range('K6', 'Margins by Client', cell_format)
ws.merge_range('J5:L5', '', cell_merge2)
ws.merge_range('J8:L8', '', cell_merge2)
ws.merge_range('J9:L9', '', cell_merge3)
ws.write('J7','', cell_format)
ws.write('J6','', cell_format)
ws.write('L7','', cell_format)
ws.write('L6','', cell_format)
ws.write('K7',('=COUNTIFS({sheet}!K:K, D3, {sheet}!A:A, L3)').format(sheet = "'Core Margin Data'"), result_format1)
ws.write('K6','Margins by Client', cell_format)

ws.merge_range('O6', 'Margins by Solution', cell_format)
ws.merge_range('N5:P5', '', cell_merge2)
ws.merge_range('N8:P8', '', cell_merge2)
ws.merge_range('N9:P9', '', cell_merge3)
ws.write('N7','', cell_format)
ws.write('N6','', cell_format)
ws.write('P7','', cell_format)
ws.write('P6','', cell_format)
ws.write('O7',('=IF(H3="All",COUNTIFS({sheet}!K:K, D3, {sheet}!A:A, L3,{sheet}!H:H, P3),COUNTIFS({sheet}!K:K, D3,{sheet}!J:J, H3,{sheet}!H:H, P3))').format(sheet = "'Core Margin Data'"), result_format1)
ws.write('O6','Margins by Solution', cell_format)

ws.merge_range('S6', 'Margins w/ Rewards', cell_format)
ws.merge_range('R5:T5', '', cell_merge2)
ws.merge_range('R8:T8', '', cell_merge2)
ws.merge_range('R9:T9', '', cell_merge3)
ws.write('R7','', cell_format)
ws.write('R6','', cell_format)
ws.write('T7','', cell_format)
ws.write('T6','', cell_format)
ws.write('S7',('=IF(H3="All",COUNTIFS({sheet}!K:K, D3, {sheet}!M:M, T3),COUNTIFS({sheet}!K:K, D3, {sheet}!J:J, H3,{sheet}!M:M, T3))').format(sheet = "'Core Margin Data'"), result_format1)#add rewards column
ws.write('S6','Margins w/ Rewards', cell_format)

ws.write('Y1','Client Name',cell_merge3)
ws.write('Z1','OFF NO',cell_merge3)
ws.set_column('Z:Z', 20)
ws.set_column('Y:Y', 40)

df_CLIENTS.fillna('',inplace=True)
df_CLIENTS = df_CLIENTS.sort_values(by = 'Client Name')
df_CLIENTS = df_CLIENTS.reset_index(drop=True)

for i, row in df_CLIENTS.iterrows():
    n = 25
    for item in row:
        REF_1 = ('{column}{row}').format(column= get_column_letter(n),row = i + 2)
        ws.write(REF_1,item,cell_format1)
        n += 1

ws.write('V1','CRM',cell_merge3)
ws.write('W1','Totals',cell_merge3)
ws.set_column('V:V', 20)

for i, row in df_CRM_TOTALS.iterrows():
    n = 22
    for item in row:
        REF_1 = ('{column}{row}').format(column= get_column_letter(n),row = i + 2)
        ws.write(REF_1,item,cell_format1)
        n += 1
        if isinstance(item, str) == False:
            ws.conditional_format(REF_1, {'type': 'cell',
                                          'criteria': '<',
                                          'value': 0,
                                          'format': format1})
            ws.conditional_format(REF_1, {'type': 'cell',
                                          'criteria': '>=',
                                          'value': 0,
                                          'format': format2})


tab_REF = ('Y1:Z{length}').format(length = len(df_CLIENTS)+1)

ws.conditional_format(tab_REF, { 'type' : 'no_blanks' , 'format' : border_format})

ws.hide_gridlines(2)

ws.hide_row_col_headers()

Margin_Summary = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

Margin_Summary.add_series({
    'name':       'Margins '+Yearpp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$B$2:$B$53",
    'line': {'width': 1.5},
    'smooth':     True
})
Margin_Summary.add_series({
    'name':       'Margins '+Yearp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$C$2:$C$53",
    'line': {'width': 1.5},
    'smooth':     True
})
Margin_Summary.add_series({
    'name':       'Margins '+Year,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$D$2:$D$53",
    'line': {'width': 1.5},
    'marker': {'type': 'automatic'},
    'smooth':     True
})

# 'trendline': {
#         'type': 'linear',
#         'name': 'Trend',
#         'forward': 0.5,
#         'backward': 0.5,
#         'line': {
#             'color': 'red',
#             'width': 1,
#             'dash_type': 'long_dash',
#         },

Margin_Summary.add_series({
    'name':       'Forecast',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$E$2:$E$53",
    'marker': {'type': 'automatic'},
    'smooth':     True,
    'line': {'width': 1,
            'dash_type': 'long_dash',
            'color': 'red'}
})

Margin_Summary.set_x_axis({'name':'Tax Week'})
Margin_Summary.set_y_axis({'name':'Margins'})

Margin_Summary.set_style(2)

#Margin_Summary.set_title({'name': 'Margin Summary'})

ws.insert_chart('A10', Margin_Summary, {'x_scale': 3.53, 'y_scale': 2})

print('___________________________________________________________________')
print('')
print('Writing Excecutive Summary...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Excecutive Summary')
ws.set_tab_color('#3F43AD')

df_C_MARGINS = df_TOTALS[df_TOTALS["Tax Week"] <= Week]
df_C_MARGINS = df_C_MARGINS[["Tax Week",f"Margins {Year}"]]

df_MARGINS_W_REWARDS = pd.pivot_table(df, values=['Margins'], index=['Tax Week'],
                                      columns=['REWARDS'], aggfunc=len, fill_value=0)

df_MARGINS_W_REWARDS.columns = df_MARGINS_W_REWARDS.columns.droplevel(0)
df_MARGINS_W_REWARDS = df_MARGINS_W_REWARDS.reset_index()
df_MARGINS_W_REWARDS = df_MARGINS_W_REWARDS.drop('No',axis=1)
df_C_MARGINS = pd.merge(df_C_MARGINS,df_MARGINS_W_REWARDS)
df_C_MARGINS = df_C_MARGINS.rename(columns = {'Yes':'REWARDS'})


df_C_MARGINS.fillna('',inplace=True)

tab_REF = ('A1:{width}{length}').format(length = len(df_C_MARGINS)+1, width = get_column_letter(len(df_C_MARGINS.columns)))

ws.write('A1','Margin Totals',cell_merge3)

for j, column in enumerate(df_C_MARGINS.columns.values):
    REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = 2)
    ws.write(REF_1,column,cell_merge3)

ws.set_column('A:A', 20)
ws.set_column('B:B', 20)
ws.set_column('E:E', 20)
ws.set_column('F:F', 20)
ws.set_column('H:H', 20)
ws.set_column('N:N', 20)
ws.set_column('L:L', 20)
ws.set_column('J:J', 20)
ws.set_column('I:I', 20)
ws.set_column('O:O', 20)
ws.set_column('P:P', 20)
ws.set_column('S:S', 20)
ws.set_column('X:X', 20)
ws.set_column('W:W', 20)
ws.set_column('T:T', 20)

for i, row in df_C_MARGINS.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 3)
        ws.write(REF_1,item,cell_format1)
        j += 1

tab_REF = ('A1:{width}{length}').format(length = len(df_C_MARGINS)+1, width = get_column_letter(len(df_C_MARGINS.columns)))
ws.conditional_format(tab_REF, {'type' : 'no_blanks' , 'format' : border_format})

# ws.add_table(tab_REF, {'name': 'Current Margins w/ Rewards',
#                        'style': 'Table Style Dark 2'})

df_MARGINS_BY_TYPE = pd.pivot_table(df, values=['Margins'], index=['Tax Week'],
                                      columns=['Type'], aggfunc=len, fill_value=0)

df_MARGINS_BY_TYPE.columns = df_MARGINS_BY_TYPE.columns.droplevel(0)
df_MARGINS_BY_TYPE = df_MARGINS_BY_TYPE.reset_index()

ws.write(('{startcol}1').format(startcol = get_column_letter(len(df_C_MARGINS.columns)+2)),'Margins by Type',cell_merge3)

for j, column in enumerate(df_MARGINS_BY_TYPE.columns.values):
    REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns)+ j + 2), row = 2)
    ws.write(REF_1,column,cell_merge3)

for i, row in df_MARGINS_BY_TYPE.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns)+ j + 2), row = i + 3)
        ws.write(REF_1,item,cell_format1)
        j += 1

tab_REF = ('{startcol}1:{width}{length}').format(startcol = get_column_letter(len(df_C_MARGINS.columns)+1),length = len(df_MARGINS_BY_TYPE)+1, width = get_column_letter(len(df_C_MARGINS.columns) + 1 + len(df_MARGINS_BY_TYPE.columns)))
ws.conditional_format(tab_REF, { 'type' : 'no_blanks' , 'format' : border_format})

# ws.add_table(tab_REF, {'name': 'Margins by Type',
#                        'style': 'Table Style Dark 2'})

df_REG_BY_SOLUTION = pd.pivot_table(df_MRRR, values=['NI Number (Contact)'], index=['Tax Week'],
                                      columns=['Solutions'], aggfunc=len, fill_value=0)

df_REG_BY_SOLUTION.columns = df_REG_BY_SOLUTION.columns.droplevel(0)
df_REG_BY_SOLUTION = df_REG_BY_SOLUTION.reset_index()

df_REG_BY_SOLUTION["Total"] = 0

for i, row in df_REG_BY_SOLUTION.iterrows():
    for column in df_REG_BY_SOLUTION.columns:
        if column not in ["Tax Week","Total"]:
            df_REG_BY_SOLUTION.at[i, "Total"] += row[column]

ws.write(('{startcol}1').format(startcol = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns)+3)),'Registrations by Solution',cell_merge3)

for j, column in enumerate(df_REG_BY_SOLUTION.columns.values):
    REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns)+ j + 3), row = 2)
    ws.write(REF_1,column,cell_merge3)

for i, row in df_REG_BY_SOLUTION.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns)+ j + 3), row = i + 3)
        ws.write(REF_1,item,cell_format1)
        j += 1

tab_REF = ('{startcol}1:{width}{length}').format(startcol = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns)+2),length = len(df_REG_BY_SOLUTION)+1, width = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns)+2 + len(df_REG_BY_SOLUTION.columns)))
ws.conditional_format(tab_REF, { 'type' : 'no_blanks' , 'format' : border_format})
# ws.add_table(tab_REF, {'name': 'Margins by Solution',
#                        'style': 'Table Style Dark 2'})

df_REG_BY_NDC = pd.pivot_table(df_MRRR, values=['NI Number (Contact)'], index=['Tax Week'],
                                      columns=['Dormant/Conversion'], aggfunc=len, fill_value=0)

df_REG_BY_NDC.columns = df_REG_BY_NDC.columns.droplevel(0)
df_REG_BY_NDC = df_REG_BY_NDC.reset_index()

df_REG_BY_NDC["Total"] = 0

for i, row in df_REG_BY_NDC.iterrows():
    for column in df_REG_BY_NDC.columns:
        if column not in ["Tax Week","Total"]:
            df_REG_BY_NDC.at[i, "Total"] += row[column]

tab_REF = ('{startcol}1:{width}{length}').format(startcol = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+3),length = len(df_REG_BY_NDC)+1, width = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+ 3 + len(df_REG_BY_NDC.columns)))

ws.write(('{startcol}1').format(startcol = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+4)),'Registrations by NDC',cell_merge3)

for j, column in enumerate(df_REG_BY_NDC.columns.values):
    REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+ j + 4), row = 2)
    ws.write(REF_1,column,cell_merge3)

for i, row in df_REG_BY_NDC.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+ j + 4), row = i + 3)
        ws.write(REF_1,item,cell_format1)
        j += 1

tab_REF = ('{startcol}1:{width}{length}').format(startcol = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+3),length = len(df_REG_BY_NDC)+1, width = get_column_letter(len(df_C_MARGINS.columns) + len(df_MARGINS_BY_TYPE.columns) + len(df_REG_BY_SOLUTION.columns)+ 3 + len(df_REG_BY_NDC.columns)))
ws.conditional_format(tab_REF, { 'type' : 'no_blanks' , 'format' : border_format})
# ws.add_table(tab_REF, {'name': 'Registrations by NDC',
#                        'style': 'Table Style Dark 2'})

New_Registrations = wb.add_chart({'type': 'line',
                                'subtype':'smooth_with_markers'})


New_Registrations.add_series({
    'name':       'New '+ Yearpp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data!$F$2:$F$53",
})
New_Registrations.add_series({
    'name':       'New '+Yearp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$J$2:$J$53",
})
New_Registrations.add_series({
    'name':       'New '+Year,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$N$2:$N$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

Conversion_Registrations = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

Conversion_Registrations.add_series({
    'name':       'Conversion '+Yearpp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$G$2:$G$53",
})
Conversion_Registrations.add_series({
    'name':       'Conversion '+Yearp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$K$2:$K$53",
})
Conversion_Registrations.add_series({
    'name':       'Conversion '+Year,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$O$2:$O$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

Dormant_Registrations = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

Dormant_Registrations.add_series({
    'name':       'Dormant '+Yearpp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$H$2:$H$53",
})
Dormant_Registrations.add_series({
    'name':       'Dormant '+Yearp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$L$2:$L$53",
})
Dormant_Registrations.add_series({
    'name':       'Dormant '+Year,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$P$2:$P$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

Total_Registrations = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

Total_Registrations.add_series({
    'name':       'Total '+Yearpp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$I$2:$I$53",
})
Total_Registrations.add_series({
    'name':       'Total '+Yearp,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$M$2:$M$53",
})
Total_Registrations.add_series({
    'name':       'Total '+Year,
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$Q$2:$Q$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

ISvCR_New = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

ISvCR_New.add_series({
    'name':       'CR New',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$R$2:$R$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})
ISvCR_New.add_series({
    'name':       'IS New',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$V$2:$V$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

ISvCR_Conversion = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

ISvCR_Conversion.add_series({
    'name':       'CR Conversion',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$S$2:$S$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})
ISvCR_Conversion.add_series({
    'name':       'IS Conversion',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$W$2:$W$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

ISvCR_Dormant = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

ISvCR_Dormant.add_series({
    'name':       'CR Dormant',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$T$2:$T$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})
ISvCR_Dormant.add_series({
    'name':       'IS Dormant',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$X$2:$X$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

ISvCR_Total = wb.add_chart({'type': 'line',
                               'subtype':'smooth_with_markers'})

ISvCR_Total.add_series({
    'name':       'CR Total',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$U$2:$U$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})
ISvCR_Total.add_series({
    'name':       'IS Total',
    'categories': "'Core Totals Data'!$A$2:$A$53",
    'values':     "'Core Totals Data'!$Y$2:$Y$53",
    'marker': {'type': 'automatic'},
    'smooth':     True
})

New_Registrations.set_title({
    'name':    'New Registrations',
    'overlay': True
})

New_Registrations.set_x_axis({'name':'Tax Week'})
New_Registrations.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA1', New_Registrations, {'x_scale': 3, 'y_scale': 1.2})

Conversion_Registrations.set_title({
    'name':    'Conversion Registrations'
})

Conversion_Registrations.set_x_axis({'name':'Tax Week'})
Conversion_Registrations.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA18', Conversion_Registrations, {'x_scale': 3, 'y_scale': 1.2})

Dormant_Registrations.set_title({
    'name':    'Dormant Registrations'
})

Dormant_Registrations.set_x_axis({'name':'Tax Week'})
Dormant_Registrations.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA35', Dormant_Registrations, {'x_scale': 3, 'y_scale': 1.2})

Total_Registrations.set_title({
    'name':    'Total Registrations'
})

Total_Registrations.set_x_axis({'name':'Tax Week'})
Total_Registrations.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA52', Total_Registrations, {'x_scale': 3, 'y_scale': 1.2})

ISvCR_New.set_title({
    'name':    'IS v CR New'
})

ISvCR_New.set_x_axis({'name':'Tax Week'})
ISvCR_New.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA69', ISvCR_New, {'x_scale': 3, 'y_scale': 1.2})

ISvCR_Conversion.set_title({
    'name':    'IS v CR Conversion'
})

ISvCR_Conversion.set_x_axis({'name':'Tax Week'})
ISvCR_Conversion.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA86', ISvCR_Conversion, {'x_scale': 3, 'y_scale': 1.2})

ISvCR_Dormant.set_title({
    'name':    'IS v CR Dormant'
})

ISvCR_Dormant.set_x_axis({'name':'Tax Week'})
ISvCR_Dormant.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA103', ISvCR_Dormant, {'x_scale': 3, 'y_scale': 1.2})

ISvCR_Total.set_title({
    'name':    'IS v CR Total'
})

ISvCR_Total.set_x_axis({'name':'Tax Week'})
ISvCR_Total.set_y_axis({'name':'Registrations'})

ws.insert_chart('AA120', ISvCR_Total, {'x_scale': 3, 'y_scale': 1.2})

ws.hide_gridlines(2)

ws.hide_row_col_headers()

print('___________________________________________________________________')
print('')
print('Writing Client Margins...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Client Margins')
ws.set_tab_color('#3F43AD')

df_MARGINS_BY_CLIENT = pd.pivot_table(df, values=['Margins'], index=['CRM','Client Name','Type'],
                                      columns=['CHQDATE'], aggfunc=len, fill_value=0)

df_MARGINS_BY_CLIENT.columns = df_MARGINS_BY_CLIENT.columns.droplevel(0)
df_MARGINS_BY_CLIENT = df_MARGINS_BY_CLIENT.reset_index()

df_MARGINS_BY_CLIENT["Total"] = 0

for i, row in df_MARGINS_BY_CLIENT.iterrows():
    for column in df_MARGINS_BY_CLIENT.columns:
        if column not in ['CRM',"Client Name","Type","Total"]:
            df_MARGINS_BY_CLIENT.at[i, "Total"] += row[column]
            
for j, column in enumerate(df_MARGINS_BY_CLIENT.columns.values):
    REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = 1)
    ws.write(REF_1,column,cell_merge3)
        
for i, row in df_MARGINS_BY_CLIENT.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
        ws.write(REF_1,item)
        j += 1

for j, column in enumerate(df_MARGINS_BY_CLIENT.columns.values):
    REF_col = ('{col}{rowstart}:{col}{rowend}').format(col = get_column_letter(j + 1), rowstart = 2,rowend = len(df_MARGINS_BY_CLIENT)+1)
    if column not in ['CRM',"Client Name","Type"]:
        ws.conditional_format(REF_col, {'type': '3_color_scale'})
        
#if isinstance(item, str) == False:

tab_REF = ('A1:{width}{length}').format(length = len(df_MARGINS_BY_CLIENT)+1, width = get_column_letter(3))

ws.add_table(tab_REF, {'name': 'ClientMargins',
                       'columns':[{'header':'CRM'},
                                  {'header':'Client Name'},
                                  {'header':'Solution'}],
                        'style': 'Table Style Dark 2'})

ws.set_column('A:A',20)
ws.set_column('B:B',20)
ws.set_column('C:C',20)

ws.hide_gridlines(2)
ws.freeze_panes(1, 3)

print('___________________________________________________________________')
print('')
print('Writing Consultant Summary...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Consultant Summary')
ws.set_tab_color('#3F43AD')

df_REG_BY_CONSULTANT = pd.pivot_table(df_MRRR, values=['NI Number (Contact)'], index=['Agency','Consultant'],
                                      columns=['Created Time'], aggfunc=len, fill_value=0)

df_REG_BY_CONSULTANT.columns = df_REG_BY_CONSULTANT.columns.droplevel(0)
df_REG_BY_CONSULTANT = df_REG_BY_CONSULTANT.reset_index()

df_REG_BY_CONSULTANT["Total"] = 0

df_REG_BY_CONSULTANT.fillna('',inplace=True)

for i, row in df_REG_BY_CONSULTANT.iterrows():
    for column in df_REG_BY_CONSULTANT.columns:
        if column not in ["Agency","Consultant","Total"]:
            df_REG_BY_CONSULTANT.at[i, "Total"] += row[column]
            
for j, column in enumerate(df_REG_BY_CONSULTANT.columns.values):
    REF_1 = ('{col}{rowstart}').format(col = get_column_letter(j + 1), rowstart = 1)
    ws.write(REF_1,column,cell_merge3)
    
for i, row in df_REG_BY_CONSULTANT.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
        ws.write(REF_1,item)
        j += 1

for j, column in enumerate(df_REG_BY_CONSULTANT.columns.values):
    REF_col = ('{col}{rowstart}:{col}{rowend}').format(col = get_column_letter(j + 1), rowstart = 2,rowend = len(df_REG_BY_CONSULTANT)+1)
    if column not in ['Agency','Consultant']:
        ws.conditional_format(REF_col, {'type': '3_color_scale'})
        
     
tab_REF = ('A1:{width}{length}').format(length = len(df_REG_BY_CONSULTANT)+1, width = get_column_letter(2))

# columns = []
# for i in df_REG_BY_CONSULTANT.columns.tolist():
#     columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'ConstultantSummary',
                       'columns':[{'header':'Agency'},
                                  {'header':'Consultant'}],
                        'style': 'Table Style Dark 2'})

#                        'data':df_REG_BY_CONSULTANT.values.tolist(),
#                       'columns':columns

ws.set_column('A:A',20)
ws.set_column('B:B',20)

ws.freeze_panes(1, 2)

ws.hide_gridlines(2)


print('___________________________________________________________________')
print('')
print('Writing Variance...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Variance')
ws.set_tab_color('#3F43AD')

col_merge = wb.add_format({'border': 2,
                           'bg_color': '#000000',
                           'font_color': '#FFFFFF',
                           'align':'centre',
                           'font_size' : 12,
                           'bold' : True})

format3 = wb.add_format({'bg_color': '#3F43AD',
                         'font_color': '#FFFFFF',
                         'font_size' : 10,
                         'font_name': 'Calibri'})

n = 3
for col_num, value in enumerate(df_VARIANCE.loc[:,df_VARIANCE.columns.get_level_values(1)=='Forecast'].columns.values):
    REF_1 = ('{columnstart}{rowstart}:{columnend}{rowend}').format(columnstart = get_column_letter(n),rowstart = 1,
                                                                   columnend = get_column_letter(n + 2),rowend = 1)  
    ws.merge_range(REF_1,value[0],col_merge)
    n += 3

for col_num, value in enumerate(df_VARIANCE.columns.values):
    REF_2 = ('{columnstart}{rowstart}').format(columnstart = get_column_letter(col_num + 3),rowstart = 2)
    ws.write(REF_2, value[1],format3)

format4 = ({'border': 2,
            'bg_color': '#000000',
            'font_color': '#FFFFFF',
            'align':'centre',
            'font_size' : 12})

ws.write('A1', '',col_merge)
ws.write('A2', 'Client Name')
ws.write('B1', 'CHQDATE',col_merge)
ws.write('B2', 'CRM')

n = 3
for i, row in df_VARIANCE.iterrows():
    REF_1 = ('A{row}').format(row = n)
    REF_2 = ('B{row}').format(row = n)
    ws.write(REF_1, i[0])
    ws.write(REF_2, i[1])
    m = 3
    for j, item in row.iteritems():
       REF = ('{column}{row}').format(row = n, column = get_column_letter(m))
       if isinstance(item, str) == False:
           if np.isnan(item) == True:
               ws.write(REF, None)
           else:
               ws.write(REF, item)
       else:
           ws.write(REF, item)
       m += 1
    ws.set_row(n-1, 40)
    n += 1

for i in range(5, len(df_VARIANCE.columns)+3,3):
    REF_df = ('{columnstart}{rowstart}:{columnend}{rowend}').format(columnstart = get_column_letter(i),
                                                                    rowstart = 3,
                                                                    columnend = get_column_letter(i),
                                                                    rowend = len(df_VARIANCE)+1)
    
    ws.conditional_format(REF_df, {'type': 'cell',
                                    'criteria': '<',
                                    'value': 0,
                                    'format': format1})

    ws.conditional_format(REF_df, {'type': 'cell',
                                    'criteria': '>=',
                                    'value': 0,
                                    'format': format2})
    
for i in range(5, len(df_VARIANCE.columns)+3,3):
    REF_df = ('{columnstart}{rowstart}:{columnend}{rowend}').format(columnstart = get_column_letter(i),
                                                                    rowstart = len(df_VARIANCE)+2,
                                                                    columnend = get_column_letter(i),
                                                                    rowend = len(df_VARIANCE)+2)

    ws.conditional_format(REF_df, {'type': 'cell',
                                    'criteria': '<',
                                    'value': 0,
                                    'format': format1})

    ws.conditional_format(REF_df, {'type': 'cell',
                                    'criteria': '>=',
                                    'value': 0,
                                    'format': format2})


format_hidden = wb.add_format({'font_color': '#FFFFFF'})

n = 3
for i, row in df_VARIANCE_TOTALS.iterrows():
    m = len(df_VARIANCE.columns) + 5
    REF_start = ('{column}{row}').format(row = n, column = get_column_letter(len(df_VARIANCE.columns) + 3))

    for j, item in row.iteritems():
       REF = ('{column}{row}').format(row = n, column = get_column_letter(m))
       if isinstance(item, str) == False:
           if np.isnan(item) == True:
               ws.write(REF, None)
           else:
               ws.write(REF, item,format_hidden)
       else:
           ws.write(REF, item,format_hidden)
       m += 1
    REF_range = ('{columnstart}{rowstart}:{columnend}{rowend}').format(rowstart = n, 
                                                                       columnstart = get_column_letter(len(df_VARIANCE.columns) + 3), 
                                                                       columnend = get_column_letter(m),
                                                                       rowend= n)
    ws.add_sparkline(REF_start,
                {'range':REF_range,
                'type': 'line',
                'negative_points': True,
                'high_point': True,
                'axis': True,
                'high_point': True})
    n += 1

ws.set_column(('{width}:{width}').format(width = get_column_letter(len(df_VARIANCE.columns)+3)), 40)

ws.set_column('A:A',20)
ws.set_column('B:B',20)

ws.add_table(('A2:B{length}').format(length=len(df_VARIANCE)+1), {'name': 'Variance',
                                                                'columns':[{'header':'Client Name'},
                                                                           {'header':'CRM'}],
                                                                'style': 'Table Style Dark 2'})

ws.freeze_panes(2, 2)

print('___________________________________________________________________')
print('')
print('Writing Forecast...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Forecast')
ws.set_tab_color('#3F43AD')

df_FORECAST = df_FORECAST.reset_index()

tab_REF = ('A1:{width}{length}').format(length = len(df_FORECAST)+1, width = get_column_letter(len(df_FORECAST.columns)))

df_FORECAST.fillna('',inplace=True)

columns = []
for i in df_FORECAST.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'Forecast',
                       'data':df_FORECAST.values.tolist(),
                       'columns':columns,
                       'style': 'Table Style Dark 2'})

ws.hide()

print('___________________________________________________________________')
print('')
print('Writing £0 Margins...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Zero Margins Report')
ws.set_tab_color('#3F43AD')

tab_REF = ('A1:{width}{length}').format(length = len(df_0_MARGINS)+1, width = get_column_letter(len(df_0_MARGINS.columns)))

df_0_MARGINS.fillna('',inplace=True)

columns = []
for i in df_0_MARGINS.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'ZeroMarginsReport',
                       'data':df_0_MARGINS.values.tolist(),
                       'columns':columns,
                       'style': 'Table Style Dark 2'})

df_0_MARGINS_ADDED.fillna('',inplace=True)

ws = wb.add_worksheet('Zero Margins Added')
ws.set_tab_color('#3F43AD')

tab_REF = ('A1:{width}{length}').format(length = len(df_0_MARGINS_ADDED)+1, width = get_column_letter(len(df_0_MARGINS_ADDED.columns)))

columns = []
for i in df_0_MARGINS_ADDED.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'ZeroMarginsAdded',
                        'data':df_0_MARGINS_ADDED.values.tolist(),
                        'columns':columns,
                        'style': 'Table Style Dark 2'})

ws.hide()

print('___________________________________________________________________')
print('')
print('Writing Core Margin Data...')
print(datetime.datetime.now().time())


ws = wb.add_worksheet('Core Margin Data')
ws.set_tab_color('#3F43AD')

tab_REF = ('A1:{width}{length}').format(length = len(df)+1, width = get_column_letter(len(df.columns)))

df.fillna('',inplace=True)

columns = []
for i in df.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'CoreMarginData',
                       'data':df.values.tolist(),
                       'columns':columns,
                       'style': 'Table Style Dark 2'})

ws.hide()

print('___________________________________________________________________')
print('')
print('Writing Core Registration Data...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Core Registration Data')
ws.set_tab_color('#3F43AD')

tab_REF = ('A1:{width}{length}').format(length = len(df_MRRR)+1, width = get_column_letter(len(df_MRRR.columns)))

df_MRRR.fillna('',inplace=True)

tab_LIST = df_MRRR.values.tolist()

columns = []
for i in df_MRRR.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'CoreRegistrationData',
                       'data':df_MRRR.values.tolist(),
                       'columns':columns,
                       'style': 'Table Style Dark 2'})

ws.hide()

print('___________________________________________________________________')
print('')
print('Writing Core Totals Data...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Core Totals Data')
ws.set_tab_color('#3F43AD')

df_TOTALS = pd.merge(df_TOTALS, df_REG_TOTALS,how='left')

df_TOTALS.fillna('',inplace=True)

tab_REF = ('A1:{width}{length}').format(length = len(df_TOTALS)+1, width = get_column_letter(len(df_TOTALS.columns)))

columns = []
for i in df_TOTALS.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'CoreTotalsData',
                       'data':df_TOTALS.values.tolist(),
                       'columns': columns,
                       'style': 'Table Style Dark 2'})

ws.hide()

print('___________________________________________________________________')
print('')
print('Writing Rewards Data...')
print(datetime.datetime.now().time())

ws = wb.add_worksheet('Rewards')
ws.set_tab_color('#3F43AD')

df_REWARDS.fillna('',inplace=True)

tab_REF = ('A1:{width}{length}').format(length = len(df_REWARDS)+1, width = get_column_letter(len(df_REWARDS.columns)))

columns = []
for i in df_REWARDS.columns.tolist():
    columns.append({'header':i})

ws.add_table(tab_REF, {'name': 'REWARDS',
                       'data': df_REWARDS.values.tolist(),
                       'columns': columns,
                       'style': 'Table Style Dark 2'})

ws.hide()


print('___________________________________________________________________')
print('')
print('Saving Workbook...')
print(datetime.datetime.now().time())

wb.close()

print('___________________________________________________________________')
print('')
print('Done.')
print(datetime.datetime.now().time())






