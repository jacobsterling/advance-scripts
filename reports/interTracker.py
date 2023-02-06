# -*- coding: utf-8 -*-
"""
Created on Fri Jan 28 14:21:33 2022

@author: jacob.sterling
"""

class intermediaryReportScript():
    def __init__(self):
        from pathlib import Path

        import numpy as np
        import pandas as pd
        import win32com.client as client
        from utils.formats import taxYear
        
        self.pd = pd
        self.np = np
        
        quarter = int(input("Enter Quarter: Q"))
        
        self.Quarter = "Q" + str(quarter)
        
        self.Quarterp = "Q" + str(quarter - 1)
        
        self.outlook = client.Dispatch('Outlook.Application')
        
        self.subject = f'Intermediary Report - {self.Quarter}'
        
        #To answer your question regarding the lack of information on the intermediary report, 
        #all category F contractors are paid on an employment basis. 
        #All of their payment details will be captured in our FPS that we submit each week to HMRC and are therefore not needed on the intermediary report. 
        #All other categories do require this additional information, 
        #however when looking at your report all contractors contained have been paid on an employment basis.

        YearFormat = taxYear().Year('-')
        YearcFormat1 = taxYear().Year_format1('-')
        YearpFormat1 = taxYear().Yearp_format1('-')#change to Year_format1
        
        self.report_prefix = 'IntermediaryReport_083_FA45839_'
        trackerName = f'{YearcFormat1}-{self.Quarter} Inter Report Tracker Python.xlsx'
        
        homePath = Path.home() / "advance.online/J Drive - Operations/Reports/Intermediary Reports"
        self.dataPath = Path.home() / rf"advance.online/J Drive - Exec Reports/Margins Reports/Margins {YearFormat}/Data"
        
        self.filePathP = homePath / rf"{YearcFormat1 if quarter > 1 else YearpFormat1}/{self.Quarterp}"
        self.filePath = homePath / rf"{YearcFormat1}/{self.Quarter}"

        self.reportsPath = self.filePath / "Outstanding Reports"
        self.trackerPath = self.filePath / trackerName
        
        self.c = True if input('Type "y" make Changes to files ?: ') == "y" else False

        self.groupsPath = Path.home() / "OneDrive - advance.online/Documents/Data/Groups.xlsx"
        
        self.missingData = pd.DataFrame([],columns = ['Name','Missing Info'])
    
    def age(self, birthdate):
        from datetime import date
        return int((date.today() - self.pd.to_datetime(birthdate, format="%d/%m/%Y").date()).days / 365)
    
    def nullQuery(self, query, row, k, report, notif = None):
        if self.pd.isnull(row[query]) or row[query] == 'N/A':
            self.missingInfo += 1
            self.details += f'missing {query}, ' if not notif else notif
            if self.c:
                try:
                    if query == 'Worker address line 2':
                        print(rf"Address Line 1: {row['Worker address line 1']}: ")
                        
                    user = input(f"Enter missing {query} for {row['Worker forename']} {row['Worker surname']}: ")
                    if user == '':
                        self.missingData = self.pd.concat([self.missingData, self.pd.DataFrame([[f"{row['Worker forename']} {row['Worker surname']}", query]], columns = self.missingData.columns)]).reset_index(drop=True)
                    else:
                        report.at[k, query] = user
                        self.changes += 1
                except KeyError:
                    user = input(f"Enter {notif} for {report.columns[1]}: ")
                    if user != '':
                        report.at[k, query] = user
                        self.changes += 1
                    else:
                        self.missingInfo += 1
                    self.missingInfo -= 1
        return report
                
    def rangeQuery(self, startCol, endCol, report_row, k, values, report, notif):
        if (report.loc[k,startCol:endCol].values != values).all():
            self.missingInfo += 1
            self.details += f'missing {notif}, '
            if self.c:
                report.loc[k,startCol:endCol] = values
                self.changes += 1
                print('')
                print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
                print(f"Changed {notif} for {report_row['Worker forename']} {report_row['Worker surname']}.")
                print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
        return report
    
    def matchingQuery(self, condition1, condition2, notif = None):
        if condition1 != condition2:
            self.details += f'{condition1} does not match {condition2}, '
            return False
        return True
    
    def readReport(self,file):
        try:
            info = self.pd.read_csv(file,nrows=5, on_bad_lines='skip').iloc[:,:2]
            info[info.columns[1]] = info[info.columns[1]].astype(str)
            info.loc[info[info.columns[1]] == 'nan',info.columns[1]] = self.np.nan
            return info, self.pd.read_csv(file,skiprows=7, on_bad_lines='skip')
        except UnicodeDecodeError:
            info = self.pd.read_csv(file,nrows=5,encoding='latin', on_bad_lines='skip').iloc[:,:2]
            info[info.columns[1]] = info[info.columns[1]].astype(str)
            info.loc[info[info.columns[1]] == 'nan',info.columns[1]] = self.np.nan
            return info, self.pd.read_csv(file,skiprows=7,encoding='latin', on_bad_lines='skip')

    def queryReport(self, info, report):
        iter_info = info
        for j, info_row in iter_info.iterrows():
            if j == 0:
                info = self.nullQuery(info.columns[1], info_row, j, info,'Missing Address, ')
        info = self.nullQuery(info.columns[1], info_row, j, info,'Missing Post Code, ')

        iter_report = report

        for k, report_row in iter_report.iterrows():
            
            report = self.nullQuery('Worker date of birth', report_row, k, report)
            
            report = self.nullQuery('Worker gender', report_row, k, report)
            
            report = self.nullQuery('Worker National Insurance number', report_row, k, report)

            report = self.nullQuery('Start date of engagement',report_row, k, report)
            
            report = self.nullQuery('Worker address line 2',report_row, k, report)
            
            report = self.nullQuery('Worker date of birth', report_row, k, report)
            
            report = self.nullQuery('Worker postcode', report_row, k, report)
            
            
            if self.age(report_row['Worker date of birth']) < 16 or self.age(report_row['Worker date of birth']) > 75 :
                self.missingInfo += 1
                self.details += 'Potentially Incorrect Date Of Birth, '
            
            if self.pd.to_datetime(report_row['End date of engagement'],format="%d/%m/%Y") == self.pd.to_datetime("19/04/2022" ,format="%d/%m/%Y"):
                self.details += 'Potentially Incorrect End Date, '
                if self.c:
                    endDate = input("Enter new End date of engagament (in format dd/mm/YYYY): ")
                    if endDate != "":
                        report.at[k, 'End date of engagement'] = self.pd.to_datetime(endDate, format="%d/%m/%Y")
                        self.changes += 1
                    else:
                        self.missingInfo += 1
                else:
                    self.missingInfo += 1
                    
            if report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'D' or report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'B':
                
                report = self.rangeQuery("Name of party paid by intermediary for worker's services",
                                "Postcode of party paid by intermediary for worker's services",
                                report_row, k,
                                ["Advance Contracting Solutions Ltd","First Floor VISTA","St David's Park","Ewloe","Chester","CH5 3DT"],
                                report,
                                'Intermediary Details, ')

                report = self.nullQuery("Amount paid for the worker's services" , report_row, k, report)

                report = self.nullQuery("Companies House registration number of party paid by intermediary for worker's services", report_row, k, report)
                    
            if report_row["Worker engagement details where intermediary didn't operate PAYE"] == 'A':
                
                report = self.rangeQuery("Name of party paid by intermediary for worker's services",
                                "Postcode of party paid by intermediary for worker's services",
                                report_row, k, 
                                ["Advance Contracting Solutions Ltd","First Floor VISTA","St David's Park","Ewloe","Chester","CH5 3DT"],
                                report,
                                'Intermediary Details, ')
                        
                report = self.nullQuery("Amount paid for the worker's services", report_row, k, report)

                report = self.nullQuery("Worker unique taxpayer reference (UTR)", report_row, k, report)
        
        numberOfErrors = len(report[report["Amount paid for the worker's services"].astype(float) == 0])
        if numberOfErrors > 0:
            self.missingInfo += numberOfErrors
            self.details += "Amount paid for the worker's services <= 0"
            if self.c:
                if input(rf'Amount paid for the workers services <= 0, Remove {numberOfErrors} rows ? ("y"): ') == 'y':
                    report = report[report["Amount paid for the worker's services"].astype(float) != 0]
                    self.changes += 1
        
        numberOfErrors = len(report[ report["Worker forename"].apply(lambda x: x.upper().__contains__(" DNU"))])
        if numberOfErrors > 0:
            self.missingInfo += numberOfErrors
            self.details += rf"Contains {numberOfErrors} DNU Worker(s), "
            if self.c:
                if input(rf'Report contains DNU Worker(s), Remove {numberOfErrors} rows ? ("y"): ') == 'y':
                    report = report[ ~report["Worker forename"].apply(lambda x: x.upper().__contains__(" DNU")) ]
                    self.changes += 1
        return info, report
    
    def createTracker(self):
        from utils.functions import tax_calcs
        
        Weekc = tax_calcs().tax_week() - 1
        
        clients = self.pd.concat([self.pd.read_csv(self.dataPath / rf"Week {Weekc}/clients io.csv", encoding = 'latin',
                                            usecols = ['Company Name                   ','OFFNO','EMAIL_DETAILS_INTER']),
                                self.pd.read_csv(self.dataPath / rf"Week {Weekc}/clients axm.csv",encoding = 'latin',
                                            usecols = ['Company Name                   ','OFFNO','EMAIL_DETAILS_INTER'])])
        
        clients.columns = ['Client Name','OFFNO','Email']
        clients['Client Name'] = clients['Client Name'].str.upper()
        clients.sort_values("Client Name", inplace = True)
        clients.drop_duplicates(subset ="Client Name",
                             keep = "last", inplace = True)
        
        accounts = self.pd.read_csv(self.dataPath / rf"Week {Weekc}/Accounts+Office.csv", na_values = '-', skiprows=6)
        accounts = accounts[['Office Number','Account Owner','Send Intermediary Report to','Account Type']]
        accounts = accounts.dropna(subset=['Office Number'])
        accounts['Office Number'] = accounts['Office Number'].astype(int)
        accounts = accounts.drop_duplicates(subset=['Office Number'], keep='first')
        
        groups = self.pd.read_excel(self.groupsPath)
        groups['Client Name'] = groups['Client Name'].str.upper()
        groups['Name Change'] = groups['Name Change'].str.upper()
        
        for file in self.filePathP.glob('*.xlsx'):
            if file.name.__contains__('Inter Report Tracker Python'):
                trackerP = self.pd.read_excel(file, usecols=['OFFNO','Merit Email', 'CRM Email'])
                break

        self.df = self.pd.DataFrame([],columns=['File','OFFNO','Client','Account','Account Type','Merit Email','CRM Email','Changes Made','Missing Info','A','F','B','D', 'Details'])
        for file in self.reportsPath.glob('*'):
            if file.is_file():
                if file.suffix == ".csv" and self.report_prefix in file.name: 
                    offno = int(file.name.split('_')[3])
                    
                    print('___________________________________________________________________')
                    print('')
                    print(f'Reading {file.name}')
                    
                    info, report = self.readReport(file)
                    
                    self.changes, self.missingInfo, self.details  = 0, 0, ''
                    
                    try:
                        clientName = clients.loc[clients['OFFNO']==offno,'Client Name'].values[0]
                        clientEmail = clients.loc[clients['OFFNO']==offno,'Email'].values[0]
                    except IndexError:
                        clientName = None
                        clientEmail = None
                        self.missingInfo += 2
                        print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
                        print(f'{offno} is missing from clients.csv')
                        self.details += f'{offno} is missing from clients.csv, '
                        print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
                        
                    try:
                        accountOwner = accounts.loc[accounts['Office Number']==offno,'Account Owner'].values[0]
                        accountType = accounts.loc[accounts['Office Number']==offno,'Account Type'].values[0]
                        crmEmail = accounts.loc[accounts['Office Number']==offno,'Send Intermediary Report to'].values[0]
                    except IndexError:
                        accountOwner = None
                        accountType = None
                        crmEmail = None
                
                    self.nullVar('account owner',accountOwner, 1)
                    self.nullVar('account type',accountType, 1)
                    self.nullVar('client email',clientEmail)
                    if self.nullVar('crm email',crmEmail, 1):
                        try:
                            crmEmail = trackerP.loc[trackerP['OFFNO']==offno,'CRM Email'].values[0]
                            if crmEmail == 0:
                                crmEmail = None 
                                self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
                            else:
                                self.details += f'Used email from previous {self.Quarterp} tracker, '
                        except IndexError:
                            try:
                                crmEmail = trackerP.loc[trackerP['OFFNO']==offno,'Merit Email'].values[0]
                                if crmEmail == 0:
                                    crmEmail = self.np.nan 
                                    self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
                                else:
                                    self.details += f'Used merit email from previous {self.Quarterp} tracker, '
                            except IndexError:
                                self.missingInfo += 1
                                self.details += f'Cannot find email in previous {self.Quarterp} tracker, '
                        
                    elif clientEmail != crmEmail and not self.pd.isnull(clientEmail):
                        self.missingInfo += 1
                        self.details += 'Merit Email does not match CRM, '
                    
                    for i, row in groups.iterrows():
                        if not self.pd.isnull(clientName):
                            if row['Client Name'] == clientName or (clientName.__contains__(row['Name Change']) and len(row['Name Change']) > 3):
                                clientName = row['Name Change']
                                accountType = 'Group'
                                self.missingInfo += 1
                                self.details += 'Needs to be Grouped, '
                                break

                    if self.pd.isnull(clientName):
                        clientName = info.columns[1].upper()
                        self.details += 'Client name used from file, '
                        
                    if not self.matchingQuery(info.columns[1].upper(), clientName):
                        if self.c and input(f"Miss Matching Client Name {info.columns[1].upper()}, type 'y' to change to {clientName}: ") == 'y':
                            self.changes += 1
                            info.columns = ['Employment intermediary name',clientName.title()]
                    
                    print('')
                    print(f'Showing Client - {info.columns[1].upper()}')
                    print('')
                    
                    info, report = self.queryReport(info, report)
                                
                    self.saveChanges(file, info, report)
                    
                    if accountType != 'Agency':
                        self.missingInfo += 1
                        self.details += 'Not agency, '

                    df_temp = self.pd.DataFrame([[file.name,
                                             offno,
                                             clientName,
                                             accountOwner,
                                             accountType,
                                             clientEmail,
                                             crmEmail,
                                             self.changes,
                                             self.missingInfo,
                                             len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'A']),
                                             len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'F']),
                                             len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'B']),
                                             len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'D']),
                                             self.details]],
                                             columns=['File','OFFNO','Client','Account','Account Type','Merit Email','CRM Email','Changes Made','Missing Info','A','F','B','D','Details'])
                    
                    self.df = self.df.append(df_temp)
                    
        self.exportTracker()
        if input('Type "y" to edit the tracker ?: ') == 'y':
            self.editTracker()
        else:
            return self.df
    
    def nullVar(self, name, var, i = 0):
        if self.pd.isnull(var):
            self.missingInfo += i
            self.details += f'Missing {name}, '
            return True
        else:
            return False
        
    def updateReport(self, i, clientName, accountOwner, accountType, clientEmail, crmEmail, report):
        
        try:
            A = len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'A'])
            F = len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'F'])
            B = len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'B'])
            D = len(report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'D'])  
        except TypeError:
            A = 0
            F = 0
            B = 0
            D = 0
        
        self.df.at[i, 'Client'] = clientName
        self.df.at[i, 'Account'] = accountOwner
        self.df.at[i, 'Account Type'] = accountType
        self.df.at[i, 'Merit Email'] = clientEmail
        self.df.at[i, 'CRM Email'] = crmEmail
        self.df.at[i, 'Changes Made'] = self.changes
        self.df.at[i, 'Missing Info'] = self.missingInfo
        self.df.at[i, 'A'] = A
        self.df.at[i, 'F'] = F
        self.df.at[i, 'B'] = B
        self.df.at[i, 'D'] = D
        self.df.at[i, 'Details'] = self.details
        
        
    def editTracker(self):
        try:
            self.send = True if input('Type "y" to setup emails ?: ') == "y" else False
            checkEmail = input("Enter Check Email: ") if input("Send Checks ?: ") == "y" else None
            checkItems = True if input("Check inbox for sent emails ?: ") == "y" else False
            
            sentItems = []
            
            if checkItems:
                namespace = self.outlook.GetNamespace("MAPI")
        
                Items = namespace.Folders["jacob.sterling@advance.online"].Folders["Sent Items"].Items
        
                for i in range(Items.Count -1, 0, -1):
                    if Items[i].Subject == self.subject:
                        sentItems.append(Items[i].To)
            
            self.df = self.pd.read_excel(self.trackerPath)
            for i, row in self.df.iterrows():
                clientName = self.df.at[i, 'Client']
                accountOwner = self.df.at[i,'Account']
                accountType = self.df.at[i,'Account Type']
                clientEmail = self.df.at[i,'Merit Email']
                crmEmail = self.df.at[i,'CRM Email']
                self.details = '' # if self.pd.isnull(row['Details']) else row['Details']
                self.missingInfo = self.df.at[i,'Missing Info']
                
                self.changes =row['Changes Made']
                
                if checkItems:
                    if crmEmail in sentItems:
                        self.changes = 'Sent'
                
                if self.changes not in ['Sent']: #,'Send'
                    file = self.reportsPath / f"{row['File']}"
                    
                    print('___________________________________________________________________')
                    print('')
                    print(f'Reading {file.name}')
                
                    info, report = self.readReport(file)
                    self.missingInfo = 0
                    
                    manualSend = True if self.changes == 'Send' else False
                    check = True if self.changes == 'Check' and checkEmail else False
                    self.changes = 0 if self.changes in ['Send', 'Check'] else self.changes
                    
                    if accountOwner == 'OFFNO not found in accounts office':
                        self.missingInfo += 1
                        self.details += 'Missing Account Owner, '
                    
                    self.nullVar('account type', accountType, 1)
                    
                    if accountType in ['Group', 'End Client'] and self.c:
                        if clientName == "MASTER PEACE RECRUITMENT":
                            try:
                                from utils.functions import tax_calcs
                                
                                tax_week_map = tax_calcs().tax_week_map()
                                quartermap = tax_week_map[tax_week_map["Quarter"].str.contains(self.Quarter)]
                                emp = self.pd.DataFrame([], columns=["Payno", "Agency","SW_SEX"])
                                
                                # report = self.pd.DataFrame([], columns=["Employment intermediary name", "Master Peace Recruitment"])
                                # report.at[0, "Employment intermediary name"] = "Employment intermediary address line 1"
                                # report.at[1, "Employment intermediary name"] = "Employment intermediary address line 2"
                                # report.at[2, "Employment intermediary name"] = "Employment intermediary address line 3"
                                # report.at[3, "Employment intermediary name"] = "Employment intermediary address line 4"
                                # report.at[4, "Employment intermediary name"] = "Employment intermediary postcode"
                                
                                for i, row in quartermap.iterrows():
                                    print(rf"Reading Emp: {row['Week']}")
                                    temp = self.pd.read_csv(self.dataPath / rf"Week {row['Week']}/emp paid by week.csv", usecols = emp.columns)
                                    emp = self.pd.concat([ emp, temp ])
                                
                                joinersPaye = self.pd.read_csv(self.dataPath / rf"Week {row['Week']}/Joiners Error Report paye.csv", usecols=["Pay No","Forenames","Surname","Date of Birth","DOJ","NI_NO","ADD1","ADD2","ADD3","ADD4","POST_CODE"]).rename(columns = {"Forenames":"Worker forename","Surname":"Worker surname","Date of Birth":"Worker date of birth","DOJ":"Start date of engagement","NI_NO":"Worker National Insurance number","ADD1":"Worker address line 1","ADD2":"Worker address line 2","ADD3":"Worker address line 3","ADD4":"Worker address line 4","POST_CODE":"Worker postcode"})
                                
                                emp = emp[~emp["Payno"].isna()].drop_duplicates(subset="Payno").reset_index(drop=True)
                                emp["Pay No"] = emp["Payno"].apply(lambda x: x.split("*")[1])
                                emp = emp.merge(joinersPaye, how = "left")
                                emp = emp[emp["Agency"] == "MASTER PEACE"].drop(columns = ["Agency", "Payno", "Pay No"]).rename(columns={"SW_SEX":"Worker gender"})
                                emp["Worker engagement details where intermediary didn't operate PAYE"] = "F"
                                
                                report = self.pd.concat([report, emp]).drop_duplicates(subset=["Worker National Insurance number", "Worker engagement details where intermediary didn't operate PAYE"]).reset_index(drop=True)
                                
                                self.changes += 1
                                accountType = 'Grouped'
                            except FileNotFoundError:
                                print("Failed to create MASTERPEICE Report")
                        else:
                            group = self.df.loc[self.df['Client'] == clientName] if accountType == 'Group' else self.df.loc[self.df['Account Type'] == 'End Client']
                            if len(group) > 1:
                                if input(f"Type 'y' to group {clientName} with groupee's: ") == 'y':
                                    changes = self.changes
                                    for j, groupee in group.iterrows():
                                        gPath = self.filePath / f"Outstanding Reports/{groupee['File']}"
                                        if gPath != file:
                                            g_info, g_report = self.readReport(gPath)
                                            self.changes = groupee['Changes Made']

                                            self.changes = 0 if self.changes == 'Send' else self.changes
                                            
                                            print('')
                                            print(f'Showing {accountType} Groupee - {g_info.columns[1].upper()}')
                                            print('')
                                            
                                            g_info, g_report = self.queryReport(g_info, g_report)
                                            
                                            self.df.at[j, 'Account Type'] = 'Groupee' if accountType == 'Group' else 'End Client Groupee'
                                            
                                            self.saveChanges(file, g_info, g_report)
        
                                            report = self.pd.concat([report,g_report])
                                            
                                    self.changes = changes + 1
                                    if accountType == 'Group':
                                        accountType = 'Grouped'
                                    else:
                                        accountType = 'End Client Grouped'
                                        clientName = rf'Advance Contracting Solutions - Formally {clientName}'
                                        info.columns = ["Employment intermediary name","Advance Contracting Solutions"]
                                        info.at[0, "Advance Contracting Solutions"] = "Ground Floor"
                                        info.at[1, "Advance Contracting Solutions"] = "VISTA"
                                        info.at[2, "Advance Contracting Solutions"] = "St David’s Park"
                                        info.at[3, "Advance Contracting Solutions"] = "Ewloe"
                                        info.at[4, "Advance Contracting Solutions"] = "CH5 3DT"
                                        report = report[report["Worker engagement details where intermediary didn't operate PAYE"] == 'A']
                                        
                                    self.details += 'Grouped, '
                                elif input("Type 'y' if already grouped: ") == 'y':
                                    accountType = 'Grouped'
                                    self.changes += 1
                    
                    if accountType not in ["Agency", 'Grouped']:
                        if accountType == "Other" and self.send:
                            if input("Type 'y' to treat as agency ?: ") != "y":
                                self.missingInfo += 1
                                self.details += 'Not Agency, '
                            else:
                                self.missingInfo += 1
                                self.details += 'Not Agency, '
                    
                    if accountType != 'Groupee':
                        if not self.matchingQuery(info.columns[1].upper(), clientName):
                            if self.c and input(f"Miss Matching Client Name {info.columns[1].upper()}, type 'y' to change to {clientName}: ") == 'y':
                                self.changes += 1
                                info.columns = ['Employment intermediary name',clientName.title()]
                        
                        print('')
                        print(f'Showing {accountType} - {info.columns[1].upper()}')
                        print('')
                            
                        info, report = self.queryReport(info, report)
                                    
                        self.saveChanges(file, info, report)
                        
                        if self.send and accountType in ["Agency", 'Grouped', "Other"] and self.changes != 'Send' and self.changes > 0:
                            print('')
                            print(f'Details - {self.details}')
                            print('')
                            self.changes = 'Send' if input(f'Type "y" to allow file with {self.changes} changes to be sent to agency (does not send): ') == 'y' else self.changes
                        
                        if self.nullVar('crm email',crmEmail, 1) and not self.nullVar('client email',clientEmail) and input(f'Type "y" to update {clientName} crm email {crmEmail} with {clientEmail}: ') == 'y':
                            crmEmail = clientEmail
                            self.missingInfo -= 1
                            self.details = self.details.replace('Missing CRM Email, ','')
                            
                        self.updateReport(i, clientName, accountOwner, accountType, clientEmail, crmEmail, report)

                        if check:
                            self.emailReport(checkEmail, file)
                            self.df.at[i,'Changes Made'] = "Check"
                        elif self.send and (manualSend or self.changes == 'Send') and accountType in ["Agency", 'Grouped', "Other"] and not self.pd.isnull(crmEmail):
                            print('')
                            print(f"A: {self.df.at[i,'A']}, F: {self.df.at[i,'F']}, B: {self.df.at[i,'B']}, D: {self.df.at[i,'D']}")
                            print('')
                            if self.missingInfo > 0:
                                s = True if input('Type "y" to if you want to compose email despite missing information: ') == 'y' else False
                            elif input(f'Type "y" to compose email to {clientName}: ') == 'y':
                                s = True
                            else:
                                s = False
                            if s:
                                self.emailReport(crmEmail,file)
                                self.df.at[i,'Changes Made'] = 'Sent'
                        elif self.changes == 'Sent':
                            print('')
                            print('Sent')
                else:
                    self.updateReport(i, clientName, accountOwner, accountType, clientEmail, crmEmail, None)
        except KeyboardInterrupt: 
            pass
        
        self.exportTracker()
        return self.df
    
    def saveChanges(self, file, info, report):
        if self.c and self.changes > 0:
            if input('Type "y" to remove duplicates on NI number: ') == 'y':
                report.drop_duplicates(subset=['Worker National Insurance number',"Worker engagement details where intermediary didn't operate PAYE"],keep = "first", inplace = True)
                
            if input(f'Type "y" to save {self.changes} changes to {info.columns[1].upper()} - {file.name} with {self.missingInfo} missing info: ') == 'y':
                
                reportCol = self.pd.DataFrame([report.columns])
                
                report.columns = range(0, len(report.columns))
                
                emptyRow = self.pd.DataFrame([[self.np.nan for i in report.columns]])
                
                infoCol = self.pd.DataFrame([info.columns])
                
                for i in range(len(info.columns), len(report.columns)):
                    info[i] = self.np.nan
                
                info.columns = report.columns
                
                df = self.pd.concat([infoCol, info, emptyRow, reportCol, report]).reset_index(drop=True)
                
                df.to_csv(file, index=False, header=False)
                
                info = info.iloc[:,:2]
                
                info.columns = infoCol.iloc[0]
                
                report.columns = reportCol.iloc[0]
            
        else:
            self.changes = 0
            if self.missingInfo == 0:
                self.changes = 'Send'

    def emailReport(self, crmEmail, report):
        email = self.outlook.CreateItem(0)
        email.To = crmEmail
        email.Subject = self.subject

        email.SentOnBehalfOfName = 'info@advance.online'
        #info
        html = rf"""
            Hi,
            <br>
            <br>
            Please find attached your intermediary report for {self.Quarter}.
            <br>
            <br>
            Please remember to report on any limited companies you may have contracted with in {self.Quarter}. 
            <br>
            <br>
            I have also attached a PDF document to that you may find useful to review on intermediary reporting.
            <br>
            <br>
            <u><i>Assumptions</i></u>
            <br>
            <br>
            Start date is the date we set the contractor up
            <br>
            <br>
            End date is the day the contractor informed us they are no longer working with ADVANCE
            <br>
            <br>
            If you have any questions please let me know. 
            <br>
            <br>
            <b>Kind regards</b>,
            <br>
            <br>
            <font color='#3F43AD'>
            <b>ADVANCE</b>
            <br>
            <br>
            Office: 01244 564 564
            <br>
            Email: info@advance.online
            <br>
            Visit: <a href='https://www.advance.online/'>www.advance.online</a>
            <br>
            <br>
            <em>Service is important to us, and we value your feedback. Please tell us how we did today by clicking <a href='https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab\#lrd=0x487ac350478ae187:0xc129ea0eaf860aee,3,,,https://www.google.com/search?rlz=1C1GCEA_enGB894GB894&ei=jRPJX56yDr2p1fAP0piSqAw&q=advance+contracting&gs_ssp=eJzj4tFP1zfMSDYsNsyxLDdgtFI1qDCxME9MNjY1MDG3SEw1tDC3MqhINjSyTE00SE1MszAzSExN9RJOTClLzEtOVUjOzyspSkwuycxLBwARHBbA&oq=advance+contract&gs_lcp=CgZwc3ktYWIQAxgAMgsILhDHARCvARCTAjIICC4QxwEQrwEyCAguEMcBEK8BMgIIADIICC4QxwEQrwEyBQgAEMkDMggILhDHARCvATICCAAyAggAMgIIADoKCAAQsQMQgwEQQzoOCC4QsQMQgwEQxwEQowI6CAgAELEDEIMBOgQIABBDOgUIABCxAzoLCC4QsQMQxwEQowI6CAguEMcBEKMCOgoILhDHARCjAhBDOg0ILhDHARCjAhBDEJMCOgsILhDHARCvARCRAjoQCC4QsQMQxwEQowIQQxCTAjoFCAAQkQI6BwgAELEDEEM6CwguELEDEMcBEK8BOg0ILhCxAxDHARCjAhBDOgoILhDHARCvARAKOgQIABAKOgIILlCFFFjCOGDAQGgCcAF4AIABxQGIAZIYkgEEMS4xN5gBAKABAaoBB2d3cy13aXrAAQE&sclient=psy-ab'>here</a> to leave us a review on Google</em>
            </font>
            <br>
            <br>
            <img src="{str(self.filePath / 'signature.png')}">
            <br>
            <br>
            <font style='color:#9C9C9C' face = 'Ariel' size='0.5'>
            The information in this email is confidential, privileged and protected by copyright. It is intended solely for the addressee. If you are not the intended recipient any disclosure, copying or distribution of this email is prohibited and may not be lawful. If you have received this transmission in error, please notify the sender by replying by email and deleting the email from all of your computer systems. The transmission of this email cannot be guaranteed to be secure or unaffected by a virus. The contents may be corrupted, delayed, intercepted or lost in transmission, and Advance cannot accept any liability for errors, omissions or consequences which may arise.
            <br>
            Registered Office; Ground Floor, VISTA, St David's Park, Ewloe, CH5 3DT. Registered in England and Wales.
            </font>
        """
        email.HTMLBody = html
        email.Attachments.Add(Source=str(report))
        email.Attachments.Add(Source=str(self.filePath / "Onshore employment intermediaries - FAQs.pdf"))
        email.Display(True)
        
    def exportTracker(self):
        import xlsxwriter
        from openpyxl.utils import get_column_letter
        
        df = self.df.fillna('').reset_index(drop=True)
        
        wb = xlsxwriter.Workbook(self.trackerPath)
        
        format1 = wb.add_format({'bg_color': '#FFC7CE',
                                       'font_color': '#9C0006'})
        
        format2 = wb.add_format({'bg_color': '#C6EFCE',
                                       'font_color': '#006100'})
        
        cell_format_column = wb.add_format({'font_size' : 16,
                                            'align': 'center',
                                            'bg_color': '#FFFF00',
                                            'border':1})
        
        ws = wb.add_worksheet('Tracker')
        
        for j, column in enumerate(df.columns.values):
            col = get_column_letter(j + 1)
            row = 1
            rowend = len(df)+1
            ws.write(f'{col}{row}',column,cell_format_column)
            ws.set_column(f'{col}:{col}', 15)
            if column == 'Missing Info':
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': '>',
                                              'value': 0,
                                              'format': format1})
                
        for i, row in df.iterrows():
            j = 0
            for item in row:
                REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
                if row.index[j] == 'File':
                    #ws.write_url(REF_1, str(self.filePath / "Outstanding Reports") + '\\' + item,string=item)
                    ws.write_url(REF_1, "Outstanding Reports/" + item, string=item)
                elif row.index[j] == 'Changes Made':
                    if item == 'Sent':
                        ws.write(REF_1,item,format2)
                    else:
                        ws.write(REF_1,item)
                        
                elif row.index[j] == 'Account' or row.index[j] == 'Account Type':
                    if item == '':
                        ws.write(REF_1,item,format1)
                    else:
                        ws.write(REF_1,item)
                else:
                    ws.write(REF_1,item)
                j += 1
        wb.close()
        
        if self.c:
            self.missingData.to_csv("missing information.csv", index = False)
            
#intermediaryReportScript().createTracker()

intermediaryReportScript().editTracker()
