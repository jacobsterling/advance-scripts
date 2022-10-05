# -*- coding: utf-8 -*-
"""
Created on Fri Jan 14 15:28:59 2022

@author: jacob.sterling
"""

import pandas as pd
import datetime
import win32com.client as win32
import xlsxwriter
win32c = win32.constants
from openpyxl.utils import get_column_letter
from pathlib import Path
from utils.formats import taxYear
from utils.functions import tax_calcs

year = taxYear().Year('-')
yearAbbr = taxYear().Year_format1("-")
yA, _ = yearAbbr.split("-")

monthNum = int(input(rf"Enter Month Number ({year}): "))
   
datetime_object = datetime.datetime.strptime(str(monthNum), "%m")
month = datetime_object.strftime("%B")

homePath = Path.home() / "advance.online"
rebatesPath = homePath / rf"J Drive - Finance/Rebates Reports/Rebates {yearAbbr}/{month} {yA}"

rebates = pd.read_excel(rebatesPath / rf"{month} py Rebates {yearAbbr}.xlsx", sheet_name= 'Core Data',usecols = ['Client Name','Surname','Forename','Margins','CHQDATE',"Email"]).rename(columns={'CHQDATE':'Date Paid','Margins':'Margin'})
rebates = rebates[(rebates['Margin'] > 0) & (rebates['Client Name'] == 'SCANTEC PERSONNEL LIMITED')].drop('Client Name',axis=1)

workers = pd.read_csv('Scantec+Workers.csv', encoding = 'latin', na_values=["-"], usecols=['Consultant',"Email"], skiprows=6).rename(columns={'Consultant':'Consultant Name in CRM'}).dropna()

consultants = pd.read_excel('scantecConsultants.xlsx')

rebates = rebates.merge(workers, how="left").merge(consultants, left_on = 'Consultant Name in CRM', right_on = 'Name', how = 'left').drop(columns=['Name','Email']).rename(columns={'Consultant_Code':'Reference'})

missingConsultant = rebates[rebates['Consultant Name in CRM'].isna()].drop_duplicates(subset=['Surname','Forename'])

missingCode = rebates[(rebates['Reference'].isnull()) & (~rebates['Consultant Name in CRM'].isna())][['Consultant Name in CRM']].drop_duplicates()
    
if len(missingConsultant) > 0:
    knownMissing = pd.read_csv("missingConsultants.csv")
    
    for i, row in missingConsultant.iterrows():
        consultant = knownMissing.loc[(knownMissing['Surname'] == row["Surname"]) & (knownMissing['Forename'] == row["Forename"]), 'Consultant Name in CRM']

        if len(consultant) > 0:
            consultant = consultant.values[0]
        else:
            consultant = input("Enter consultant for worker {row['Forename'] + ' ' + row['Surname']}: ")
            row['Consultant Name in CRM'] = consultant
            pd.concat([knownMissing, row])
        
        if not pd.isnull(consultant): 
            code = consultants.loc[consultants['Name'] == consultant, 'Consultant_Code'].values[0]
            rebates.loc[(rebates['Surname'] == row["Surname"]) & (rebates['Forename'] == row["Forename"]), 'Consultant Name in CRM'] = consultant
            rebates.loc[(rebates['Surname'] == row["Surname"]) & (rebates['Forename'] == row["Forename"]), 'Reference'] = code
        
    knownMissing.to_csv("missingConsultants.csv", index = False)
    
if len(missingCode) > 0:
    for i, row in missingCode.iterrows():
        code = input(rf"Enter missing code for {row['Consultant Name in CRM']}:")

        consultants = pd.concat([consultants, [row['Consultant Name in CRM'], code]])

        rebates.loc[rebates['Consultant Name in CRM'] == row['Consultant Name in CRM'], 'Reference'] = code
        
    consultants.to_excel('scantecConsultants.xlsx', index = False)
            
wb = xlsxwriter.Workbook(rebatesPath / rf"Scantec {month} {yA}.xlsx")

cellValue = wb.add_format({'font_size' : 16,
                                'align': 'center',
                                'border':1})

currencyFormat = wb.add_format({'font_size' : 16,
                                'align': 'center',
                                'num_format': '£#,##0.00',
                                'border':1})

currencyTotals = wb.add_format({'font_size' : 16,
                                'align': 'center',
                                'num_format': '£#,##0.00',
                                'bg_color': '#FFFF00',
                                'border':1})

cellHeader = wb.add_format({'bold':     True,
                                'font_size' : 16,
                                'align': 'center',
                                'font_color': '#FF0000'})

cellTotals = wb.add_format({'font_size' : 16,
                                'align': 'center',
                                'bg_color': '#FFFF00',
                                'border':1})

cellColumn = wb.add_format({'font_size' : 16,
                                'align': 'center',
                                'bg_color': '#92D050',
                                'border':1})

chqdates = rebates['Date Paid'].unique()

totals = pd.DataFrame([],columns=['','£4.75','£3'],index=pd.DataFrame(chqdates)[0].apply(lambda x: tax_calcs().tax_week(x)).sort_values())
print(totals)

for chqdate in chqdates:
    chqdate = pd.to_datetime(chqdate)
    date = chqdate.strftime('%d.%m')
    week = tax_calcs().tax_week(chqdate)
    ws = wb.add_worksheet(f'Week {week} {date}')

    df = rebates[rebates['Date Paid'] == chqdate].reset_index(drop = True)
    df['Margin'] = 1
    df['Date Paid'] = chqdate.strftime('%d/%m/%Y')

    totals.at[week,''] = len(df)
    totals.at[week,'£4.75'] = len(df)*4.75
    totals.at[week,'£3'] = 0

    tab_REF = ('A1:{width}{length}').format(length = len(df)+1, width = get_column_letter(len(df.columns)))
    df.fillna('',inplace=True)

    for j, column in enumerate(df.columns.values):
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = 1)
        ws.write(REF_1,column, cellHeader)
        ws.set_column(('{col}:{col}').format(col = get_column_letter(j + 1)), 15)

    for k, row in df.iterrows():
        j = 0
        for item in row:
            REF_1 = ('{col}{rowf}').format(col = get_column_letter(j + 1), rowf = k + 2)
            ws.write(REF_1,item)
            j += 1

ws = wb.add_worksheet('Totals')

totals = totals.dropna(axis=0).reset_index()

ws.write('B5','Week', cellHeader)
ws.write('C5','Total numbers paid', cellHeader)
ws.write('D6','UMB/CIS', cellValue)
ws.write('E6','UMB/CIS', cellValue)

ws.write('C14','Subtotals', cellTotals)
ws.write('D14',sum(totals['£4.75']), currencyTotals)
ws.write('E14',sum(totals['£3']), currencyTotals)
ws.write('C17','Total', cellTotals)
ws.write('D17',sum(totals['£4.75'])+sum(totals['£3']), currencyTotals)

ws.write('D7','£4.75',cellColumn)
ws.write('E7','£3.00',cellColumn)
ws.write('C7','',cellValue)

for i, row in totals.iterrows():
    j = 1
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 8)
        if get_column_letter(j + 1) == 'C' or get_column_letter(j + 1) == 'B':
            ws.write(REF_1,item,cellValue)
        else:
            ws.write(REF_1,item,currencyFormat)
        j += 1
    
ws.set_column('B:B', 15)
ws.set_column('C:C', 40)
ws.set_column('D:D', 20)

wb.close()