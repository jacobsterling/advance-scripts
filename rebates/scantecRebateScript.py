# -*- coding: utf-8 -*-
"""
Created on Fri Jan 14 15:28:59 2022

@author: jacob.sterling
"""

#importing external modules
import pandas as pd
import datetime
import win32com.client as win32
import xlsxwriter
win32c = win32.constants
from openpyxl.utils import get_column_letter

from utils import formats
from utils import functions


Year = formats.taxYear().Year('-')

datetime.datetime.now().time()
    
Year = '2022'

df1_path: str = r"C:\Users\jacob.sterling\advance.online\J Drive - Exec Reports\Margins Reports\Margins 2021-2022\Margins Report 21-22.xlsx"
df_path: str = r"C:\Users\jacob.sterling\advance.online\J Drive - Exec Reports\Margins Reports\Margins 2022-2023\Margins Report 22-23.xlsx"
#data_path = Path(r"C:\Users\jacob.sterling\OneDrive - advance.online\Documents\Data")

###########run scantec workers from CRM

# Email Address: grace.webber@advance.online
# Password: Advance2018!

# df1_ = pd.read_excel(df1_path, sheet_name= 'Core Data',usecols = ['Client Name','Surname','Forename','Margins','CHQDATE',"Email Address"]).rename(columns={'CHQDATE':'Date Paid','Margins':'Margin'})
# df1_ = df1_[df1_['Margin'] > 0]
# df1_['Client Name'] = df1_['Client Name'].str.upper()
# df1_ = df1_[df1_['Client Name'] == 'SCANTEC PERSONNEL LIMITED'].drop('Client Name',axis=1)
# df1_['Date Paid'] = pd.to_datetime(df1_['Date Paid'],format='%d/%m/%Y')

df_ = pd.read_excel(df_path, sheet_name= 'Core Data',usecols = ['Client Name','Surname','Forename','Margins','CHQDATE',"Email"]).rename(columns={'CHQDATE':'Date Paid','Margins':'Margin'})
df_ = df_[df_['Margin'] > 0]
df_['Client Name'] = df_['Client Name'].str.upper()
df_ = df_[df_['Client Name'] == 'SCANTEC PERSONNEL LIMITED'].drop('Client Name',axis=1)
df_['Date Paid'] = pd.to_datetime(df_['Date Paid'],format='%d/%m/%Y')

Scantec_Workers = pd.read_csv('Scantec+Workers.csv', encoding = 'latin', na_values=["-"])
#Scantec_Workers['Full Name'] =  Scantec_Workers['Last Name']+ ' ' + Scantec_Workers['First Name'] # 'Full Name',
Scantec_Workers = Scantec_Workers[['Consultant',"Email"]].rename(columns={'Consultant':'Consultant Name in CRM'}).dropna()#.drop_duplicates(subset = 'Full Name')

Scantec_Consultants = pd.read_excel('Scantec Consultants.xlsx',usecols = ['Name','Consultant_Code']).drop_duplicates(subset = 'Name')

df_missing_con = pd.DataFrame()

for n in range(8,9,1):######################change range
    if len(str(n)) == 1:
        month_num = '0'+ str(n)
    else:
        month_num = str(n)
    
    datetime_object = datetime.datetime.strptime(month_num, "%m")
    month_name = datetime_object.strftime("%b")
    full_month_name = datetime_object.strftime("%B")
    
    df_path_new: str = rf"Scantec Rebates\Scantec {month_name} {Year}.xlsx"
    
    # df1 = df1_[df1_['Date Paid'].dt.year == int(Year)]
    # df1 = df1[df1['Date Paid'].dt.month == int(month_num)]
    # df1['Full Name'] = df1['Surname'] + ' ' + df1['Forename']

    df = df_[df_['Date Paid'].dt.month == int(month_num)]
    df['Full Name'] = df['Surname'] + ' ' + df['Forename']

    # df = pd.concat([df,df1])

    df = pd.merge(df, Scantec_Workers, how="left").drop('Email',axis = 1)
    df = pd.merge(df, Scantec_Consultants, left_on = 'Consultant Name in CRM', right_on = 'Name', how = 'left').rename(columns={'Consultant_Code':'Reference'}).drop('Full Name',axis = 1).drop('Name',axis = 1)
    
    df_missing_con = df_missing_con.append(df.loc[df['Consultant Name in CRM'].isnull()])
    
    wb = xlsxwriter.Workbook(df_path_new)

    cell_format_value = wb.add_format({'font_size' : 16,
                                   'align': 'center',
                                   'border':1})

    currency_format = wb.add_format({'font_size' : 16,
                                 'align': 'center',
                                 'num_format': '£#,##0.00',
                                 'border':1})

    currency_format_totals = wb.add_format({'font_size' : 16,
                                 'align': 'center',
                                 'num_format': '£#,##0.00',
                                 'bg_color': '#FFFF00',
                                 'border':1})

    cell_format_header = wb.add_format({'bold':     True,
                                    'font_size' : 16,
                                    'align': 'center',
                                    'font_color': '#FF0000'})

    cell_format_totals = wb.add_format({'font_size' : 16,
                                    'align': 'center',
                                    'bg_color': '#FFFF00',
                                    'border':1})

    cell_format_column = wb.add_format({'font_size' : 16,
                                    'align': 'center',
                                    'bg_color': '#92D050',
                                    'border':1})

    idx = pd.DataFrame(df['Date Paid'].unique())[0].apply(lambda x: functions.tax_calcs().tax_week_calc(x)).sort_values()

    df_Totals = pd.DataFrame([],columns=['','£4.75','£3'],index=idx)

    for i, row in pd.DataFrame(df['Date Paid'].unique()).iterrows():
        CHQDATE = row[0]
        Week = functions.tax_calcs().tax_week_calc(CHQDATE)
        if not Week:
            Week = 52
        str_date = CHQDATE.strftime('%d.%m')
        ws = wb.add_worksheet(f'Week {Week} {str_date}')
    
        df_CHQDATE = df[df['Date Paid'] == CHQDATE].reset_index(drop = True)
        df_CHQDATE['Margin'] = 1
        df_CHQDATE['Date Paid'] = CHQDATE.strftime('%d/%m/%Y')
    
        df_Totals.at[Week,''] = len(df_CHQDATE)
        df_Totals.at[Week,'£4.75'] = len(df_CHQDATE)*4.75
        df_Totals.at[Week,'£3'] = 0
    
        tab_REF = ('A1:{width}{length}').format(length = len(df_CHQDATE)+1, width = get_column_letter(len(df_CHQDATE.columns)))
        df_CHQDATE.fillna('',inplace=True)

        for j, column in enumerate(df_CHQDATE.columns.values):
            REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = 1)
            ws.write(REF_1,column,cell_format_header)
            ws.set_column(('{col}:{col}').format(col = get_column_letter(j + 1)), 15)

        for k, row in df_CHQDATE.iterrows():
            j = 0
            for item in row:
                REF_1 = ('{col}{rowf}').format(col = get_column_letter(j + 1), rowf = k + 2)
                ws.write(REF_1,item)
                j += 1
    
    ws = wb.add_worksheet('Totals')

    df_Totals = df_Totals.dropna(axis=0).reset_index()

    ws.write('B5','Week', cell_format_header)
    ws.write('C5','Total numbers paid', cell_format_header)
    ws.write('D6','UMB/CIS', cell_format_value)
    ws.write('E6','UMB/CIS', cell_format_value)

    ws.write('C14','Subtotals', cell_format_totals)
    ws.write('D14',sum(df_Totals['£4.75']), currency_format_totals)
    ws.write('E14',sum(df_Totals['£3']), currency_format_totals)
    ws.write('C17','Total', cell_format_totals)
    ws.write('D17',sum(df_Totals['£4.75'])+sum(df_Totals['£3']), currency_format_totals)

    ws.write('D7','£4.75',cell_format_column)
    ws.write('E7','£3.00',cell_format_column)
    ws.write('C7','',cell_format_value)

    for i, row in df_Totals.iterrows():
        j = 1
        for item in row:
            REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 8)
            if get_column_letter(j + 1) == 'C' or get_column_letter(j + 1) == 'B':
                ws.write(REF_1,item,cell_format_value)
            else:
                ws.write(REF_1,item,currency_format)
            j += 1
        
    ws.set_column('B:B', 15)
    ws.set_column('C:C', 40)
    ws.set_column('D:D', 20)

    wb.close()
    
df_missing_con = df_missing_con.drop_duplicates(subset=['Surname','Forename'])
print(df_missing_con)
df_missing_code = df.loc[df['Reference'].isnull(),'Consultant Name in CRM']
print(df_missing_code)