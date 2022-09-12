# -*- coding: utf-8 -*-
"""
Created on Wed Mar 30 16:38:44 2022

@author: jacob.sterling
"""

class RTW_Report:
    from pathlib import Path
    def __init__(self, path = Path.home() / 'advance.online'):
        self.path = self.Path(path)
        from Formats import taxYear
        import pandas as pd
        self.pd = pd
        import numpy as np
        import datetime
        self.datetime = datetime
        from Functions import PAYNO_Check
        self.re = __import__('re')
        self.namePattern = r'([A-Z][a-z]{2,}).+([A-Z][a-z]{2,})'
        self.paynoPattern = r"([0-9]{4,})"
        self.solutions = ['CIS','PAYE','SE', 'Swedish Degrogation', 'Umbrella']
        self.fileExceptions = []
        
        Week = input('Enter Week Number: ')
        self.Week = Week
        
        Year = taxYear().Yearp('-')
        
        df_path = self.path / rf"J Drive - Exec Reports\Margins Reports\Margins {Year}"
        data_path = df_path / f'Data/Week {Week}'
        
        df_io = False
        df_axm = False
        joiners_io = False
        joiners_axm = False
        
        for file in data_path.glob('*'):
            if file.is_file() and file.name.__contains__('emp'):
                emp = pd.read_csv(file,usecols=['Agency','Payno','Name','Week']).drop_duplicates(subset=['Payno']).dropna(subset=['Payno'], axis = 0).rename(columns={'Agency':'Client Name','Payno':'PAYNO','Week':'Week Number'})
                emp[['OFFNO','PAYNO']] = emp['PAYNO'].str.split('*',expand=True)
                emp[['Forename','Surname','Surname2']] = emp['Name'].str.split(' ',expand=True)
                emp.loc[emp['Surname2'].isnull() == False, 'Surname'] = emp.loc[emp['Surname2'].isnull() == False, 'Surname'] + ' ' + emp.loc[emp['Surname2'].isnull() == False, 'Surname2']
                emp = emp.drop('OFFNO',axis=1).drop('Name',axis=1).drop('Surname2',axis=1)
                emp['Margins'] = 1
                emp['Solution'] = 'PAYE'
                emp = emp.reindex(columns=['Client Name', 'PAYNO','Surname', 'Forename','Solution', 'Margins', 'CHQDATE', 'Week Number'])
            elif file.is_file() and file.name.__contains__('oiners'):
                if file.name.__contains__('axm'):
                    joiners_axm = pd.read_csv(file,usecols=['Pay No','Type','WS_ID_RECEIVED'],encoding = 'latin').rename(columns={'Pay No':'PAYNO'})
                else:
                    joiners_io = pd.read_csv(file,usecols=['Pay No','Type','WS_ID_RECEIVED'],encoding = 'latin').rename(columns={'Pay No':'PAYNO'})
            elif file.is_file() and file.name.__contains__('retained'):
                if file.name.__contains__('axm'):
                    df_axm = pd.read_csv(file)
                else:
                    df_io = pd.read_csv(file)
        
        if joiners_io:
            joiners = pd.concat([joiners_io, joiners_axm]) if joiners_axm else joiners_io
        
        if df_io:
            df = pd.concat([df_io,df_axm]).rename(columns={'Management Fee':'Margins'}) if df_axm else df_io.rename(columns={'Management Fee':'Margins'})
            
        
        if df['Client Name'].str.contains('ALEXANDER MANN').any():
            df['Week Number'] = Week
            df = df.dropna(subset=['PAYNO'], axis = 0).drop_duplicates(subset=['PAYNO']).reset_index(drop=True)
        else:
            print(f'Reading Margins Report {Year}.xlsx.....')
            print(datetime.datetime.now().time())
            
            df_list = pd.read_excel(df_path / f'Margins Report {Year}.xlsx', sheet_name= ['Core Data','Joiners Compliance'])
            joiners = df_list['Joiners Compliance'][['Pay No','Type','WS_ID_RECEIVED']].rename(columns={'Pay No':'PAYNO'})
            df = df_list['Core Data'][['Client Name','PAYNO','Surname','Forename','Margins','CHQDATE','Week Number']]
            df = df[df['Week Number'] == int(Week)].dropna(subset=['PAYNO'], axis = 0).drop_duplicates(subset=['PAYNO']).reset_index(drop=True)
        
        print('Formatting Data.....')
        print(datetime.datetime.now().time())
        
        df['Solution'], emp['CHQDATE'] = np.nan, df.at[0, 'CHQDATE']
        df = pd.concat([df, emp])
        df['Surname'] = df['Surname'].str.title()
        df['Forename'] = df['Forename'].str.title()
        df['PAYNO'] = df['PAYNO'].astype(int)
        joiners = joiners[joiners['PAYNO'].apply(PAYNO_Check)]
        joiners['PAYNO'] = joiners['PAYNO'].astype(int)
        df = df.merge(joiners, how='left', validate='one_to_many', copy=False)
        df.loc[df['Type'] == 'PAYE' ,'Type'] = 'Umbrella'
        df.loc[df['Solution'].isnull() ,'Solution'] = df.loc[df['Solution'].isnull() ,'Type']
        df['RTW Types'] = ''
        df['ETW Types'] = ''
        df['Number of RTW'] = 0
        df['Method'] = ''
        
        self.df = df.drop('Type',axis = 1)
        
    def exportReport(self):
        df = self.df.fillna('').reset_index(drop=True)
        import xlsxwriter
        from openpyxl.utils import get_column_letter
        #df.to_excel(f'RTW Report Week {self.Week}.xlsx',index=False)
        
        wb = xlsxwriter.Workbook(f'RTW Report Week {self.Week}.xlsx')
        
        format1 = wb.add_format({'bg_color': '#FFC7CE',
                                        'font_color': '#9C0006'})
        
        format2 = wb.add_format({'bg_color': '#C6EFCE',
                                        'font_color': '#006100'})
        
        format3 = wb.add_format({'bg_color': '#6CA6CD',
                                        'font_color': '#9C0006'})
        
        cell_format_column = wb.add_format({'font_size' : 12,
                                            'align': 'center',
                                            'border':1})
        
        ws = wb.add_worksheet('RTW Report')
        
        for j, column in enumerate(df.columns.values):
            col = get_column_letter(j + 1)
            row = 1
            rowend = len(df)+1
            ws.write(f'{col}{row}',column,cell_format_column)
            ws.set_column(f'{col}:{col}', 15)
            if column == 'Number of RTW':
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': '==',
                                              'value': 0,
                                              'format': format1})
            
            if column == 'WS_ID_RECEIVED':
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': '==',
                                              'value': 'Yes',
                                              'format': format2})
                
                # ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                #                               'criteria': 'blanks',
                #                               'format': format2})
                
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': '==',
                                              'value': 'No',
                                              'format': format2})
                
            if column == 'Method':
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': 'containing',
                                              'value': 'Name w/ Solution',
                                              'format': format1})
                
                ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                              'criteria': 'containing',
                                              'value': 'Shared',
                                              'format': format1})
            
            # if column == 'Solution':
            #     ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
            #                                   'criteria': 'blanks',
            #                                   'format': format3})
                
        for i, row in df.iterrows():
            j = 0
            for item in row:
                REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
                ws.write(REF_1,item)
                j += 1
            
        wb.close()
        
        return self.df
    
    def emailReport(self):
        import win32com.client as client
        outlook = client.Dispatch('Outlook.Application')
        email = outlook.CreateItem(0)
        email.Display()
        email.To = 'hannah.jarvis@advance.online'
        email.CC = 'jacob.sterling@advance.online; joshua.richards@advance.online'
        email.Subject = (f'RTW Report Week {self.Week}')
        # html = """
        #     </div>
        # """
        email.Attachments.Add(Source=f'RTW Report Week {self.Week}.xlsx')
        email.Send()
    
    @staticmethod
    def RTW_Reader(subfile):
        RTWcsv, n = '', 0
        for RTW in subfile.glob('*'):
            if RTW.suffix != '':
                RTWcsv = RTWcsv + RTW.suffix + ', '
                n += 1
        return RTWcsv, n
    
    def Interpreter(self, fileName):
        forename, surname, othernames, n = False, False, '', 0
        for name in self.re.findall(self.namePattern,fileName):
            forename = name if n == 0 else forename
            surname = name if n == 1 else surname
            othernames = othernames + ' ' + name.upper() if n > 1 else othernames
            n += 1
        try:
           return forename, surname, othernames, int(self.re.findall(self.paynoPattern,fileName)[0])
        except IndexError:
            return forename, surname, othernames, False
    
    def identifier(self, payno, forename, surname, solution):

        if payno and solution:
            if solution == 'Swedish Degrogation':
                index = self.df[(self.df['PAYNO'] == payno) & (self.df['Client Name'] == 'NRL WASHINGTON')].index.values
            else:
                index = self.df[(self.df['PAYNO'] == payno) & (self.df['Solution'] == solution)].index.values
            if len(index) > 0:
                return 'PAYNO w/ Solution, ', self.multiIndexIdentifier(index)
                
        if payno:
            index = self.df[(self.df['PAYNO'] == payno)].index.values
            if len(index) > 0:
                return 'PAYNO no Solution, ', self.multiIndexIdentifier(index)
                
        if forename and surname and solution:
            if solution == 'Swedish Degrogation':
                index = self.df.loc[(self.df['Surname'].str.contains(surname)) & (self.df['Forename'].str.contains(forename)) & (self.df['Client Name'] == 'NRL WASHINGTON')].index.values
            else:
                index = self.df.loc[(self.df['Surname'].str.contains(surname)) & (self.df['Forename'].str.contains(forename)) & (self.df['Solution'] == solution)].index.values
            if len(index) > 0:
                return 'Name w/ Solution, ', self.multiIndexIdentifier(index)
            
        # if forename and surname:
        #     index = self.df.loc[(self.df['Surname'].str.contains(surname)) & (self.df['Forename'].str.contains(forename))].index.values
        #     return self.multiIndexIdentifier(index, 'Name no Solution, ')
        
        return False, []
    
    def multiIndexIdentifier(self, index):
        for i in index:
            if len(index) > 1:
                self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + 'Shared, '
        return index
    
    def write(self,i, RTWcsv, n, method = 'w', ftype = 'RTW'):
        if method == 'o':
            self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + 'Overwritten, '
            self.df.at[i, 'Number of RTW'] = self.df.at[i, 'Number of RTW'] + n - self.df.at[i, f'{ftype} Types'].count(',')
            self.df.at[i, f'{ftype} Types'] = RTWcsv
        elif method == 'w':
            self.df.at[i, 'Number of RTW'] = self.df.at[i, 'Number of RTW'] + n
            self.df.at[i, f'{ftype} Types'] = RTWcsv

    def search(self,file,f = False):
        for subfile in file.glob('*'):
            if (subfile.name.__contains__('ETW') or subfile.name.__contains__('RTW')) and subfile.suffix == '':
                ftype = 'ETW' if subfile.name.__contains__('ETW') else 'RTW'
                RTWcsv, n = self.RTW_Reader(subfile)
                forename, surname, othernames, payno = self.Interpreter(file.name)
                sol = False
                for solution in self.solutions:
                    if str(file).__contains__(solution):
                        sol = solution 
                        break
                
                method, index = self.identifier(payno, forename, surname, sol)
                if method and n > 0:
                    for i in index:
                        if self.df.at[i, f'{ftype} Types'] != '':
                            if self.df.at[i, 'Method'].__contains__('PAYNO w/ Solution, '):
                                if method == 'PAYNO w/ Solution, ':
                                    self.write(i, RTWcsv, n, 'o', ftype)
                                    self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + method
                                
                            elif self.df.at[i, 'Method'].__contains__('PAYNO no Solution, '):
                                if method == 'PAYNO w/ Solution, ' or method == 'PAYNO no Solution, ':
                                    self.write(i, RTWcsv, n, 'o', ftype)
                                    self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + method
                                    
                            elif self.df.at[i, 'Method'].__contains__('Name w/ Solution, '):
                                self.write(i, RTWcsv, n, 'o', ftype)
                                self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + method
                        else:
                            self.write(i, RTWcsv, n, 'w', ftype)
                            self.df.at[i, 'Method'] = self.df.at[i, 'Method'] + method
                        print(f"+{n} {ftype} found for payno {self.df.at[i, 'PAYNO']}!!")
                        f = True
        self.fileExceptions.append(file)
            
    def run(self):
        start = self.datetime.datetime.now()
        print(self.datetime.datetime.now().time())
        try:
            for contractorFile in self.path.glob('*'):
                if contractorFile.name.__contains__('Contractors Files'):
                    print(f'Searching {contractorFile.name}.....')
                    print(self.datetime.datetime.now().time())
                    for solution in contractorFile.glob('*'):
                        if solution.name in self.solutions:
                            if solution.name.__contains__('Umbrella') or solution.name.__contains__('Swedish'):
                                if contractorFile.name.__contains__('2015') or contractorFile.name.__contains__('2016'):
                                    globSol = '*'
                                else:
                                    globSol = '*/*'
                            else:
                                globSol = '*'
                            for file in solution.glob(globSol):
                                if (self.re.search(self.namePattern, file.name) or self.re.search(self.paynoPattern, file.name)) and file not in self.fileExceptions:
                                    self.search(file)
        except KeyboardInterrupt:       
            pass
        print('Done.')
        print(self.datetime.datetime.now().time())
        print(start - self.datetime.datetime.now())
        df = self.exportReport()
        self.emailReport()
        return df, self.fileExceptions
        
df, exceptions = RTW_Report().run()


    
