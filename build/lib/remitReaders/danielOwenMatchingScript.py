# -*- coding: utf-8 -*-
"""
Created on Thu May 12 17:13:05 2022

@author: jacob.sterling
"""

import pandas as pd
from pathlib import Path
from Formats import taxYear
import pdfplumber
import re
import xlsxwriter
from openpyxl.utils import get_column_letter
    
homePath = Path.home() / r"advance.online"
Year = taxYear().Year(' - ')
filePath = homePath / rf"J Drive - Operations\Remittances and invoices\Daniel Owen\Tax Year {Year}"

def output(Week: int = None, nominal: str = None):
    Week = input("Enter Week number: ") if not Week else Week
    
    joinersPath = homePath / rf"J Drive - Exec Reports\Margins Reports\Margins {taxYear().Year('-')}\Data\Week {Week}"
    
    for file in joinersPath.glob("*"):
        if file.name.__contains__("oiners"):
            joiners = pd.read_csv(file, encoding="latin", usecols=["Pay No","Forenames","Surname","CONT_AGENCY_NAME","Date of Birth","Email Address","MOBILE"])
            if file.name.__contains__("io"):
                break
    
    invoicePath = Path("M:\Work\PSF\Invoice")
    
    for file in invoicePath.glob("*"):
        if file.name.__contains__(f"inv_IO{Week}_{taxYear().yearc}"): 
            invoice = pd.read_csv(file, header = None)
    
    invoice = invoice[invoice[7] == "S"].drop([3,7,10,11,12,13], axis = 1).reset_index(drop=True)
    invoice.columns = ["Invoice","Invoice Number","Invoice Date","Pay No","Net","Full Name","Week","Nominal Code"]
    invoice["Merit Gross"] = invoice["Net"]*-1.2
    invoice["Pay No"] = invoice["Pay No"].astype(int)
    invoice = invoice.merge(joiners, how = "left")
    
    invoice["Full Name"] = invoice["Full Name"].apply(lambda x: nameStrip(x).title())
    #pertempCodes = ["56","87","92","103","214","254","295","355","480","515","522","523","524","542","592","NETW02","PERT01","PERT03","THEE01"]

    return invoice[invoice["Nominal Code"] == nominal ] if nominal else invoice

def matching(Week: int = None):
    Week = input("Enter Week number: ") if not Week else Week
    
    invoice = output(Week, "317")
        
    result, totals = runWeek(Week)
    
    def removeMiddleName(name):
        splitName = name.split(" ")
        return splitName[0] + " " + splitName[-1]
    
    result["PDF Gross"] = result["Amount"]
    
    #totals["Total"] = totals["Total"]*1.2
    
    resultGrouped = result.groupby(["Worker Name"]).agg({"PDF Gross":sum}).reset_index()
    
    resultGrouped["Worker Name"] = resultGrouped["Worker Name"].apply(lambda x: removeMiddleName(x).title())
    
    invoiceGrouped = invoice.groupby(["Full Name"]).agg({"Merit Gross":sum}).reset_index()
    
    invoiceGrouped["Full Name"] = invoiceGrouped["Full Name"].apply(lambda x: removeMiddleName(x).title())
    
    matching = invoiceGrouped.merge(resultGrouped, left_on="Full Name", right_on="Worker Name", how = "outer")
            
    matching["Difference"] = matching["Merit Gross"] - matching["PDF Gross"]
    
    matching = matching[(matching["Difference"].abs() > 0.02) | (matching["Difference"].isna())]
    
    matching = matching.fillna('').reset_index(drop=True)
    
    wb = xlsxwriter.Workbook(filePath / rf'Week {Week}/Daniel Owen Py Matching Script Week {Week}.xlsx')
    
    format1 = wb.add_format({'bg_color': '#FFC7CE',
                                    'font_color': '#9C0006'})
    
    format2 = wb.add_format({'bg_color': '#C6EFCE',
                                    'font_color': '#006100'})
    
    cell_format_column = wb.add_format({'font_size' : 12,
                                        'align': 'center',
                                        'border':1})
    
    money_fmt = wb.add_format({'num_format':'[$£]#,##0.00'})
    
    ws = wb.add_worksheet('Matching')
    
    for j, column in enumerate(matching.columns.values):
        col = get_column_letter(j + 1)
        row = 1
        rowend = len(matching)+1
        ws.write(f'{col}{row}',column,cell_format_column)
        ws.set_column(f'{col}:{col}', 15)
        if column == "Difference":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                          'criteria': 'between',
                                          'minimum': -0.02,
                                          'maximum': 0.02,
                                          'format': format2})
            
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                          'criteria': 'not between',
                                          'minimum': -0.02,
                                          'maximum': 0.02,
                                          'format': format1})
            
            ws.set_column(f'{col}:{col}', 12, money_fmt)
            
        if column == 'Worker Name':
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
        
        if column == "Merit Gross":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
            ws.set_column(f'{col}:{col}', 12, money_fmt)
            
        if column == "Full Name":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
        
        if column == "PDF Gross":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
            ws.set_column(f'{col}:{col}', 12, money_fmt)
            
        if column == "PDF Name":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
    for i, row in matching.iterrows():
        j = 1
        for item in row:
            REF_1 = ('{col}{row}').format(col = get_column_letter(j), row = i + 2)
            if j == len(matching.columns):
                ws.write_formula(REF_1,f"=SUM(B{i + 2} - D{i + 2})")
            else:
                ws.write(REF_1,item)
            j += 1
    
    ws = wb.add_worksheet('Invoice Data')
    
    invoice = invoice.fillna('').reset_index(drop=True)
    
    for j, column in enumerate(invoice.columns.values):
        col = get_column_letter(j + 1)
        row = 1
        rowend = len(invoice)+1
        ws.write(f'{col}{row}',column,cell_format_column)
        ws.set_column(f'{col}:{col}', 15)
            
        if column == "Merit Gross":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
            ws.set_column(f'{col}:{col}', 12, money_fmt)
        
    for i, row in invoice.iterrows():
        j = 1
        for item in row:
            REF_1 = ('{col}{row}').format(col = get_column_letter(j), row = i + 2)
            ws.write(REF_1,item)
            j += 1
        
    ws = wb.add_worksheet('PDF Data')
    
    result = result.fillna('').reset_index(drop=True)
    
    for j, column in enumerate(result.columns.values):
        col = get_column_letter(j + 1)
        row = 1
        rowend = len(result)+1
        ws.write(f'{col}{row}',column,cell_format_column)
        ws.set_column(f'{col}:{col}', 15)
        
        if column == "PDF Gross":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type':'blanks',
                                          'format': format1})
            
            ws.set_column(f'{col}:{col}', 12, money_fmt)
            
    for i, row in result.iterrows():
        j = 1
        for item in row:
            REF_1 = ('{col}{row}').format(col = get_column_letter(j), row = i + 2)
            ws.write(REF_1,item)
            j += 1
            
    ws = wb.add_worksheet('Totals')
    
    totals = totals.fillna('').reset_index(drop=True)
    
    for j, column in enumerate(totals.columns.values):
        col = get_column_letter(j + 1)
        row = 1
        rowend = len(totals)+1
        ws.write(f'{col}{row}',column,cell_format_column)
        ws.set_column(f'{col}:{col}', 15)
        if column == "Difference":
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                          'criteria': 'between',
                                          'minimum': -0.02,
                                          'maximum': 0.02,
                                          'format': format2})
            
            ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                          'criteria': 'not between',
                                          'minimum': -0.02,
                                          'maximum': 0.02,
                                          'format': format1})
    
    ws.set_column('B:D', 12, money_fmt)
    
    for i, row in totals.iterrows():
        j = 1
        for item in row:
            REF_1 = ('{col}{row}').format(col = get_column_letter(j), row = i + 2)
            if j == len(totals.columns):
                ws.write_formula(REF_1,f"=SUM(B{i + 2} - C{i + 2})")
            else:
                ws.write(REF_1,item)
            j += 1
            
    wb.close()
    return matching

def nameStrip(name):
    newName = []
    n = 0
    for char in name.upper():
        if char.isupper():
            newName.append(char)
            n = 1
        elif n == 1:
            newName.append(char)
            n = 0
        else:
            n = 0
    if newName[-1] == " ":
        del newName[-1]
    return "".join(newName)

def runWeek(Week: int = None):
    Week = input("Enter Week number: ") if not Week else Week
    
    weekPath = filePath / rf"Week {Week}"
    
    result = pd.DataFrame([], columns = ["Worker Name", "Date", "Hours", "Rate", "Amount","PDF Name"])
    totals = pd.DataFrame([], columns = ["PDF Name", "Total", "PDF Total", "Difference"])
    for file in weekPath.glob("*"):
        if file.suffix in [".PDF", ".pdf"] and file.name.__contains__("PL"):
            print(f'Reading {file.name}......')
            df, total = danielOwenPDFReader(file)
            #df["Gross"] = df["Amount"]*1.2
            df["PDF Name"] = file.name
            result = pd.concat([result, df]).reset_index(drop = True)
            temp = pd.DataFrame([[file.name, df["Amount"].sum(), total*1.2, df["Amount"].sum() - total*1.2]], columns = totals.columns)
            totals = pd.concat([totals, temp]).reset_index(drop = True)
    #result.to_csv(weekPath / rf"Daniel Owen Py Import Week {Week}.csv")
    return result, totals

def danielOwenPDFReader(pdf):
    result = pd.DataFrame([], columns = ["Worker Name", "Date", "Hours", "Rate", "Amount"])
    newLinePattern = re.compile(r"^(.*), (.*) on  ?([A-Z][a-z]+ [0-9]+ [0-9]+) (-?[1-9],?[0-9]*\.[0-9]+)+ (-?[1-9],?[0-9]*\.[0-9]+)+ (-?[1-9],?[0-9]*\.[0-9]+)+$")
    totalPattern = re.compile(r"Amount payable ([1-9][0-9]+,?[0-9]+\.[0-9]{2})")
    total = 0
    pdf = pdfplumber.open(pdf)
    for page in pdf.pages:
        text = page.extract_text()
        for line in text.split("\n"):
            if newLinePattern.match(line):
                groups = re.search(newLinePattern, line).groups()
                forename = groups[1]
                surnames = groups[0]
                date = groups[2]
                hours = float(groups[3].replace(",",""))
                rate = float(groups[4].replace(",",""))
                amount = float(groups[5].replace(",",""))
                
                result = pd.concat([result, pd.DataFrame([[forename + " " + surnames,
                                                           date,
                                                           hours,
                                                           rate,
                                                           amount
                    ]], columns = result.columns)]).reset_index(drop=True)
            elif totalPattern.match(line):
                total = float(re.search(totalPattern, line).groups()[0].replace(",","").replace(" ",""))
                
    return result, total

df = matching(5)