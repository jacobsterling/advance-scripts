# -*- coding: utf-8 -*-
"""
Created on Wed Jan 26 09:25:00 2022

@author: jacob.sterling
"""

#importing modules
#pandas is used to create dataframes import and export data (export has limitations)
import pandas as pd
#datetime module used to handle dates and calculate tax week
from datetime import date
import datetime
#pathlib is usefull for its windows path functions
from pathlib import Path
#tabula is used to read tables from PDF's
from tabula import read_pdf
#xlsxwriter is used to create excel files
import xlsxwriter
#used to translate a number an excel column letter
from openpyxl.utils import get_column_letter

#gets todays date in date time format
datetime.datetime.now()
today = date.today()

#creating a function called tax_week_calc to calculate the tax week from the isocalendar week
def tax_week_calc(week):
    if week < 15:
        week = week + 38
    else:
        week = week - 14
    return week

#today.isocalendar()[1] gives the current week number and tax_week_calc converts it to the tax week
Week = tax_week_calc(today.isocalendar()[1])

#asks the user for the tax week number you want to match the remits for
Week = int(input('Enter Week Number: '))

#calculates the previous week
if Week > 1:
    Weekp = Week - 1
else:
    Weekp = 52

#gives the current year
Year = today.year

#calculating previous and concurrent years
yearppp = Year - 3
yearpp = Year - 2
yearp = Year - 1
yearc = Year
yearcc = Year + 1

#calculates current tax year range using above year calculations
if today.isocalendar()[1] < 39:
    Yearpp = (f'{yearpp} - {yearp}')
    Yearp = (f'{yearp} - {yearc}')
    Year = (f'{yearc} - {yearcc}')
else:
    Yearpp = (f'{yearppp} - {yearpp}')
    Yearp = (f'{yearpp} - {yearp}')
    Year = (f'{yearp} - {yearc}')


#defining file paths
##############################################################################

file_path = Path(rf"O:\Remittances and invoices\NRL\Tax Year {Year}\Week {Week}\CIS")

##############################################################################

import_path = Path(rf"{file_path}\NRL CIS Python Import Wk {Week}.csv")
review_path = Path(rf"{file_path}\NRL CIS Python Matching Wk {Week}.xlsx")

##############################################################################

#asks user for the report file name (since its name varies)
Report_Name = input('Enter Report File Name: ')
report_path = Path(rf"{file_path}\{Report_Name}.xlsx")

##############################################################################

#reads the report and gives as a dataframe
df_report = pd.read_excel(report_path,usecols = ['Worker','Worker Ref','Period Ending','Rate Name','Units','Rate','Total','NI Number'])
#changes the rate names that contain overtime into overtime
df_report.loc[df_report['Rate Name'].str.contains('overtime'),'Rate Name'] = 'Overtime'
df_report.loc[df_report['Rate Name'].str.contains('Overtime'),'Rate Name'] = 'Overtime'

#takes the worker and total column from the report
df = df_report[['Worker','Total']]
#renames the total column and fills empty values with 0
df = df.rename(columns = {'Total':'Net Total Report'}).fillna(0)
#creates a new column of 0's
df['Net Total Remit'] = 0
#groups the dataframe by the worker name and sums the total column to 2 significant figures
df = df.groupby(['Worker']).sum().round(2)

#creates an empty dataframe that gives clarity to what part of the pdf was used
df_ADDED = pd.DataFrame([],columns=['Rows Added','Pdf'])

#glob iterates through the files in file_path 
for file in file_path.glob('*'):
    #checks if it is a file
    if file.is_file():
        #checks if it is a pdf that begins with SB-
        if file.suffix == ".pdf" and 'SB-' in file.name:
            #reads the 1st page of the pdf, guess=False prevents the function from guessing where the table is in the pdf, squeeze converts it from a dataframe to a series
            df_pdf = read_pdf(file,pages=1, guess=False)[0].squeeze().fillna('')
            try:
                #gets the workers full name from the pdf
                fullname = df_pdf.name.split(' - ')[0].split(' ')
            except AttributeError:
                #if it cant find the full name it will iterate through the pdf to find it
                for col in df_pdf.columns:
                    if col != df_pdf.columns[0]:
                        df_pdf[df_pdf.columns[0]] = df_pdf[df_pdf.columns[0]] + ' ' + df_pdf[col]
                        df_pdf = df_pdf.drop(col,axis=1)
                df_pdf = df_pdf.squeeze()
                fullname = df_pdf.name.split(' - ')[0].split(' ')
            
            #takes the net total from the pdf and adds it to the dataframe
            df.loc[fullname[-1] + ', ' + fullname[0],'Net Total Remit'] = float(df_pdf[df_pdf.str.contains('Net')].values[1].split(' ')[-1].replace(',',''))
            #adds the line in the pdf where the net total was taken from and adds it to the below dataframe
            df_ADDED = df_ADDED.append(pd.DataFrame([[df_pdf[df_pdf.str.contains('Net')].values[1],file.name]],index=[fullname[-1] + ', ' + fullname[0]],columns=['Rows Added','Pdf']))

#fills empty values with 0's
df['Net Total Report'] = df['Net Total Report'].fillna(0)
df['Net Total Remit'] = df['Net Total Remit'].fillna(0)

#calculates the variance
df['Variance'] = df['Net Total Report'] - df['Net Total Remit']

#creates the summary table by adding the above columns
df_sum = df.sum().reset_index().rename(columns = {0:'Net Total'}).fillna('')
#merges the 2 dataframes by worker name
df = pd.merge(df,df_ADDED.drop('Rows Added',axis = 1),left_index=True,right_index=True,how = 'left')

#exports the report to a csv file for merit upload
df_report.to_csv(import_path,index=False)

#turning the dataframes index into a column so it can write it to the excel file
df = df.reset_index().rename(columns = {'index':'Worker'}).fillna('')

#creates a excel workbook at the review_path
wb = xlsxwriter.Workbook(review_path)

#creating the red and green cell colour formats for the conditonal formatting
format1 = wb.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006'})

format2 = wb.add_format({'bg_color': '#C6EFCE',
                               'font_color': '#006100'})

#creating a sheet for the summary in the workbook
ws = wb.add_worksheet('NRL CIS Summary')

#writing the column names to the workbook and adds conditional formatting to the variance columns
for j, column in enumerate(df.columns.values):
    col = get_column_letter(j + 1)
    row = 1
    rowend = len(df)+1
    ws.write(f'{col}{row}',column)
    ws.set_column(f'{col}:{col}', 15)
    if column == 'Variance':
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '<',
                                      'value': 0,
                                      'format': format1})
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '>',
                                      'value': 0,
                                      'format': format1})
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '=',
                                      'value': 0,
                                      'format': format2})

#writing the values to the workbook
for i, row in df.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
        ws.write(REF_1,item)
        j += 1

#writing the summary columns to the workbook
for j, column in enumerate(df_sum.columns.values):
    col = get_column_letter(j + len(df.columns)+ 3)
    row = 1
    ws.write(f'{col}{row}',column)
    ws.set_column(f'{col}:{col}', 15)

#writing the summary values to the workbook
for i, row in df_sum.iterrows():
    j = len(df.columns)+3
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j), row = i + 2)
        ws.write(REF_1,item)
        j += 1 

wb.close()