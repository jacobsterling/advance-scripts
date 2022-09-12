# -*- coding: utf-8 -*-
"""
Created on Thu Jan 20 16:57:19 2022

@author: jacob.sterling
"""

import pandas as pd
import numpy as np
from datetime import date
import xlsxwriter
from openpyxl.utils import get_column_letter

def age(birthdate):
    DOB = str(birthdate).split('/')
    today = date.today()
    age = today.year - int(DOB[2]) - ((today.month, today.day) < (int(DOB[1]), int(DOB[0])))
    return age

PAY_DESC = dict([('Company Income',1),
                 ('Day rate EDU - TEN',6),
                 ('Day Rate EDU - TEN',6),
                 ('Daily Rate',7.5),
                 ('Overtime',1),
                 ('Overtime 1',1),
                 ('Overtime 2',1),
                 ('Day rate EDU',6.5),
                 ('Day Rate EDU',6.5),
                 ('Day Rate EDU - GSL',7),
                 ('Day Rate EDU - Coba',6),
                 ('RSS DAY',10),
                 ('Standard',1),
                 ('Standard Rate',1),
                 ('Standard rate',1),
                 ('Income Other',0),
                 ('Expense.',0),
                 ('Mileage',0),
                 ('Bonus',0),
                 ('Rate Adj',0),
                 ('Accomodatin non VAT',0),
                 ('Expense No VAT',0),
                 ('Holiday Pay',0),
                 ('SSP',0),
                 ('SNP',0),
                 ('SMP',0),
                 ('RSS day',0),
                 ('Margin Refund',0),
                 (np.nan,0)])

PAY_DESC_DAY_RATE = ['Day rate EDU - TEN',
                     'Day Rate EDU - TEN',
                     'Daily Rate',
                     'Day rate EDU',
                     'Day Rate EDU',
                     'Day Rate EDU - GSL',
                     'Day Rate EDU - Coba',
                     'RSS DAY']

#reads csv file
df_MCR = pd.read_csv("Timesheets Hist.csv",
                     encoding = 'latin')


df_Joiners_Error_Report = pd.read_csv("Joiners Error Report.csv",
                                      encoding = 'latin',
                                      usecols = ['Pay No','Sdc Option', 'Type', 'Date of Birth','DOJ','FEE_TYPE','REWARDS'])



df_Salary_Sacrifice = pd.read_csv('Salary Sacrifice.csv',
                                  encoding = 'latin')

df_MCR['PAYNO'] = df_MCR['PAYNO'].apply(pd.to_numeric, errors='ignore', downcast='float')
df_Joiners_Error_Report['Pay No'] = df_Joiners_Error_Report['Pay No'].apply(pd.to_numeric, errors='ignore', downcast='float')
df_Salary_Sacrifice ['PAYNO'] = df_Salary_Sacrifice ['PAYNO'].apply(pd.to_numeric, errors='ignore', downcast='float')

#df_Joiners_Error_Report['DOJ'] = pd.to_datetime(df_Joiners_Error_Report['DOJ'], format='%d/%m/%Y')

df_Salary_Sacrifice = pd.DataFrame(df_Salary_Sacrifice, columns= ['PAYNO', 'DED_ONGOING'])
df_Salary_Sacrifice = df_Salary_Sacrifice.dropna()

df_Salary_Sacrifice['DED_ONGOING']=df_Salary_Sacrifice['DED_ONGOING'].str.replace(',', '').astype(int)/100
df_Salary_Sacrifice = df_Salary_Sacrifice.apply(pd.to_numeric, errors='ignore', downcast='float')

df_MCR = pd.merge(df_MCR, df_Salary_Sacrifice, left_on = 'PAYNO', right_on = 'PAYNO', how = 'left')



df_MCR['DED_ONGOING'] = df_MCR['DED_ONGOING'].replace(np.nan, 0)

df = pd.DataFrame([(0,0,0,0,0,0,0,0,0,0,0,0)] ,columns = ['PAYNO', 'T/S NUMBER', 'TEMPNAME', 'COMPNAME', 
                                                    'TOTAL HOURS', 'TOTAL PAY','CONTRACTING RATE', 
                                                    'COMPANY INCOME TOTAL','DAY RATE TOTAL','DAY RATE TYPE',
                                                    'SALARY SACRIFICE','Week'])
n = 0

for i, item in df_MCR.iterrows():
    if item['PAY_DESC'] in PAY_DESC:
        TOTAL_HOURS = item['HOURS']*PAY_DESC[item['PAY_DESC']]
        if TOTAL_HOURS > 0:
            TOTAL_PAY = item['HOURS']*item['PAY_RATE']
            CONTRACTING_RATE = (TOTAL_PAY - item['DED_ONGOING'])/TOTAL_HOURS
        else:
            TOTAL_PAY = 0
            CONTRACTING_RATE = 0
        if pd.isnull(item['PAYNO']) == False:
                if item['PAY_DESC'] == 'Company Income':
                    COMPANY_INCOME_TOTAL = TOTAL_PAY
                else:
                    COMPANY_INCOME_TOTAL = 0
                if item['PAY_DESC'] in PAY_DESC_DAY_RATE:
                    DAY_RATE_TOTAL = item['PAY_RATE']
                    DAY_RATE_TYPE = PAY_DESC[item['PAY_DESC']]
                else:
                    DAY_RATE_TOTAL = 0
                    DAY_RATE_TYPE = 0
                ls = [item['PAYNO'], item['T/S Number'], item['TEMPNAME'], 
                      item['COMPNAME'], TOTAL_HOURS, TOTAL_PAY, CONTRACTING_RATE, 
                      COMPANY_INCOME_TOTAL, DAY_RATE_TOTAL, DAY_RATE_TYPE, item['DED_ONGOING'],item['Week']]
                row = pd.Series(ls, index=df.columns)
                df = df.append(row, ignore_index=True)
                n += 1
        else:
            TOTAL_HOURS = df.at[n, 'TOTAL HOURS'] + TOTAL_HOURS
            if TOTAL_HOURS > 0:
                TOTAL_PAY = df.at[n, 'TOTAL PAY'] + TOTAL_PAY
                df.at[n, 'CONTRACTING RATE'] = (TOTAL_PAY - df.at[n, 'SALARY SACRIFICE'])/TOTAL_HOURS
            else:
                TOTAL_PAY = 0
                df.at[n, 'CONTRACTING RATE'] = 0
                
            df.at[n, 'TOTAL HOURS'] = TOTAL_HOURS
            df.at[n, 'TOTAL PAY'] = TOTAL_PAY
            
            if item['PAY_DESC'] == 'Company Income':
                df.at[n, 'COMPANY INCOME TOTAL'] = df.at[n, 'COMPANY INCOME TOTAL'] + TOTAL_PAY 
                
            if item['PAY_DESC'] in PAY_DESC_DAY_RATE:
                if item['PAY_RATE'] < df.at[n, 'DAY RATE TOTAL'] or df.at[n, 'DAY RATE TOTAL'] == 0:
                    df.at[n, 'DAY RATE TOTAL'] = item['PAY_RATE']
                    df.at[n, 'DAY RATE TYPE'] = PAY_DESC[item['PAY_DESC']]
    else:
        print('Error - Undefined Pay Description : ',item['PAY_DESC'])
df = df.drop(0 ,axis = 0)
df.reset_index(drop=True)

df['COMPANY INCOME TOTAL'] = df['COMPANY INCOME TOTAL']/df['TOTAL HOURS']
df = df[df['TOTAL HOURS'] > 0]
df = df[df['TOTAL PAY'] > 0]

df['TOTAL HOURS'] = df['TOTAL HOURS'].round(decimals=1)
df['TOTAL PAY'] = df['TOTAL PAY'].round(decimals=2)
df['CONTRACTING RATE'] = df['CONTRACTING RATE'].round(decimals=2)
df['SALARY SACRIFICE'] = df['SALARY SACRIFICE'].round(decimals=2)

df = pd.merge(df, df_Joiners_Error_Report, left_on = 'PAYNO', right_on = 'Pay No', how = 'left')
df = df.drop(['Pay No'], axis = 1)
df['PAYNO'] = df['PAYNO'].astype(int).round()
df_path: str = r"C:\Users\jacob.sterling\OneDrive - advance.online\Documents\ACR\fees retained ts.csv"
df_margins = pd.read_csv(df_path)[['TSNO','Management Fee']]
df_margins[['OFFNO','TSNO']] = df_margins['TSNO'].str.split('*',expand = True)
df_margins = df_margins.drop('OFFNO',axis=1)
df_margins = df_margins.drop_duplicates(subset='TSNO', keep='first')
df = pd.merge(df, df_margins, left_on='T/S NUMBER', right_on ='TSNO', how = 'left').drop('TSNO',axis=1)

df = df.sort_values(by=['COMPNAME','PAYNO','Week']).reset_index(drop=True).fillna('0')
df['Management Fee'] = df['Management Fee'].astype(float)
df['Fee'] = float(0.00)

for i, row in df.iterrows():
    if row['FEE_TYPE'] == '#REF!':
        FEE_TYPE_VALUE = 0
    elif row['FEE_TYPE'] == '£12.95 WEE':
        FEE_TYPE_VALUE = 12.95
    elif row['FEE_TYPE'] == 'EDU':
        FEE_TYPE_VALUE = row['TOTAL HOURS']/6*2.88
        WEEKS_WORKED = round(row['TOTAL HOURS']/25)
        if FEE_TYPE_VALUE > 12*WEEKS_WORKED and WEEKS_WORKED > 0:
            FEE_TYPE_VALUE = 12*WEEKS_WORKED
    elif row['FEE_TYPE'] == 'SCAT':
        FEE_TYPE_VALUE = 15.95
    elif row['FEE_TYPE'] == 'INIT' or row['FEE_TYPE'] == 'KFOX':
        FEE_TYPE_VALUE = 15
    elif row['FEE_TYPE'] == 'RSS1':
        if row['TOTAL HOURS'] < 12:
            FEE_TYPE_VALUE = 6.90
        elif 12 <= row['TOTAL HOURS'] < 16:
            FEE_TYPE_VALUE = 11.70
        elif 16 <= row['TOTAL HOURS'] < 20:
            FEE_TYPE_VALUE = 11.95
        elif 20 <= row['TOTAL HOURS'] < 24:
            FEE_TYPE_VALUE = 15.45
        elif 24 <= row['TOTAL HOURS'] < 32:
            FEE_TYPE_VALUE = 18.45
        elif 32 <= row['TOTAL HOURS'] < 40:
            FEE_TYPE_VALUE = 20.95
        elif 40 <= row['TOTAL HOURS'] < 48:
            FEE_TYPE_VALUE = 23.45
        elif 48 <= row['TOTAL HOURS'] < 55:
            FEE_TYPE_VALUE = 24.45
        else:
            FEE_TYPE_VALUE = 26.95
    elif row['FEE_TYPE'] == 'RSS':
        if row['TOTAL HOURS'] < 12:
            FEE_TYPE_VALUE = 4.95
        elif 12 <= row['TOTAL HOURS'] < 16:
            FEE_TYPE_VALUE = 9.75
        elif 16 <= row['TOTAL HOURS'] < 20:
            FEE_TYPE_VALUE = 10
        elif 20 <= row['TOTAL HOURS'] < 24:
            FEE_TYPE_VALUE = 13.5
        elif 24 <= row['TOTAL HOURS'] < 32:
            FEE_TYPE_VALUE = 16.5
        elif 32 <= row['TOTAL HOURS'] < 40:
            FEE_TYPE_VALUE = 19
        elif 40 <= row['TOTAL HOURS'] < 48:
            FEE_TYPE_VALUE = 21.5
        elif 48 <= row['TOTAL HOURS'] < 55:
            FEE_TYPE_VALUE = 22.5
        else:
            FEE_TYPE_VALUE = 25
    elif row['FEE_TYPE'] == 'SCANT':
        if row['TOTAL HOURS'] <= 24:
            FEE_TYPE_VALUE = 14
        elif row['CONTRACTING RATE'] < 15:
            FEE_TYPE_VALUE = 21.9
        else:
            FEE_TYPE_VALUE = 26.9
    elif row['FEE_TYPE'] == '3.98 PER D':
        FEE_TYPE_VALUE = round(row['TOTAL HOURS']/24)*3.98
        if round(row['TOTAL HOURS']/24) == 0 and row['TOTAL HOURS']>0:
            FEE_TYPE_VALUE = 3.98
    elif row['FEE_TYPE'] == '£18PH+ MAR':
        FEE_TYPE_VALUE = 24.95
    elif row['FEE_TYPE'] == 'TEN' or row['FEE_TYPE'] == 'TENED':
        if row['TOTAL HOURS'] < 6:
            FEE_TYPE_VALUE = 2.88
        elif 6 <= row['TOTAL HOURS'] < 12:
            FEE_TYPE_VALUE = 5.76
        elif 12 <= row['TOTAL HOURS'] < 18:
            FEE_TYPE_VALUE = 8.64
        elif 18 <= row['TOTAL HOURS'] < 24:
            FEE_TYPE_VALUE = 11.52
        else:
            FEE_TYPE_VALUE = 12
    elif row['FEE_TYPE'] == '£2.88 PER ':
        FEE_TYPE_VALUE = round(row['TOTAL HOURS']/24)*2.88
        if FEE_TYPE_VALUE > 12:
            FEE_TYPE_VALUE = 12
    elif row['FEE_TYPE'] == '£4 A DAY C':
        FEE_TYPE_VALUE = round(row['TOTAL HOURS']/24)*4
        if FEE_TYPE_VALUE > 36:
            FEE_TYPE_VALUE = 36
    else:
        FEE_TYPE_VALUE = round(float(row['FEE_TYPE'].replace('£','').replace('P','').replace('T','').replace(' ','')),2)
        if FEE_TYPE_VALUE > 100:
            FEE_TYPE_VALUE = FEE_TYPE_VALUE/100
    #if row['REWARDS'] == 'Yes':
       # FEE_TYPE_VALUE += 1.99
    df.loc[i,'Fee'] = FEE_TYPE_VALUE

df['Difference'] = df['Management Fee'] - df['Fee']
df['Difference'] = df['Difference'].round(2)

# df = df[df['Management Fee'] != 10]
# df = df[df['Management Fee'] != 0]
# df = df[df['Difference'] != 0]
# df = df[df['Difference'] != 1.99]

df = df.sort_values(by=['COMPNAME','PAYNO','Week','Difference']).reset_index(drop=True)
df['No. of Timesheets w/ Missed Revenue'] = 0
df['No. of Timesheets w/ Extra Revenue'] = 0
df.loc[df['Difference'] >= 0, 'No. of Timesheets w/ Extra Revenue'] = 1
df.loc[df['Difference'] < 0, 'No. of Timesheets w/ Missed Revenue'] = 1

df_pivot = pd.pivot_table(df, values=['Difference','No. of Timesheets w/ Extra Revenue','No. of Timesheets w/ Missed Revenue'], index=['COMPNAME','PAYNO'],aggfunc={'Difference': np.sum, 'No. of Timesheets w/ Missed Revenue': np.sum, 'No. of Timesheets w/ Extra Revenue': np.sum}, fill_value=0, margins=True).reset_index()
df = df.drop(['No. of Timesheets w/ Extra Revenue','No. of Timesheets w/ Missed Revenue'],axis=1)
df_pivot['Total Workers'] = 1
df_pivot = df_pivot.groupby(['COMPNAME'], as_index=False).agg({'Difference':np.sum,'No. of Timesheets w/ Extra Revenue':np.sum,'No. of Timesheets w/ Missed Revenue':np.sum,'Total Workers':np.sum})

df_pivot['Total Timesheets'] = df_pivot['No. of Timesheets w/ Missed Revenue'] + df_pivot['No. of Timesheets w/ Extra Revenue']
df_pivot['% of Timesheets w/ missed Revenue'] = (df_pivot['No. of Timesheets w/ Missed Revenue'] / df_pivot['Total Timesheets'])*100

wb = xlsxwriter.Workbook('ACR data.xlsx')

format1 = wb.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006'})

format2 = wb.add_format({'bg_color': '#C6EFCE',
                               'font_color': '#006100'})

cell_format_column = wb.add_format({'bold':     True,
                                    'font_size' : 16,
                                    'align': 'center',
                                    'font_color': '#FF0000'})

ws = wb.add_worksheet('Difference in Fees Data')

for j, column in enumerate(df.columns.values):
    col = get_column_letter(j + 1)
    row = 1
    rowend = len(df)+1
    ws.write(f'{col}{row}',column,cell_format_column)
    ws.set_column(f'{col}:{col}', 15)
    if column == 'Difference':
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '<',
                                      'value': 0,
                                      'format': format1})
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '>',
                                      'value': 0,
                                      'format': format2})
        
for i, row in df.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
        ws.write(REF_1, item)
        j += 1

ws = wb.add_worksheet('Difference in Fees Summary')

for j, column in enumerate(df_pivot.columns.values):
    col = get_column_letter(j + 1)
    row = 1
    rowend = len(df)+1
    ws.write(f'{col}{row}',column,cell_format_column)
    ws.set_column(f'{col}:{col}', 15)
    if column == 'Difference':
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '<',
                                      'value': 0,
                                      'format': format1})
        ws.conditional_format(f'{col}{row + 1}:{col}{rowend}', {'type': 'cell',
                                      'criteria': '>',
                                      'value': 0,
                                      'format': format2})
        
for i, row in df_pivot.iterrows():
    j = 0
    for item in row:
        REF_1 = ('{col}{row}').format(col = get_column_letter(j + 1), row = i + 2)
        ws.write(REF_1, item)
        j += 1
wb.close()

# df_MEAN = df.groupby(df['PAYNO']).mean()
# df_MEAN = df_MEAN['CONTRACTING RATE'].reset_index().rename(columns={'CONTRACTING RATE':'Average CR','index':'PAYNO'})
# df_MEAN['PAYNO'] = df_MEAN['PAYNO'].astype(int).round()

# df1 = pd.merge(df,df_MEAN)
# df1 = df1[df1['Type'] == 'PAYE'].groupby(['PAYNO'], as_index=False).agg({'COMPNAME':'first','Average CR':np.mean})

# df1['< £14.50'] = 0
# df1['>= £14.50'] = 0

# df1.loc[df1['Average CR'] >= 14.5, '>= £14.50'] = 1
# df1.loc[df1['Average CR'] < 14.5, '< £14.50'] = 1

# df_pivot = pd.pivot_table(df1, values=['Average CR','>= £14.50','< £14.50'], index=['COMPNAME'],aggfunc={'Average CR': np.mean, '>= £14.50': np.sum, '< £14.50': np.sum}, fill_value=0, margins=True)

# df_pivot['Total'] = df_pivot['< £14.50'] + df_pivot['>= £14.50']
# df_pivot['% below £14.50'] = (df_pivot['< £14.50'] / df_pivot['Total'])*100

